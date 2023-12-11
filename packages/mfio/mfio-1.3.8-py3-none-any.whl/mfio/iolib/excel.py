# -*- coding: utf-8 -*-#
# -------------------------------------------------------------------------------
# Author:       yunhgu
# Date:         2023/8/3
# -------------------------------------------------------------------------------
from collections import defaultdict

from openpyxl import Workbook as opwb
from xlrd import open_workbook
from xlwt import Workbook

from .base import IO


class Excel(IO):
    @classmethod
    def read(cls, path) -> dict:
        """表格读取,xlrd==1.2.0支持xlsx,xls

        Args:
            path: 表格路径

        Returns:
            {"sheet1":[[x,x,x],[x,x],[x,x,xx]]}
        """
        data = open_workbook(path)
        sheet_names = data.sheet_names()
        sheet_data_dic = defaultdict(list)
        for name in sheet_names:
            table = data.sheet_by_name(name)
            for row in range(table.nrows):
                columns = table.row_values(row)  # 某一行数据
                sheet_data_dic[name].append(columns)
        return sheet_data_dic

    @classmethod
    def write(cls, path, content, sheet_name="Sheet1"):
        """表格写入

        Args:
            path: 表格保存路径
            content: [[a,b,c],[d,e,f]]
            sheet_name: 表格sheet名字. Defaults to "Sheet1".
        """
        wd = Workbook()
        sheet = wd.add_sheet(sheet_name)
        for row, contents in enumerate(content):
            for column, value in enumerate(contents):
                sheet.write(row, column, value)
        wd.save(path)

    @classmethod
    def writes(cls, path, content_dic: dict):
        """表格写入
        Args:
            path: 表格保存路径
            content: {"Sheet1":[[a,b,c],[d,e,f]],"Sheet2":[[a,b,c],[d,e,f]]}
        """
        wd = Workbook()
        for sheet_name, content in content_dic.items():
            sheet = wd.add_sheet(sheet_name)
            for row, contents in enumerate(content):
                for column, value in enumerate(contents):
                    sheet.write(row, column, value)
        wd.save(path)

    @classmethod
    def write_openpyxl(cls, path, content_dic: dict):
        """表格写入
        Args:
            path: 表格保存路径
            content: {"Sheet1":[[a,b,c],[d,e,f]],"Sheet2":[[a,b,c],[d,e,f]]}
        """
        wb = opwb()  # 新建工作簿
        # ws = wb.active # 获取工作表
        # ws.append(['姓名', '学号', '年龄']) # 追加一行数据
        # ws.append(['张三', "1101", 17]) # 追加一行数据
        # ws.append(['李四', "1102", 18]) # 追加一行数据
        index = 0
        for sheet_name, content in content_dic.items():
            # 创建并返回一个工作表对象，默认位置最后，0代表第一个
            sheet = wb.create_sheet(sheet_name, index=index)
            for lines in content:
                sheet.append(lines)
            index += 1
        wb.save(path)

    @staticmethod
    def name():
        """
        :return: string with name of geometry
        """
        return "excel"


if __name__ == '__main__':
    value_dic = {
        "sheet1": [[1, 2, 3], [4, 5, 67, 8, 9]],
        "表格": [[1, 2, 3], [4, 5, 67, 8, 9]],
    }
    Excel.write_openpyxl("a.xlsx", value_dic)
