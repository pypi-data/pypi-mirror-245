import time
import json

from clickhouse_driver import Client
from openpyxl import Workbook
import random
import hashlib
import os
import base64
import pickle
import datetime
import zlib

from shapely.geometry import Point, Polygon

def get_version():
    return "0.2.0"


# Класс для работы с таблицами - результатами выполнения запросов
class ResultSet(object):
    def __init__(self):
        self._redis = None
        self._cacheExpire = 21600  # 6 hours
        self.data = []
        self.columns = []
        self.rowCount = 0
        self.keyField = None
        self.hashKey = None

    def setParams(self, params):
        self._redis = None if 'redis' not in params else params['redis']
        self._cacheExpire = None if 'expiration' not in params else params['expiration']

    def isCached(self):
        if not self._redis:
            return False
        else:
            return self._redis.exists(self.hashKey)

    def putToCache(self):
        # if not self.isCached():
        self._redis.set(self.hashKey,
                        zlib.compress(
                            base64.b64encode(pickle.dumps(
                                [self.columns,
                                 self.data,
                                 self.keyField,
                                 self.rowCount,
                                 self.hashKey])
                            ), 9)
                        )
        self._redis.expire(self.hashKey, self._cacheExpire)

    def getFromCache(self):
        tempHash = self.hashKey
        if not self._redis or not self.isCached():
            return None
        else:
            storedResult = self._redis.get(self.hashKey)
            if not storedResult:
                return None
            else:
                decompressedString = zlib.decompress(storedResult)
                if not decompressedString:
                    return None
                else:
                    decodedResult = base64.b64decode(decompressedString)
                    if not decodedResult:
                        return None
                    else:
                        storedObj = pickle.loads(decodedResult)
                        if not storedObj:
                            return None
                        else:
                            self.columns, self.data, self.keyField, self.rowCount = storedObj[0:4]
                            self.hashKey = tempHash
                            return self

    def findColumn(self, name):
        if self.columns:
            for pos, column in enumerate(self.columns):
                if column == name:
                    return pos
            return -1
        else:
            return -1

    def printOut(self):
        print(*self.columns, sep=', ')
        print(*self.data, sep=', ')
        print(f'rows {self.rowCount}')

    def exportToMIF(self, mifname):
        if not self.data or not self.columns:
            return None
        else:
            with open(mifname + '.mid', 'w') as midFile:
                with open(mifname + '.mif', 'w') as mifFile:
                    columnsIndexes = [self.findColumn('longitude'),
                                      self.findColumn('latitude'),
                                      self.findColumn(self.keyField)]
                    if any(columnIdx == -1 for columnIdx in columnsIndexes):
                        return None
                    else:
                        # Check type of keyfield
                        i_type = 2
                        for row in self.data:
                            testValue = row[columnsIndexes[2]]
                            if testValue:
                                v_type = type(testValue)
                                if v_type is int:
                                    v_type = 2
                                elif v_type is float:
                                    v_type = 1
                                elif v_type is str:
                                    v_type = 0
                                else:
                                    v_type = 0
                                i_type = min(i_type, v_type)
                                if i_type == 0:
                                    break

                        p_type = "Char(128)"
                        if i_type == 2:
                            p_type = "Integer"
                        elif i_type == 1:
                            p_type = "Float"
                        elif i_type == 0:
                            p_type = "Char(128)"

                        # write standard MIF-header
                        mifFile.write('Version   450\n')
                        mifFile.write('Charset "WindowsCyrillic"\n')
                        mifFile.write('Delimiter ","\n')
                        mifFile.write('CoordSys Earth Projection 1, 104\n')
                        mifFile.write('Columns 3\n')
                        mifFile.write('  Lon Float\n')
                        mifFile.write('  Lat Float\n')
                        mifFile.write('  {} {}\n'.format('Parameter', p_type))
                        mifFile.write('Data\n')
                        mifFile.write('\n')

                        for row in self.data:
                            lon, lat = row[columnsIndexes[0]], row[columnsIndexes[1]]
                            parameter = row[columnsIndexes[2]]
                            mifFile.write('Point {:.6f} {:.6f}\n'.format(lon, lat))
                            mifFile.write('    Symbol (33,0,4,"Map Symbols",0,0)\n')
                            midFile.write(f'{lon:.6f},{lat:.6f},{parameter}\n')


class ResultSetStream(object):
    def __init__(self, logs, keyfield, query, redis, redisHashList):
        self.logFiles = logs  # Log-file list
        self.curLogFile = '' if len(logs) == 0 else logs[0]  # Current log-file (name)
        self.curLogFileIdx = -1 if len(logs) == 0 else 0  # Current log-file (index)
        self.commonSQL = query.definition  # Common SQL for whole dataset
        self.query = query  # clickhouse DB
        self.redis = redis  # Redis cache DB
        self.redis_hash = redisHashList  # set of hashes of resultsets
        self.keyField = keyfield  # key field of resultset
        self.curRow = -1  # Current row number of current loaded Resultset

    def getFirstRecord(self):
        for self.curLogFileIdx in range(0, len(self.logFiles)):
            self.curLogFile = self.logFiles[self.curLogFileIdx]
            rs = self.query._Query__resultset
            # добываем данные по текущему логу
            # ... либо напрямую из базы
            if self.curLogFileIdx not in self.redis_hash:
                self.query.define({'sql': self.commonSQL})

                # TODO: ошибка! resultSet - не член класса ResultSetStream
                rs = self.query.run([self.curLogFile])
                if rs:
                    self.redis_hash[self.curLogFileIdx] = rs.hashKey

            # ...либо из кеша redis
            else:
                rs.hashKey = self.redis_hash[self.curLogFileIdx]
                rs.getFromCache()

            # если данные в конечном счете удалось получить, возвращаем стартовую строку
            # и выставляем позицию текущей строки на 0.
            if rs and rs.rowCount != 0:
                self.curRow = 0
                return rs.data[self.curRow], True
            else:
                continue

        # данных нет ни в одном из логов
        return None, None

    def getHashList(self):
        for self.curLogFileIdx in range(0, len(self.logFiles)):
            self.curLogFile = self.logFiles[self.curLogFileIdx]

            # добываем данные по текущему логу
            # ... либо напрямую из базы
            if self.curLogFileIdx not in self.redis_hash:
                self.query.define({'sql': self.commonSQL})
                rs = self.query.run([self.curLogFile])
                if rs:
                    self.redis_hash[self.curLogFileIdx] = rs.hashKey

            # ...либо из кеша redis
            else:
                rs.hashKey = self.redis_hash[self.curLogFileIdx]
                rs.getFromCache()

            # если данные в конечном счете удалось получить, возвращаем стартовую строку
            # и выставляем позицию текущей строки на 0.
            if rs and rs.rowCount != 0:
                self.curRow = 0
                return rs.data[self.curRow], True
            else:
                continue

        # данных нет ни в одном из логов
        return None, None

    def getRecordWithPromote(self):
        # если индекс текущего файла находится в пределах списка
        if self.curLogFileIdx in range(0, len(self.logFiles)):
            if self.query._Query__resultset.rowCount > 0 and self.curRow < self.query._Query__resultset.rowCount:
                row = self.query._Query__resultset.data[self.curRow]
                self.curRow += 1
                return row, True
            else:
                self.curLogFileIdx += 1
                while self.curLogFileIdx < len(self.logFiles):
                    self.curLogFile = self.logFiles[self.curLogFileIdx]

                    if self.curLogFileIdx not in self.redis_hash:
                        self.query.define({'sql': self.commonSQL})
                        self.query.run([self.curLogFile])
                        self.redis_hash[self.curLogFileIdx] = self.query._Query__resultset.hashKey
                        self.query._Query__resultset.putToCache()
                    else:
                        self.query._Query__resultset.hashKey = self.redis_hash[self.curLogFileIdx]
                        self.query._Query__resultset.getFromCache()

                    if self.query._Query__resultset and self.query._Query__resultset.rowCount > 0:
                        self.curRow = 1  # передвинуть указатель сразу на следующую запись...
                        return self.query._Query__resultset.data[0], False  # а вернуть при этом самую первую (нулевую)
                    else:
                        self.curLogFileIdx += 1
                else:
                    return None, None
        else:
            return None, None

    def exportRawData(self, filename):
        rawFile = open(filename + '.txt', 'w')
        row, _ = self.getFirstRecord()
        columns = self.query._Query__resultset.columns
        rawFile.write(f'{columns}\n')
        while True:
            row, sameFileFlag = self.getRecordWithPromote()
            if not row:
                break
            else:
                rawFile.write(f'{row}\n')

    def exportExcelData(self, filename):
        workbook = Workbook()
        sheet = workbook.active
        row, _ = self.getFirstRecord()
        columns = self.query._Query__resultset.columns
        for columnIdx in range(0, len(columns)):
            sheet.cell(row=1, column=columnIdx + 1).value = columns[columnIdx]
        rowIdx = 2
        while True:
            row, _ = self.getRecordWithPromote()
            if not row:
                break
            else:
                for columnIdx in range(0, len(columns)):
                    sheet.cell(row=rowIdx, column=columnIdx + 1).value = row[columnIdx]
                rowIdx += 1
        workbook.save(filename)

    def exportToMIF(self, keyfield, mifname):
        row, _ = self.getFirstRecord()
        rs = self.query._Query__resultset
        columns = rs.columns

        # открываем файлы на запись
        with open(mifname + '.mid', 'w') as midFile:
            with open(mifname + '.mif', 'w') as mifFile:
                if not row or not columns:
                    # формируем пустой MIF-файл
                    mifFile.write('Version   450\n')
                    mifFile.write('Charset "WindowsCyrillic"\n')
                    mifFile.write('Delimiter ","\n')
                    mifFile.write('CoordSys Earth Projection 1, 104\n')
                    mifFile.write('Columns 3\n')
                    mifFile.write('  Lon Float\n')
                    mifFile.write('  Lat Float\n')
                    mifFile.write('  {} {}\n'.format('Parameter', "Char(128)"))
                    mifFile.write('Data\n')
                    mifFile.write('\n')
                    return
                else:
                    # отыскать ключевые поля (lon, lat, keyField) и поместить их индексы в список
                    # разыскиваем поля с долготой и широтой, а также с целевым параметром и записываем
                    # их позиции в массив columnsIndexes
                    columnsIndexes = [rs.findColumn('longitude'),
                                      rs.findColumn('latitude'),
                                      rs.findColumn(keyfield)]

                    # проверка, все ли колонки найдены
                    if all(c != -1 for c in columnsIndexes):
                        # вычисляем тип данных на ключевом поле
                        i_type = 2
                        while True:
                            row, _ = self.getRecordWithPromote()
                            if not row:
                                break
                            else:
                                testValue = row[columnsIndexes[2]]
                                if testValue:
                                    v_type = type(testValue)
                                    if v_type is int:
                                        v_type = 2
                                    elif v_type is float:
                                        v_type = 1
                                    elif v_type is str:
                                        v_type = 0
                                    else:
                                        v_type = 0
                                    i_type = min(i_type, v_type)
                                    if i_type == 0:
                                        break
                        # преобразуем полученный тип в стандартный mapinfo
                        p_type = "Char(128)"
                        if i_type == 2:
                            p_type = "Integer"
                        elif i_type == 1:
                            p_type = "Float"
                        elif i_type == 0:
                            p_type = "Char(128)"

                        # формируем заголовок MIF-файла
                        mifFile.write('Version   450\n')
                        mifFile.write('Charset "WindowsCyrillic"\n')
                        mifFile.write('Delimiter ","\n')
                        mifFile.write('CoordSys Earth Projection 1, 104\n')
                        mifFile.write('Columns 3\n')
                        mifFile.write('  Lon Float\n')
                        mifFile.write('  Lat Float\n')
                        mifFile.write('  {} {}\n'.format('Parameter', p_type))
                        mifFile.write('Data\n')
                        mifFile.write('\n')

                        # перебираем все данные и экспортируем их
                        row, _ = self.getFirstRecord()
                        while True:
                            row, _ = self.getRecordWithPromote()
                            if not row:
                                break
                            else:
                                lon, lat = row[columnsIndexes[0]], row[columnsIndexes[1]]
                                parameter = row[columnsIndexes[2]]
                                mifFile.write('Point {:.6f} {:.6f}\n'.format(lon, lat))
                                mifFile.write('    Symbol (33,0,4,"Map Symbols",0,0)\n')
                                midFile.write(f'{lon:.6f},{lat:.6f},{parameter}\n')


# Класс для работы с запросами ClickHouse
class Query(object):
    def __init__(self, h, p, u, pw, db, redis, shouldCaching, cacheExpire, logger):
        # соединяемся с базой данных
        try:
            # do not change this parameters directly!
            self.__сlient = Client(host=h, port=p, user=u, password=pw, secure=False, verify=True,
                                   database=db, compression=True)

            self.__redis = redis
            self.__logger = logger
            self.__resultset = ResultSet()
            self.__directUsage = 0
            self.__cacheUsage = 0
            self.__shouldCaching = shouldCaching

            # this parameters can be responsible changed during runtime
            self._cacheExpire = cacheExpire
            self._projectId = -1
            self._udr = None

            # volatile value. you must set it before every call
            self.definition = None

        except:
            print('There is no connection with database')
            exit(-1)

    def __del__(self):
        self.__сlient.disconnect()

    def incDirectUsage(self):
        self.__directUsage += 1

    def incCacheUsage(self):
        self.__cacheUsage += 1

    def getUsageStat(self):
        return self.__directUsage, self.__cacheUsage

    # найти текст запроса в базе. Результат сохраняется поиска в объекте Query
    def getCemonaKPI(self, name):
        self.definition = f"SELECT sql FROM dictionary._queries where name = '{name}'"
        self.sqlExecute(False)
        if self.__resultset.data:
            self.definition = self.__resultset.data[0][0].strip(";")  # обрезать ";", если он в конце закрался
            return True
        else:
            self.definition = ''
            return False

    def define(self, params):
        self.definition = self.definition if 'sql' not in params else params['sql']
        self._udr = self._udr if 'udr' not in params else params['udr']
        self._projectId = self._projectId if 'projectId' not in params else params['projectId']

    def sqlCommand(self):
        self.__сlient.execute(self.definition)

    def sqlExecute(self, with_column_types=False):
        if not with_column_types:
            self.__resultset.data = self.__сlient.execute(self.definition)
            self.__resultset.columns = None
            self.__resultset.rowCount = len(self.__resultset.data)
        else:
            raw_data = self.__сlient.execute(self.definition, with_column_types=True)
            self.__resultset.data = raw_data[0]
            self.__resultset.columns = [tup[0] for tup in raw_data[1]]
            self.__resultset.rowCount = len(raw_data[0])

            if self.__shouldCaching:
                self.__resultset.keyField = 'unknown'
                self.__resultset.putToCache()

    def getHash(self, logFilesString=None):
        if self._Query__сlient:
            if logFilesString:
                logFilesString = str(logFilesString).replace("[", "").replace("]", "").replace("(", "").replace(")", "")
                sqlForOneFile = self.definition.replace("'dummy_filelist'", logFilesString)
                return hashlib.sha1(f'{sqlForOneFile}_{self._projectId}'.encode()).hexdigest()
            else:
                return None
        else:
            return None

    # query with log-file list and udr
    def run(self, logFilesString=None):
        if self.__сlient:
            # save basic query SQL
            _savedSQL = self.definition
            if logFilesString:
                logFilesString = str(logFilesString).replace("[", "").replace("]", "").replace("(", "").replace(")", "")
                sqlForOneFile = self.definition.replace("'dummy_filelist'", logFilesString)
                self.__resultset.hashKey = self.getHash(logFilesString)
                self.__resultset._redis = self.__redis
                if self.__resultset.isCached():
                    self.incCacheUsage()
                    self.__resultset.getFromCache()

                else:
                    self.incDirectUsage()
                    if not self._udr:
                        self.definition = sqlForOneFile
                        self.sqlExecute(True)
                    else:
                        def uniqueid():
                            seed = random.getrandbits(32)
                            while True:
                                yield seed
                                seed += 1

                        unique_sequence = uniqueid()
                        unicID = next(unique_sequence)
                        id_name = f'tmp.N{str(unicID)}'

                        # 1. create temporary table
                        self.definition = f"create table {id_name}" \
                                          f"(hash UInt64, longitude Float64, latitude Float64) " \
                                          f"engine = Memory()"
                        self.sqlCommand()

                        # 2. extract coordinates from source table into temporary table

                        self.definition = f'insert into {id_name} (hash, longitude, latitude)' \
                                          'select halfMD5(filename, number_line) hash, ' \
                                          'toFloat64(longitude) longitude, toFloat64(latitude) latitude ' \
                                          f'from ({sqlForOneFile}) where longitude is not null and latitude is not null'
                        self.sqlCommand()

                        # 2,3
                        self.definition = \
                            f"select * from ({sqlForOneFile}) main " \
                            f"inner join (select hash from {id_name} a, " \
                            f"(select key[1][1] as poly from dictionary._polygons where name_en = '{self._udr}') b " \
                            "where pointInPolygon(tuple(longitude, latitude), poly) " \
                            "group by hash" \
                            ") z on  halfMD5(filename, number_line) = z.hash " \
                            "order by date_time"

                        self.sqlExecute(True)

                        # 4. drop temporary table
                        self.definition = f"drop table {id_name}"
                        self.sqlCommand()

            self.definition = _savedSQL
            return self.__resultset

    def getUDRBounds(self):
        if not self._udr:
            return None, None, None, None
        else:
            params = {}
            params[
                'sql'] = f"SELECT min_lon, min_lat, max_lon, max_lat FROM dictionary.udr where name_en ='{self._udr}'"
            self.define(params)
            try:
                self.sqlExecute()
            except Exception:
                print(Exception.mro())
                return None, None, None, None

            if len(self.__resultset.data) >= 1:
                min_lon, min_lat, max_lon, max_lat = self.__resultset.data[0][0:4]
                return min_lon, min_lat, max_lon, max_lat
            else:
                return None, None, None, None


# класс, который предназначен для обработки датасетов, в частности фильтрации,
# а также вычисления агрегированных значений заданных полей:
# суммы, среднего, количества, процента от общего и пр.
class Processor(object):
    def __init__(self):
        self.condition = ''

    @staticmethod
    def __channelListParser(inputText):
        final_list = set()
        if not inputText:
            return []
        commaSplitted = inputText.replace(' ', '').split(',')
        for lexem in commaSplitted:
            dashSplitted = lexem.split('-')
            if len(dashSplitted) == 2:
                value1, value2 = int(dashSplitted[0]), int(dashSplitted[1])
                if value1 <= value2:
                    for v in range(value1, value2):
                        final_list.add(v)
                    final_list.add(value2)
                else:
                    for v in range(value2, value1):
                        final_list.add(v)
                    final_list.add(value1)
            else:
                if dashSplitted[0] != '':
                    final_list.add(int(dashSplitted[0]))
        return sorted(list(final_list))

    def _dump(source_rss, target_rss, data, columns, keyfield, signature):
        target_rss.query._Query__resultset = ResultSet()
        target_rss.query._Query__resultset._redis = target_rss.redis
        target_rss.query._Query__resultset.data = data.copy()
        target_rss.query._Query__resultset.rowCount = len(data)
        target_rss.query._Query__resultset.columns = columns
        target_rss.query._Query__resultset.keyField = source_rss.keyField
        target_rss.commonSQL = source_rss.commonSQL

        hashIndex = source_rss.curLogFileIdx - 1
        if hashIndex not in source_rss.redis_hash.keys():
            currentLog = source_rss.logFiles[hashIndex]
            target_rss.query.definition = target_rss.commonSQL
            hash = target_rss.query.getHash(currentLog)
            source_rss.redis_hash[hashIndex] = hash
        derivedHash = f'{source_rss.redis_hash[hashIndex]}' \
                      f'_{signature}_' \
                      f'{source_rss.logFiles[hashIndex]}'
        target_rss.query._Query__resultset.hashKey = hashlib.sha1(derivedHash.encode()).hexdigest()
        target_rss.redis_hash[source_rss.curLogFileIdx - 1] = target_rss.query._Query__resultset.hashKey
        target_rss.query._Query__resultset.putToCache()

    def SetCondition(self, cond=''):
        self.condition = cond.strip()

    def isConditionTrue(self, columns, row):
        if self.condition == '':
            return True
        else:
            for col in range(len(columns)):
                field_name = columns[col]
                locals()[field_name] = row[col]
            try:
                if eval(self.condition):
                    return True
                else:
                    return False
            except:
                return False

    @staticmethod
    def Sum(rss, param):
        cond = "" if "cond" not in param else param["cond"]
        local_cpu = Processor()
        local_cpu.SetCondition(cond)
        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns
        keyFieldIndex = rss.query._Query__resultset.findColumn(rss.keyField)

        accum = 0
        while True:
            row, _ = rss.getRecordWithPromote()
            if not row:
                break
            else:
                if local_cpu.isConditionTrue(columns, row):
                    value = row[keyFieldIndex]
                    if value is not None:
                        accum += value
                else:
                    pass
        return accum

    @staticmethod
    def Average(rss, param):
        cond = "" if "cond" not in param else param["cond"]
        local_cpu = Processor()
        local_cpu.SetCondition(cond)
        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns
        keyFieldIndex = rss.query._Query__resultset.findColumn(rss.keyField)

        accum = 0
        count = 0
        while row:
            row, _ = rss.getRecordWithPromote()
            if not row:
                break
            else:
                if local_cpu.isConditionTrue(columns, row):
                    value = row[keyFieldIndex]
                    if value is not None:
                        accum += value
                        count += 1

        if count != 0:
            print(count)
            return float(accum) / float(count)
        else:
            return 0.0

    @staticmethod
    def Threshold(rss, param):
        cond = "" if "cond" not in param else param["cond"]
        local_cpu = Processor()
        local_cpu.SetCondition(cond)
        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns
        keyFieldIndex = rss.query._Query__resultset.findColumn(rss.keyField)

        count = 0
        totalCount = 0

        while True:
            row, _ = rss.getRecordWithPromote()
            if not row:
                break
            else:
                totalCount += 1
                if local_cpu.isConditionTrue(columns, row):
                    count += 1

        if totalCount != 0:
            return float(count) / float(totalCount)
        else:
            return 0.0

    @staticmethod
    def Count(rss, param):
        #  подсчет уникальных групп значений согласно ключам из Cond
        cond = "" if "cond" not in param else param["cond"]

        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns
        unic_fields_raw = cond.split(',')
        unic_fields = []
        for u in unic_fields_raw:
            unic_fields.append(u.strip())
        unic_set = set()

        while True:
            row, _ = rss.getRecordWithPromote()
            if not row:
                break
            else:
                value = []
                for col in range(len(columns)):
                    field_name = columns[col]
                    if field_name in unic_fields:
                        value.append(row[col])
                unic_set.add(tuple(value))
        count = len(unic_set)
        return count

    @staticmethod
    def RouteLength(rss, param=None):
        row, _ = rss.getFirstRecord()

        keyFieldIndex = rss.query._Query__resultset.findColumn(rss.keyField)
        DistanceAccumulator = 0.0
        LastDistanceCtr = row[keyFieldIndex]

        while True:
            row, _ = rss.getRecordWithPromote()
            if not row:
                break
            else:
                CurrentDistanceCtr = row[keyFieldIndex]
                if CurrentDistanceCtr is not None:
                    # дистанция последнего перемещения
                    LastHopDelta = abs(CurrentDistanceCtr - LastDistanceCtr)
                    if LastHopDelta < 100:  # если нет разрыва непрерывности трассы
                        # накапливаем расстояние в аккумуляторе
                        DistanceAccumulator += LastHopDelta

                    # обновляем счетчик дистанции по текущему значению
                    LastDistanceCtr = CurrentDistanceCtr

        return DistanceAccumulator / 1000.0

    @staticmethod
    def GetBestGSM(rss, param):
        getBest = True if 'getBest' not in param else param['getBest']
        shortChannelList = '' if 'ch_list' not in param else param['ch_list']
        channelList = Processor.__channelListParser(shortChannelList)
        bandTriggerAllowed = False if 'bandTriggerAllowed' not in param else param['bandTriggerAllowed']
        zero_list_allow_all = False if 'zero_list_allow_all' not in param else param['zero_list_allow_all']

        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns

        def pocketDump():
            if getBest:  # если нас просят считать Best,
                maxValue = -200
                maxValueChannel = -1
                # перебираем весь карман в поисках максимума

                for channel in sorted(list(pocket.keys())):
                    if (len(channelList) == 0 and zero_list_allow_all) or (channel in channelList):
                        rxlev = pocket[channel][1]
                        if rxlev > maxValue:
                            maxValue = rxlev
                            maxValueChannel = channel
                if len(pocket.keys()) != 0 and maxValueChannel != -1:
                    buf.append([pocket[maxValueChannel][0],
                                maxValueChannel,
                                pocket[maxValueChannel][1],
                                pocket[maxValueChannel][2],
                                pocket[maxValueChannel][3]])

            else:  # иначе просто построчно сбрасываем в буфер (с учетом списка допустимых каналов, если он непуст)
                for key in sorted(list(pocket.keys())):
                    if (len(channelList) == 0) or (key in channelList):
                        buf.append([keyDate, key, pocket[key][1], pocket[key][2], pocket[key][3]])

        # Алгоритм:
        # в процессе прохода по resultset`у надо отслеживать факт повторного появления уже ранее известного канала
        # при этом следует также отслеживать смену band`а, и реагировать на эти события следуюшим образом:
        # 1. фиксировать datetime последнего (до повтора канала или смены бэнда) события
        # 2. всем прошедшим событиям назначать этот последний datetime

        # вычисляем положение ключевых колонок
        cols = [rss.query._Query__resultset.findColumn('date_time'),
                rss.query._Query__resultset.findColumn('arfcn'),
                rss.query._Query__resultset.findColumn('rx_level'),
                rss.query._Query__resultset.findColumn('longitude'),
                rss.query._Query__resultset.findColumn('latitude')]

        clnt = rss.query._Query__сlient

        h = clnt.connection.hosts[0][0]
        p = clnt.connection.hosts[0][1]
        u = clnt.connection.user
        pw = clnt.connection.password
        db = clnt.connection.database
        rd = rss.query._Query__redis
        sc = rss.query._Query__shouldCaching
        ce = rss.query._cacheExpire
        lg = rss.query._Query__logger

        q1 = Query(h, p, u, pw, db, rd, sc, ce, lg)
        out_rss = ResultSetStream(rss.logFiles, rss.keyField, q1, rss.redis, rss.redis_hash)

        if not row:
            return out_rss
        else:
            buf = []
            pocket = {}
            lastBand = None
            keyDate = None

            if any(c == -1 for c in cols) or (len(channelList) == 0 and not zero_list_allow_all):
                for rss.curLogFileIdx in range(1, len(rss.logFiles) + 1):
                    # def _dump(source_rss, target_rss, data, columns, keyfield, signature):
                    columns_reduced = ['date_time', 'arfcn', 'rx_level', 'longitude', 'latitude']
                    Processor._dump(rss, out_rss, [], columns_reduced, rss.keyField, 'GetBestGSM')
                return out_rss
            else:
                while True:
                    row, sameFileFlag = rss.getRecordWithPromote()
                    if not row:
                        break
                    else:
                        if sameFileFlag:
                            # получаем ключевые данные
                            date, arfcn, rxlevel, lon, lat = \
                                row[cols[0]], row[cols[1]], row[cols[2]], row[cols[3]], row[cols[4]]

                            # определяем текущий band
                            currentBand = 1800 if (512 <= arfcn <= 886) else 900

                            # проверяем, не начали ли новый цикл (по наличию записи в pocket с текущим arfcn)
                            # или не сменился ли band (при разрешенном триггере по смене Band)
                            if (arfcn in pocket) or (bandTriggerAllowed and (currentBand != lastBand)):
                                # прежде чем сбросить "карман" в ноль, выводим его накопленное содержимое
                                # (при первом заходе ничего не будет выведено, что нормально)
                                pocketDump()
                                keyDate = date
                                lastBand = currentBand
                                pocket.clear()
                                pocket[arfcn] = (date, rxlevel, lon, lat)
                            else:
                                pocket[arfcn] = (date, rxlevel, lon, lat)

                        else:
                            # если перешли на новый файл, то надо сохранить результаты предыдущего
                            out_columns = [columns[c] for c in cols]
                            Processor._dump(rss, out_rss, buf, out_columns, rss.keyField, 'GetBestGSM')
                            buf.clear()
                            date, arfcn, rxlevel, lon, lat = \
                                row[cols[0]], row[cols[1]], row[cols[2]], row[cols[3]], row[cols[4]]

                            currentBand = 1800 if (512 <= arfcn <= 886) else 900

                            # проверяем, не начали ли новый цикл (по наличию записи в pocket с текущим arfcn)
                            # или не сменился ли band (при разрешенном триггере по смене Band)
                            if (arfcn in pocket) or (bandTriggerAllowed and (currentBand != lastBand)):
                                # прежде чем сбросить "карман" в ноль, выводим его накопленное содержимое
                                # (при первом заходе ничего не будет выведено, что нормально)
                                pocketDump()
                                keyDate = date
                                lastBand = currentBand
                                pocket.clear()
                                pocket[arfcn] = (date, rxlevel, lon, lat)
                            else:
                                pocket[arfcn] = (date, rxlevel, lon, lat)
                pocketDump()

                out_columns = [columns[c] for c in cols]
                Processor._dump(rss, out_rss, buf, out_columns, rss.keyField, 'GetBestGSM')

                return out_rss

    @staticmethod
    def GetBestUMTS(rss, param):
        getBest = True if 'getBest' not in param else param['getBest']
        shortChannelList = '' if 'ch_list' not in param else param['ch_list']
        channelList = Processor.__channelListParser(shortChannelList)
        bandTriggerAllowed = False if 'bandTriggerAllowed' not in param else param['bandTriggerAllowed']
        zero_list_allow_all = False if 'zero_list_allow_all' not in param else param['zero_list_allow_all']

        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns

        def pocketDump():
            if getBest:  # если нас просят считать Best,
                maxValue = -200
                maxValueIndex = (None, None)
                # перебираем весь карман в поисках максимума
                key_list = sorted(list(pocket.keys()))
                for key in key_list:
                    if (len(channelList) == 0 and zero_list_allow_all) or (key[0] in channelList):
                        rscp = pocket[key][1]
                        if rscp > maxValue:
                            maxValue = rscp
                            maxValueIndex = key

                if len(key_list) != 0 and maxValueIndex != (None, None):
                    buf.append([pocket[maxValueIndex][0], maxValueIndex[0], maxValueIndex[1],
                                pocket[maxValueIndex][1], pocket[maxValueIndex][2], pocket[maxValueIndex][3]])

            else:  # иначе просто построчно сбрасываем в буфер (с учетом списка допустимых каналов, если он непуст)
                for key in sorted(list(pocket.keys())):
                    if (len(channelList) == 0) or (key[0] in channelList):
                        buf.append([keyDate, key[0], key[1], pocket[key][1], pocket[key][2], pocket[key][3]])

        # Алгоритм:
        # в процессе прохода по resultset`у надо отслеживать факт повторного появления уже ранее известного канала
        # при этом следует также отслеживать смену band`а, и реагировать на эти события следуюшим образом:
        # 1. фиксировать datetime последнего (до повтора канала или смены бэнда) события
        # 2. всем прошедшим событиям назначать этот последний datetime

        # вычисляем положение ключевых колонок
        cols = [rss.query._Query__resultset.findColumn('date_time'),
                rss.query._Query__resultset.findColumn('ch'),
                rss.query._Query__resultset.findColumn('sc'),
                rss.query._Query__resultset.findColumn('rscp'),
                rss.query._Query__resultset.findColumn('longitude'),
                rss.query._Query__resultset.findColumn('latitude')]

        clnt = rss.query._Query__сlient

        h = clnt.connection.hosts[0][0]
        p = clnt.connection.hosts[0][1]
        u = clnt.connection.user
        pw = clnt.connection.password
        db = clnt.connection.database
        rd = rss.query._Query__redis
        sc = rss.query._Query__shouldCaching
        ce = rss.query._cacheExpire
        lg = rss.query._Query__logger

        q1 = Query(h, p, u, pw, db, rd, sc, ce, lg)
        out_rss = ResultSetStream(rss.logFiles, rss.keyField, q1, rss.redis, rss.redis_hash)

        if not row:
            return out_rss
        else:
            buf = []
            pocket = {}
            currentBand = None
            lastBand = None
            keyDate = None
            last_ch_sc = None

            if any(c == -1 for c in cols) or (len(channelList) == 0 and not zero_list_allow_all):
                for rss.curLogFileIdx in range(1, len(rss.logFiles) + 1):
                    columns_reduced = ['date_time', 'ch', 'sc', 'rscp', 'longitude', 'latitude']
                    Processor._dump(rss, out_rss, [], columns_reduced, rss.keyField, 'GetBestUMTS')
                return out_rss

            else:
                while True:
                    row, sameFileFlag = rss.getRecordWithPromote()
                    if not row:
                        break
                    else:
                        if sameFileFlag:
                            # получаем ключевые данные
                            date, ch, sc, rscp, lon, lat = \
                                row[cols[0]], row[cols[1]], row[cols[2]], row[cols[3]], row[cols[4]], row[cols[5]]

                            if not ch or not sc or ch not in channelList:
                                continue

                            umtsBands = {
                                range(10562, 10838): '2100',
                                range(2937, 3088): '900'
                            }
                            for band in umtsBands:
                                if ch in band:
                                    currentBand = umtsBands[band]
                                    break

                            # проверяем, не начали ли новый цикл (по наличию записи в pocket с текущим channel number (ch))
                            # или не сменился ли band
                            # при этом, если номер канала остался прежним от предудущей итерации, ничего не предпринимаем -
                            # это текущее измерение, а не заход на новый цикл
                            if ((ch, sc) in pocket and (ch, sc) != last_ch_sc) or (
                                    bandTriggerAllowed and (currentBand != lastBand)):
                                # прежде чем сбросить "карман" в ноль, выводим его накопленное содержимое
                                # (при первом заходе ничего не будет выведено, что нормально)
                                pocketDump()
                                keyDate = date
                                last_ch_sc = ch, sc
                                lastBand = currentBand
                                pocket.clear()
                                pocket[(ch, sc)] = (date, rscp, lon, lat)
                            else:
                                pocket[(ch, sc)] = (date, rscp, lon, lat)
                        else:
                            # если перешли на новый файл, то надо сохранить результаты предыдущего
                            out_columns = [columns[c] for c in cols]
                            Processor._dump(rss, out_rss, buf, out_columns, rss.keyField, 'GetBestUMTS')

                            buf.clear()
                            date, ch, sc, rscp, lon, lat = \
                                row[cols[0]], row[cols[1]], row[cols[2]], row[cols[3]], row[cols[4]], row[cols[5]]

                            if not ch or not sc or ch not in channelList:
                                continue

                            umtsBands = {
                                range(10562, 10838): '2100',
                                range(2937, 3088): '900'
                            }
                            for band in umtsBands:
                                if ch in band:
                                    currentBand = umtsBands[band]
                                    break

                            # проверяем, не начали ли новый цикл (по наличию записи в pocket с текущим channel number (ch))
                            # или не сменился ли band
                            # при этом, если номер канала остался прежним от предудущей итерации, ничего не предпринимаем -
                            # это текущее измерение, а не заход на новый цикл
                            if ((ch, sc) in pocket and (ch, sc) != last_ch_sc) or (
                                    bandTriggerAllowed and (currentBand != lastBand)):
                                # прежде чем сбросить "карман" в ноль, выводим его накопленное содержимое
                                # (при первом заходе ничего не будет выведено, что нормально)
                                pocketDump()
                                keyDate = date
                                last_ch_sc = ch, sc
                                lastBand = currentBand
                                pocket.clear()
                                pocket[(ch, sc)] = (date, rscp, lon, lat)
                            else:
                                pocket[(ch, sc)] = (date, rscp, lon, lat)

                pocketDump()
                out_columns = [columns[c] for c in cols]
                Processor._dump(rss, out_rss, buf, out_columns, rss.keyField, 'GetBestUMTS')

                return out_rss

    @staticmethod
    def GetBestLTE(rss, param):
        getBest = True if 'getBest' not in param else param['getBest']
        shortChannelList = '' if 'ch_list' not in param else param['ch_list']
        channelList = Processor.__channelListParser(shortChannelList)
        bandTriggerAllowed = False if 'bandTriggerAllowed' not in param else param['bandTriggerAllowed']
        dl_bw = None if 'dl_bw' not in param else int(param['dl_bw']) * 5
        zero_list_allow_all = False if 'zero_list_allow_all' not in param else param['zero_list_allow_all']

        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns

        def pocketDump():
            if getBest:  # если нас просят считать Best,
                maxValue = -200
                maxValueIndex = (None, None)

                # перебираем весь карман в поисках максимума
                for key in sorted(list(pocket.keys())):
                    if (len(channelList) == 0) or (key[0] in channelList):
                        rsrp = pocket[key][2]
                        if rsrp > maxValue:
                            maxValue = rsrp
                            maxValueIndex = key

                if len(pocket) != 0 and maxValue != (None, None):
                    buf.append([pocket[maxValueIndex][0],
                                maxValueIndex[0],
                                pocket[maxValueIndex][1],
                                maxValueIndex[1],
                                pocket[maxValueIndex][2],
                                pocket[maxValueIndex][3],
                                pocket[maxValueIndex][4]])

            else:
                # иначе просто построчно сбрасываем в буфер
                # (с учетом списка допустимых каналов, если он непуст)
                for key in sorted(list(pocket.keys())):
                    if (len(channelList) == 0) or (key[0] in channelList):
                        buf.append(
                            [keyDate,
                             key[0],
                             pocket[key][1],
                             key[1],
                             pocket[key][2],
                             pocket[key][3],
                             pocket[key][4]])

        # Алгоритм:
        # в процессе прохода по resultset`у надо отслеживать факт повторного появления уже ранее известного канала
        # при этом следует также отслеживать смену band`а, и реагировать на эти события следуюшим образом:
        # 1. фиксировать datetime последнего (до повтора канала или смены бэнда) события
        # 2. всем прошедшим событиям назначать этот последний datetime

        # вычисляем положение ключевых колонок
        cols = [rss.query._Query__resultset.findColumn('date_time'),
                rss.query._Query__resultset.findColumn('ch'),
                rss.query._Query__resultset.findColumn('dl_bw'),
                rss.query._Query__resultset.findColumn('pci'),
                rss.query._Query__resultset.findColumn('rsrp'),
                rss.query._Query__resultset.findColumn('longitude'),
                rss.query._Query__resultset.findColumn('latitude')]

        clnt = rss.query._Query__сlient

        h = clnt.connection.hosts[0][0]
        p = clnt.connection.hosts[0][1]
        u = clnt.connection.user
        pw = clnt.connection.password
        db = clnt.connection.database
        rd = rss.query._Query__redis
        sc = rss.query._Query__shouldCaching
        ce = rss.query._cacheExpire
        lg = rss.query._Query__logger

        q1 = Query(h, p, u, pw, db, rd, sc, ce, lg)
        out_rss = ResultSetStream(rss.logFiles, rss.keyField, q1, rss.redis, rss.redis_hash)

        if not row:
            return out_rss
        else:

            buf = []
            pocket = {}
            currentBand = None
            lastBand = None
            keyDate = None
            last_ch_pci = None

            if any(c == -1 for c in cols) or (len(channelList) == 0 and not zero_list_allow_all):
                for rss.curLogFileIdx in range(1, len(rss.logFiles) + 1):
                    columns_reduced = ['date_time', 'ch', 'dl_bw', 'pci', 'rsrp', 'longitude', 'latitude']
                    Processor._dump(rss, out_rss, [], columns_reduced, rss.keyField, 'GetBestLTE')
                return out_rss

            else:
                while True:
                    row, sameFileFlag = rss.getRecordWithPromote()
                    if not row:
                        break
                    else:
                        if sameFileFlag:
                            date, ch, bw, pci, rsrp, lon, lat = \
                                row[cols[0]], row[cols[1]], row[cols[2]], \
                                row[cols[3]], row[cols[4]], row[cols[5]], row[cols[6]]

                            if not ch or not pci:
                                continue
                            if ch not in channelList:
                                continue

                            lteBands = {
                                range(6150, 6450): '800',
                                range(3450, 3800): '900',
                                range(1200, 1950): '1800',
                                range(0, 600): '2100',
                                range(38650, 39650): '2300',
                                range(2750, 3450): '2600',
                                range(37750, 38250): '2600TDD'
                            }
                            for band in lteBands:
                                if ch in band:
                                    currentBand = lteBands[band]
                                    break

                            if ((ch, pci) in pocket and (ch, pci) != last_ch_pci) or \
                                    (bandTriggerAllowed and (currentBand != lastBand)):
                                pocketDump()
                                keyDate = date
                                last_ch_pci = ch, pci
                                lastBand = currentBand
                                pocket.clear()
                                pocket[(ch, pci)] = (date, bw, rsrp, lon, lat)
                            else:
                                pocket[(ch, pci)] = (date, bw, rsrp, lon, lat)

                        else:
                            # если перешли на новый файл, то надо сохранить результаты предыдущего
                            out_columns = [columns[c] for c in cols]
                            Processor._dump(rss, out_rss, buf, out_columns, rss.keyField, 'GetBestLTE')

                            buf.clear()

                            date, ch, bw, pci, rsrp, lon, lat = \
                                row[cols[0]], row[cols[1]], row[cols[2]], row[cols[3]], \
                                row[cols[4]], row[cols[5]], row[cols[6]]

                            if not ch or not pci:
                                continue
                            if ch not in channelList:
                                continue

                            lteBands = {
                                range(6150, 6450): '800',
                                range(3450, 3800): '900',
                                range(1200, 1950): '1800',
                                range(0, 600): '2100',
                                range(38650, 39650): '2300',
                                range(2750, 3450): '2600',
                                range(37750, 38250): '2600TDD'
                            }
                            for band in lteBands:
                                if ch in band:
                                    currentBand = lteBands[band]
                                    break

                            if ((ch, pci) in pocket and (ch, pci) != last_ch_pci) or \
                                    (bandTriggerAllowed and (currentBand != lastBand)):
                                pocketDump()
                                keyDate = date
                                last_ch_pci = ch, pci
                                lastBand = currentBand
                                pocket.clear()
                                pocket[(ch, pci)] = (date, bw, rsrp, lon, lat)
                            else:
                                pocket[(ch, pci)] = (date, bw, rsrp, lon, lat)

                # сбрасываем остатки
                pocketDump()

                out_columns = [columns[c] for c in cols]
                Processor._dump(rss, out_rss, buf, out_columns, rss.keyField, 'GetBestLTE')

                proc = Processor()
                if dl_bw:
                    out_rss = proc.SimpleFilter(out_rss, param={'cond': f'dl_bw == {dl_bw}'})

                return out_rss

    @staticmethod
    def SimpleFilter(rss, param):
        cond = '' if 'cond' not in param else param['cond']
        local_cpu = Processor()
        local_cpu.SetCondition(cond)
        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns
        keyFieldIndex = rss.query._Query__resultset.findColumn(rss.keyField)

        # out_rss = ResultSetStream(rss.logFiles, rss.keyField, rss.query, rss.redis, rss.redis_hash)

        clnt = rss.query._Query__сlient

        h = clnt.connection.hosts[0][0]
        p = clnt.connection.hosts[0][1]
        u = clnt.connection.user
        pw = clnt.connection.password
        db = clnt.connection.database
        rd = rss.query._Query__redis
        sc = rss.query._Query__shouldCaching
        ce = rss.query._cacheExpire
        lg = rss.query._Query__logger

        q1 = Query(h, p, u, pw, db, rd, sc, ce, lg)
        out_rss = ResultSetStream(rss.logFiles, rss.keyField, q1, rss.redis, rss.redis_hash)

        if not row:
            return out_rss
        else:
            out_rss.commonSQL = 'SQL not valid after simplefilter applying'
            buf = []
            # если целевая колонка обнаружена во входном resultsetе
            if keyFieldIndex == -1 or not row:
                # если ключевое поле не найдено, возвращаем пустой набор данных
                for out_rss.curLogFileIdx in range(0, len(rss.logFiles)):
                    Processor._dump(rss, out_rss, [], columns, rss.keyField, 'SimpleFilter')
                return out_rss
            else:
                while True:
                    row, sameFileFlag = rss.getRecordWithPromote()
                    if not row:
                        break
                    else:
                        if sameFileFlag:
                            if local_cpu.isConditionTrue(columns, row):
                                buf.append(row)
                        else:
                            Processor._dump(rss, out_rss, buf, columns, rss.keyField, 'SimpleFilter')
                            buf.clear()  # сбросить буфер
                            if local_cpu.isConditionTrue(columns, row):
                                buf.append(row)
                Processor._dump(rss, out_rss, buf, columns, rss.keyField, 'SimpleFilter')
                return out_rss

    @staticmethod
    def Max(rss, param):
        cond = '' if 'cond' not in param else param['cond']
        local_cpu = Processor()
        local_cpu.SetCondition(cond)
        outfield = '' if 'outfield' not in param else param['outfield']
        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns
        if rss.keyField != 'date_time':
            maxValue = -float('inf')
        else:
            maxValue = datetime.datetime(1, 1, 1, 0, 0)
        extraValue = None
        count = 0
        keyFieldIndex = rss.query._Query__resultset.findColumn(rss.keyField)
        outFieldIndex = -1 if outfield == '' else rss.query._Query__resultset.findColumn(outfield)
        while True:
            row, _ = rss.getRecordWithPromote()
            if not row:
                break
            else:
                if local_cpu.isConditionTrue(columns, row):
                    value = row[keyFieldIndex]
                    if value and isinstance(value, (int, float, datetime.datetime)):
                        if maxValue < value:
                            maxValue = value
                            if outfield != "":
                                extraValue = row[outFieldIndex]

        if outfield != "":
            return extraValue
        else:
            if rss.keyField == 'date_time':
                maxValue = datetime.datetime(year=maxValue.year,
                                             month=maxValue.month,
                                             day=maxValue.day,
                                             hour=maxValue.hour,
                                             minute=maxValue.minute,
                                             second=maxValue.second)

                return maxValue
            else:
                return maxValue

    @staticmethod
    def Min(rss, param):
        cond = '' if 'cond' not in param else param['cond']
        local_cpu = Processor()
        local_cpu.SetCondition(cond)
        outfield = '' if 'outfield' not in param else param['outfield']
        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns
        if rss.keyField != 'date_time':
            minValue = float('inf')
        else:
            minValue = datetime.datetime(9999, 1, 1, 0, 0)
        extraValue = None
        count = 0
        keyFieldIndex = rss.query._Query__resultset.findColumn(rss.keyField)
        outFieldIndex = -1 if outfield == '' else rss.query._Query__resultset.findColumn(outfield)
        while True:
            row, _ = rss.getRecordWithPromote()
            if not row:
                break
            else:
                if local_cpu.isConditionTrue(columns, row):
                    value = row[keyFieldIndex]
                    if value and isinstance(value, (int, float, datetime.datetime)):
                        if minValue > value:
                            minValue = value
                            if outfield != "":
                                extraValue = row[outFieldIndex]

        if outfield != "":
            return extraValue
        else:
            if rss.keyField == 'date_time':
                minValue = datetime.datetime(year=minValue.year,
                                             month=minValue.month,
                                             day=minValue.day,
                                             hour=minValue.hour,
                                             minute=minValue.minute,
                                             second=minValue.second)

                return minValue
            else:
                return minValue

    @staticmethod
    def StateMachine(rss,
                     param={'KeyField': 'app_rate_dl_kb',
                            'States': ['Normal', 'Fall', 'Stall'],
                            'EnabledStates': ['Normal', 'Fall'],
                            'Transitions': [
                                {
                                    'from': 'Normal',
                                    'to': 'Fall',
                                    'trigger': 'app_rate_dl_kb < 10000'
                                },
                                {
                                    'from': 'Fall',
                                    'to': 'Stall',
                                    'trigger': 'app_rate_dl_kb < 6000',
                                    'ttt': 5000
                                },
                                {
                                    'from': 'Fall',
                                    'to': 'Normal',
                                    'trigger': 'app_rate_dl_kb > 15000'
                                },
                                {
                                    'from': 'Stall',
                                    'to': 'Normal',
                                    'trigger': 'app_rate_dl_kb > 15000'
                                }
                            ],
                            'InitialState': 'Normal'
                            }
                     ):

        keyField = None if 'KeyField' not in param else param['KeyField']
        State = None if 'InitialState' not in param else param['InitialState']
        EnabledStates = None if 'EnabledStates' not in param else param['EnabledStates']
        Transitions = None if 'Transitions' not in param else param['Transitions']

        row, _ = rss.getFirstRecord()
        columns = rss.query._Query__resultset.columns
        keyColumn = rss.query._Query__resultset.findColumn(keyField)
        timeColumn = rss.query._Query__resultset.findColumn('date_time')

        clnt = rss.query._Query__сlient

        h = clnt.connection.hosts[0][0]
        p = clnt.connection.hosts[0][1]
        u = clnt.connection.user
        pw = clnt.connection.password
        db = clnt.connection.database
        rd = rss.query._Query__redis
        sc = rss.query._Query__shouldCaching
        ce = rss.query._cacheExpire
        lg = rss.query._Query__logger

        q1 = Query(h, p, u, pw, db, rd, sc, ce, lg)
        out_rss = ResultSetStream(rss.logFiles, rss.keyField, q1, rss.redis, rss.redis_hash)

        if not row:
            return out_rss
        else:
            # если не определены релевантные колонки, возвращаем исходный resultset
            if keyColumn == -1 or timeColumn == -1:
                return rss
            else:
                # готовим буфер для отбора данных
                buf = []

                for t in Transitions:
                    t['triggered'] = False
                    t['baseTime'] = row[timeColumn]

                while True:
                    row, sameFileFlag = rss.getRecordWithPromote()
                    if not row:
                        break
                    else:
                        currentTime = row[timeColumn]

                        for col in range(len(columns)):
                            field_name = columns[col]
                            locals()[field_name] = row[col]

                        value = row[keyColumn]

                        if State in EnabledStates:
                            buf.append(row)

                        for t in Transitions:
                            if t['from'] == State:
                                timeToTrigger = None if 'ttt' not in t else int(t['ttt'])
                                try:
                                    if not eval(t['trigger'].strip()):
                                        t['triggered'] = False
                                    else:
                                        if not timeToTrigger:
                                            State = t['to']
                                            baseTime = currentTime
                                            break
                                        else:
                                            if not t['triggered']:
                                                t['triggered'] = True
                                                t['baseTime'] = row[timeColumn]
                                                break
                                            else:
                                                currentTime = row[timeColumn]
                                                timeDifference = (currentTime - t['baseTime']).total_seconds()
                                                if timeDifference >= timeToTrigger / 1000:
                                                    State = t['to']
                                                    t['triggered'] = False
                                                    t['baseTime'] = currentTime
                                                    break
                                except:
                                    pass

                Processor._dump(rss, out_rss, buf, columns, rss.keyField, 'StateMachine')

                return out_rss

    @staticmethod
    def Correlate(rss, param):

        master = rss
        subs = [] if 'subs' not in param else param['subs']
        final_buffer = []
        if master.rowCount == 0:
            return rss
        else:
            master_datetime_index = master.FindColumn('date_time')
            master_size = len(master.data)
            master_row = list(master.data[0])

            subs_datetime_indexes = []  # номера datetime-колонок
            subs_size = []  # размеры sub-таблиц
            subs_iterators = []  # итераторы sub-таблиц
            sub_rows = []  # текущие строки данных из sub-таблиц
            prev_sub_datetimes = []  # стартовые datetime по sub-таблицам
            for s in range(len(subs)):
                subs_datetime_indexes.append(subs[s].findColumn('date_time'))
                subs_size.append(len(subs[s].data))
                subs_iterators.append(0)
                sub_rows.append(list(subs[s].data[0]))
                prev_sub_datetimes.append(subs[s].data[0][subs_datetime_indexes[s]])

            # цикл по мастер-таблице
            for next_master_row in range(master_size):
                current_master_datetime = master.data[next_master_row][master_datetime_index]
                current_sub_datetimes = []
                for s in range(len(subs)):
                    current_sub_datetimes.append(subs[s].data[subs_iterators[s]][subs_datetime_indexes[s]])

                final_buffer.append(list(master.data[next_master_row]))
                for s in range(len(subs)):
                    if current_master_datetime >= current_sub_datetimes[s]:
                        # найти ближайший предшествующий datetime к master datetime
                        while current_master_datetime >= current_sub_datetimes[s]:
                            prev_sub_datetimes[s] = current_sub_datetimes[s]
                            subs_iterators[s] += 1
                            if subs_iterators[s] < subs_size[s]:
                                current_sub_datetimes[s] = subs[s].data[subs_iterators[s]][subs_datetime_indexes[s]]
                            else:
                                break
                        subs_iterators[s] -= 1
                        # на выходе из цикла в prev_sub_datetimes будет ближайшее время
                        # а в subs_iterators[s] номер соответствующей ему строки

                        final_buffer[next_master_row].extend(subs[s].data[subs_iterators[s]][2::])

        return final_buffer


class AbstractProcessor:
    def __init__(self):
        self.columns = []
        self.nextnode = None
        self.cond = None
        self.usage = 0

    def Configure(self, config):
        self.nextnode = self.nextnode if "nextnode" not in config else config["nextnode"]
        self.columns = self.columns if "columns" not in config else config["columns"]
        if self.nextnode:
            config['columns'] = self.columns
            self.nextnode.Configure(config)

    def Execute(self):
        self.usage += 1

    def FindColumn(self, column_name):
        try:
            return self.columns.index(column_name)
        except:
            return -1

class DBClient(AbstractProcessor):
    def __init__(self):
        super().__init__()
        self._host = None
        self._port = None
        self._user = None
        self._pwd = None
        self._dbname = None
        self.query = None
        self.logfilelist = []
        self.kpi_name = ''

    def Configure(self, config):
        self._host = self._host if "host" not in config else config["host"]
        self._port = self._port if "port" not in config else config["port"]
        self._user = self._user if "user" not in config else config["user"]
        self._pwd = self._pwd if "pwd" not in config else config["pwd"]
        self._dbname = self._dbname if "dbname" not in config else config["dbname"]
        self.logfilelist = self.logfilelist if "logfilelist" not in config else config["logfilelist"]
        self.kpi_name = self.kpi_name if "kpi_name" not in config else config["kpi_name"]

        self.query = Query(self._host, self._port, self._user, self._pwd, self._dbname,
                           redis=None, shouldCaching=False, cacheExpire=-1, logger=None)

        self.query.getCemonaKPI(self.kpi_name)
        raw_data = self.query._Query__сlient.execute(self.query.definition, with_column_types=True)
        self.columns = [tup[0] for tup in raw_data[1]]
        config['columns'] = self.columns
        AbstractProcessor.Configure(self, config)

    def Execute(self):
        AbstractProcessor.Execute(self)
        tmp_sql = self.query.definition
        if self.nextnode:
            for log in self.logfilelist:
                self.query.definition = tmp_sql.replace('dummy_filelist', log)
                settings = {'max_block_size': 100000}
                rows_gen = self.query._Query__сlient.execute_iter(self.query.definition, settings = settings)
                for row in rows_gen:
                    self.nextnode.Execute(row)
            self.nextnode.Execute(None)

class UDRChecker(AbstractProcessor):
    def __init__(self):
        super().__init__()
        self._host = None
        self._port = None
        self._user = None
        self._pwd = None
        self._dbname = None
        self.query = None
        self.udr = None
        self.geom = None
        self.bounds = []
        self.longitudeIdx = -1
        self.latitudeIdx = -1
        self.bypass = False

    def Configure(self, config):
        self._host = self._host if "host" not in config else config["host"]
        self._port = self._port if "port" not in config else config["port"]
        self._user = self._user if "user" not in config else config["user"]
        self._pwd = self._pwd if "pwd" not in config else config["pwd"]
        self._dbname = self._dbname if "dbname" not in config else config["dbname"]

        if self._host and self._port and self._user and self._pwd and self._dbname:
            self.query = Query(self._host, self._port, self._user, self._pwd, self._dbname,
                               redis=None, shouldCaching=False, cacheExpire=-1, logger=None)

        self.udr = self.udr if "udr" not in config else config["udr"]
        if self.udr:
            sql = f"SELECT key, min_lon, min_lat, max_lon, max_lat FROM dictionary.udr where name_en = '{self.udr}'"
            raw_data = self.query._Query__сlient.execute(sql, with_column_types=False)
            self.geom = raw_data[0][0]
            self.bounds = [raw_data[0][1], raw_data[0][2], raw_data[0][3], raw_data[0][4]]
            self.points = self.geom[0][0]

            self.longitudeIdx = self.FindColumn('longitude')
            self.latitudeIdx = self.FindColumn('latitude')
            self.bypass = False
        else:
            self.bypass = True
        AbstractProcessor.Configure(self, config)

    def Execute(self, record):
        AbstractProcessor.Execute(self)
        if not self.bypass:
            if record:
                lon = record[self.longitudeIdx]
                lat = record[self.latitudeIdx]
                #сначала грубая проверка на попадание в MBR
                if (self.bounds[0] <= lon <= self.bounds[2]) and \
                    (self.bounds[1] <= lat <= self.bounds[3]):

                    # потом, когда большинство лишних точек отсеялось, проверяем на попадание в UDR
                    pointtocheck = Point(lon, lat)
                    poly = Polygon(self.points)
                    if pointtocheck.within(poly):
                        self.nextnode.Execute(record)
                else:
                    pass
            else:
                self.nextnode.Execute(None)
        else:
            self.nextnode.Execute(record)


class SourceFromFile(AbstractProcessor):
    def __init__(self, sf):
        super().__init__()
        self.filename = sf
        self.sourceFile = open(sf, 'r')

    def Configure(self, config):
        with self.sourceFile as file:
            first_line = file.readline()
            self.columns = json.loads(first_line)
        AbstractProcessor.Configure(self, config)

    def Execute(self):
        super().Execute()
        self.sourceFile = open(self.filename, 'r')
        with self.sourceFile as file:
            file.readline() #skip first record (header)
            for line in file:
                record = json.loads(line)
                if self.nextnode is not None:
                    self.nextnode.Execute(record)
            self.nextnode.Execute(None)

class Output(AbstractProcessor):
    def __init__(self):
        super().__init__()
        self.result = None

    def Configure(self, config):
        AbstractProcessor.Configure(self, config)

    def Execute(self, record):
        AbstractProcessor.Execute(self)
        if record:
            print(record)
            self.result = record

class TXTExport(AbstractProcessor):
    def __init__(self, outfile):
        super().__init__()
        self.outfile = open(outfile, 'w')
        self.isheaderdumped = False

    def Configure(self, config):
        if self.columns and not self.isheaderdumped and self.columns != [None]:
            self.outfile.write(json.dumps(self.columns))
            self.isheaderdumped = True
        AbstractProcessor.Configure(self, config)


    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            self.outfile.write('\n')
            self.outfile.write(json.dumps(record, default=str))

class MIFExport(AbstractProcessor):
    def __init__(self):
        super().__init__()
        self.rootname = ''
        self.midfilename = ''
        self.midfile = None

        self.miffilename = ''
        self.miffile = None

        self.tmpmiffilename = ''
        self.tmpmiffile = None

        #self.isheaderdumped = False
        self.keyfield = ''
        self.valid = True
        self.columnsIndexes = []
        self.i_type = 2
        self.workdir = ''

    def Configure(self, config):
        self.workdir = self.workdir if 'workdir' not in config else config['workdir']
        self.rootname = self.rootname if 'rootname' not in config else config['rootname']
        self.keyfield = self.keyfield if 'keyfield' not in config else config['keyfield']
        self.columns = self.columns if ('columns' not in config or len(config['columns'])==0) else config['columns']
        if self.rootname != '':
            self.tmpmiffilename = os.path.join(self.workdir, self.rootname+'_tmp.mif')
            self.midfilename = os.path.join(self.workdir, self.rootname+'.mid')

            self.columnsIndexes = [self.FindColumn('longitude'),
                              self.FindColumn('latitude'),
                              self.FindColumn(self.keyfield)]
            if any(columnIdx == -1 for columnIdx in self.columnsIndexes):
                self.valid = False
            else:
                self.valid = True
        if self.valid:
            self.tmpmiffile = open(self.tmpmiffilename, 'w')
            self.midfile = open(self.midfilename, 'w')
        AbstractProcessor.Configure(self, config)


    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)

            lon, lat = record[self.columnsIndexes[0]], record[self.columnsIndexes[1]]
            parameter = record[self.columnsIndexes[2]]
            self.tmpmiffile.write('Point {:.6f} {:.6f}\n'.format(lon, lat))
            self.tmpmiffile.write('    Symbol (33,17,1,"Map Symbols",0,0)\n')
            self.midfile.write(f'{lon:.6f},{lat:.6f},{parameter}\n')

            if self.i_type > 0:
                # Check type of keyfield
                if parameter:
                    v_type = type(parameter)
                    if v_type is int:
                        v_type = 2
                    elif v_type is float:
                        v_type = 1
                    elif v_type is str:
                        v_type = 0
                    else:
                        v_type = 0
                    self.i_type = min(self.i_type, v_type)
        else:
            type_dict = {
                2: "Integer",
                1: "Float",
                0: "Char(128)"
            }
            p_type = type_dict.get(self.i_type, "Char(128)")

            self.tmpmiffile.close()
            self.midfile.close()

            if self.rootname != '':
                self.miffilename = os.path.join(self.workdir, self.rootname + '.mif')

            self.miffile = open(self.miffilename, 'w')
            # write standard MIF-header
            self.miffile.write('Version   450\n')
            self.miffile.write('Charset "WindowsCyrillic"\n')
            self.miffile.write('Delimiter ","\n')
            self.miffile.write('CoordSys Earth Projection 1, 104\n')
            self.miffile.write('Columns 3\n')
            self.miffile.write('  Lon Float\n')
            self.miffile.write('  Lat Float\n')
            self.miffile.write('  {} {}\n'.format('Parameter', p_type))
            self.miffile.write('Data\n')
            self.miffile.write('\n')

            with open(self.tmpmiffilename, 'r') as mif_content:
                # Читаем все строки первого файла
                lines = mif_content.readlines()
                # Записываем каждую строку первого файла в конец второго файла
                for line in lines:
                    self.miffile.write(line)

class Filter(AbstractProcessor):
    def __init__(self):
        super().__init__()
        self.usageCounter = 0
        self.cond = ''

    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        AbstractProcessor.Configure(self, config)

    def isConditionMet(self, record):
        if len(self.columns) == 0:
            return False
        if self.cond == '' or not self.cond:
            return True
        else:
            for column in range(len(self.columns)):
                field_name = self.columns[column]
                locals()[field_name] = record[column]
            try:
                if eval(self.cond):
                    return True
                else:
                    return False
            except:
                return False

    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            if self.isConditionMet(record):
                self.nextnode.Execute(record)
        else:
            if self.nextnode:
                self.nextnode.Execute(None)


class Max(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0
        self.keyfield = None
        self.outfield = None
        self.keyfieldIdx = -1
        self.outfieldIdx = -1
        self.keyValue = -float("inf")
        self.extraValue = None

    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        self.columns = self.columns if "columns" not in config else config["columns"]
        self.keyfield = self.keyfield if 'keyfield' not in config else config['keyfield']
        self.outfield = self.outfield if 'outfield' not in config else config['outfield']
        self.keyfieldIdx = super().FindColumn(self.keyfield)
        self.outfieldIdx = super().FindColumn(self.outfield)
        self.keyValue = -float("inf") if self.keyfield != 'date_time' else datetime.datetime(1, 1, 1, 0, 0)
        self.extraValue = None
        config['columns'] = [f'Max({self.keyfield})']
        if self.nextnode:
            self.nextnode.Configure(config)


    #TODO: переписать метод Flush внутрь Execute
    def Flush(self):
        if self.outfieldIdx != -1:
            self.nextnode.Execute(self.extraValue)
        else:
            if self.keyfield == 'date_time':
                maxValue = datetime.datetime(year=self.keyValue.year,
                                             month=self.keyValue.month,
                                             day=self.keyValue.day,
                                             hour=self.keyValue.hour,
                                             minute=self.keyValue.minute,
                                             second=self.keyValue.second)

                self.nextnode.Execute([maxValue])
            else:
                self.nextnode.Execute([self.keyValue])

    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            if self.isConditionMet(record):
                if self.keyfieldIdx != -1:
                    try:
                        value = record[self.keyfieldIdx]
                        if value and isinstance(value, (int, float, datetime.datetime)):
                            if self.keyValue < value:
                                self.keyValue = value
                                if self.outfieldIdx != -1:
                                    self.extraValue = record[self.outfieldIdx]
                    except:
                        pass
                else:
                    pass
        else:
            self.Flush()


class Min(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0
        self.keyfield = None
        self.outfield = None
        self.keyfieldIdx = -1
        self.outfieldIdx = -1
        self.keyValue = float("inf")
        self.extraValue = None

    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        self.columns = self.columns if "columns" not in config else config["columns"]
        self.keyfield = self.keyfield if 'keyfield' not in config else config['keyfield']
        self.outfield = self.outfield if 'outfield' not in config else config['outfield']
        self.keyfieldIdx = super().FindColumn(self.keyfield)
        self.outfieldIdx = super().FindColumn(self.outfield)
        self.keyValue = float("inf") if self.keyfield != 'date_time' else datetime.datetime(9999, 1, 1, 0, 0)
        self.extraValue = None
        config['columns'] = [f'Min({self.keyfield})']
        if self.nextnode:
            self.nextnode.Configure(config)

    #TODO: переписать метод Flush внутрь Execute
    #TODO: переписать обработку входных данных с учетом datetime (src-узлы)
    def Flush(self):
        if self.outfieldIdx != -1:
            self.nextnode.Execute(self.extraValue)
        else:
            if self.keyfield == 'date_time':
                minValue = datetime.datetime(year=self.keyValue.year,
                                             month=self.keyValue.month,
                                             day=self.keyValue.day,
                                             hour=self.keyValue.hour,
                                             minute=self.keyValue.minute,
                                             second=self.keyValue.second)

                self.nextnode.Execute([minValue])
            else:
                self.nextnode.Execute([self.keyValue])

    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            if self.isConditionMet(record):
                if self.keyfieldIdx != -1:
                    try:
                        value = record[self.keyfieldIdx]
                        if value and isinstance(value, (int, float, datetime.datetime)):
                            if self.keyValue > value:
                                self.keyValue = value
                                if self.outfieldIdx != -1:
                                    self.extraValue = record[self.outfieldIdx]
                    except:
                        pass
                else:
                    pass
        else:
            self.Flush()


class Avg(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0
        self.keyfield = None
        self.keyfieldIdx = -1
        self.accumValue = 0
        self.count = 0

    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        self.columns = self.columns if "columns" not in config else config["columns"]
        self.keyfield = self.keyfield if 'keyfield' not in config else config['keyfield']
        self.keyfieldIdx = super().FindColumn(self.keyfield)
        config['columns'] = [f'Avg({self.keyfield})']
        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            if self.isConditionMet(record):
                if self.keyfieldIdx != -1:
                    try:
                        value = record[self.keyfieldIdx]
                        if value and isinstance(value, (int, float)):
                            self.accumValue += value
                            self.count += 1
                    except:
                        pass
            else:
                pass
        else:
            try:
                record = [self.accumValue / self.count]
                self.nextnode.Execute(record)
            except:
                record = ["0.0"]
                self.nextnode.Execute(record)

class Sum(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0
        self.keyfield = None
        self.keyfieldIdx = -1
        self.accumValue = 0

    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        self.columns = self.columns if "columns" not in config else config["columns"]
        self.keyfield = self.keyfield if 'keyfield' not in config else config['keyfield']
        self.keyfieldIdx = super().FindColumn(self.keyfield)
        config['columns'] = [f'Sum({self.keyfield})']
        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            if self.isConditionMet(record):
                if self.keyfieldIdx != -1:
                    try:
                        value = record[self.keyfieldIdx]
                        if value and isinstance(value, (int, float)):
                            self.accumValue += value
                    except:
                        pass
            else:
                pass
        else:
            try:
                record = [self.accumValue]
                self.nextnode.Execute(record)
            except:
                record = ["0.0"]
                self.nextnode.Execute(record)

class Count(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0
        self.keyfield = None
        self.keyfieldIdx = -1
        self.accumValue = 0
        self.unic_fields = []
        self.unic_set = set()

    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        if self.cond:
            unic_fields_raw = self.cond.split(',')
            for u in unic_fields_raw:
                self.unic_fields.append(u.strip())

        self.columns = self.columns if "columns" not in config else config["columns"]
        self.keyfield = self.keyfield if 'keyfield' not in config else config['keyfield']
        self.keyfieldIdx = super().FindColumn(self.keyfield)
        config['columns'] = [f'Sum({self.keyfield})']
        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            value = []
            for col in range(len(self.columns)):
                field_name = self.columns[col]
                if field_name in self.unic_fields:
                    value.append(record[col])
            self.unic_set.add(tuple(value))
        else:
            count = len(self.unic_set)
            self.nextnode.Execute([str(count)])

class RouteLength(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0
        self.keyfield = None
        self.keyfieldIdx = -1
        self.distance_accumulator = 0.0
        self.last_dist_counter = 0.0
        self.last_hop_delta = 0.0

    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        self.columns = self.columns if "columns" not in config else config["columns"]
        self.keyfield = self.keyfield if 'keyfield' not in config else config['keyfield']
        self.keyfieldIdx = super().FindColumn(self.keyfield)
        config['columns'] = [f'Sum({self.keyfield})']
        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            CurrentDistanceCtr = record[self.keyfieldIdx]
            if CurrentDistanceCtr is not None:
                # дистанция последнего перемещения
                self.last_hop_delta = abs(CurrentDistanceCtr - self.last_dist_counter)
                if self.last_hop_delta  < 100:  # если нет разрыва непрерывности трассы
                    # накапливаем расстояние в аккумуляторе
                    self.distance_accumulator += self.last_hop_delta

                # обновляем счетчик дистанции по текущему значению
                self.last_dist_counter = CurrentDistanceCtr
        else:
            dist_in_km = self.distance_accumulator / 1000.0
            self.nextnode.Execute([str(dist_in_km)])


class Threshold(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0
        self.keyfield = None
        self.keyfieldIdx = -1
        self.thrcount = 0
        self.total = 0

    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        self.columns = self.columns if "columns" not in config else config["columns"]
        self.keyfield = self.keyfield if 'keyfield' not in config else config['keyfield']
        self.keyfieldIdx = super().FindColumn(self.keyfield)
        config['columns'] = [f'Thr({self.keyfield})']
        if self.nextnode:
            self.nextnode.Configure(config)


    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            self.total += 1
            if self.isConditionMet(record):
                self.thrcount += 1
        else:
            try:
                record = [self.thrcount / self.total]
                print(self.thrcount)
                self.nextnode.Execute(record)
            except:
                record = ["0.0"]
                self.nextnode.Execute(record)


def channelListParser(inputText):
    final_list = set()
    if not inputText:
        return []
    commaSplitted = inputText.replace(' ', '').split(',')
    for lexem in commaSplitted:
        dashSplitted = lexem.split('-')
        if len(dashSplitted) == 2:
            value1, value2 = int(dashSplitted[0]), int(dashSplitted[1])
            if value1 <= value2:
                for v in range(value1, value2):
                    final_list.add(v)
                final_list.add(value2)
            else:
                for v in range(value2, value1):
                    final_list.add(v)
                final_list.add(value1)
        else:
            if dashSplitted[0] != '':
                final_list.add(int(dashSplitted[0]))
    return sorted(list(final_list))


'''class BestGSM(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0
        self.maxValue = -float("inf")
        self.maxValueChannel = -1
        self.pocket = {}
        self.lastBand = None
        self.keyDate = None
        self.isProcessorValid = True
        self.filename = None
        self.date = None
        self.arfcn = None
        self.rxlevel = None
        self.lon = None
        self.lat = None
        self.lastUsedLog = None
        self.channelList= []

    def pocketDump(self):
        if self.getBest:  # если нас просят считать Best,
            self.maxValue = -float("inf")
            self.maxValueChannel = -1
            # перебираем весь карман в поисках максимума
            for channel in sorted(list(self.pocket.keys())):
                if (len(self.channelList) == 0 and self.zero_list_allow_all) or (channel in self.channelList):
                    rxlev = self.pocket[channel][2]
                    if rxlev > self.maxValue:
                        self.maxValue = rxlev
                        self.maxValueChannel = channel
            if len(self.pocket.keys()) != 0 and self.maxValueChannel != -1:
                return  [self.pocket[self.maxValueChannel][0],
                         self.pocket[self.maxValueChannel][1],
                         self.maxValueChannel,
                         self.pocket[self.maxValueChannel][2],
                         self.pocket[self.maxValueChannel][3],
                         self.pocket[self.maxValueChannel][4]]

        #TODO: проверить эту ветку
        else:  # иначе просто построчно сбрасываем в буфер (с учетом списка допустимых каналов, если он непуст)
            for key in sorted(list(self.pocket.keys())):
                if (len(self.channelList) == 0) or (key in self.channelList):
                    return [self.pocket[key][0],
                            self.keyDate,
                            key,
                            self.pocket[key][2],
                            self.pocket[key][3],
                            self.pocket[key][4]]

    def Configure(self, config):
        self.isProcessorValid = True
        self.cond = self.cond if "cond" not in config else config["cond"]
        self.getBest = True if 'getBest' not in config else config['getBest']
        shortChannelList = '' if 'ch_list' not in config else config['ch_list']
        self.channelList = self.channelList if 'ch_list' not in config else channelListParser(shortChannelList)
        self.bandTriggerAllowed = False if 'bandTriggerAllowed' not in config else config['bandTriggerAllowed']
        self.zero_list_allow_all = False if 'zero_list_allow_all' not in config else config['zero_list_allow_all']
        AbstractProcessor.Configure(self, config)
        # вычисляем положение ключевых колонок
        self.cols = [self.FindColumn('filename'),
                     self.FindColumn('date_time'),
                     self.FindColumn('arfcn'),
                     self.FindColumn('rx_level'),
                     self.FindColumn('longitude'),
                     self.FindColumn('latitude')]

        self.columns_reduced = ['filename', 'date_time', 'arfcn', 'rx_level', 'longitude', 'latitude']
        if any(c == -1 for c in self.cols) or (len(self.channelList) == 0 and not self.zero_list_allow_all):
            self.isProcessorValid = False
        config['columns'] = self.columns_reduced
        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if not record:
            self.maxValue = -float("inf")
            self.maxValueChannel = -1
            best_result = self.pocketDump()
            if self.nextnode:
                self.nextnode.Execute(best_result)
                self.nextnode.Execute(None)
        else:
            AbstractProcessor.Execute(self)
            # получаем ключевые данные
            self.filename, self.date, self.arfcn, self.rxlevel, self.lon, self.lat = \
                record[self.cols[0]], record[self.cols[1]], record[self.cols[2]], \
                record[self.cols[3]], record[self.cols[4]], record[self.cols[5]]

            if not self.lastUsedLog:                # выставляем на обработку первый лог
                self.lastUsedLog = self.filename

            # определяем текущий band
            currentBand = 1800 if (512 <= self.arfcn <= 886) else 900

            # проверяем, не начали ли новый цикл (по наличию записи в pocket с текущим arfcn)
            # или не сменился ли band (при разрешенном триггере по смене Band) или лог
            if (self.arfcn in self.pocket) or (self.bandTriggerAllowed and (
                    currentBand != self.lastBand) or (self.filename != self.lastUsedLog)):
                self.lastUsedLog = self.filename
                # прежде чем сбросить "карман" в ноль, выводим его накопленное содержимое
                # (при первом заходе ничего не будет выведено, что нормально)
                best_result = self.pocketDump()
                if self.nextnode:
                    self.nextnode.Execute(best_result)
                self.keyDate = self.date
                self.lastBand = currentBand
                self.pocket.clear()
                self.pocket[self.arfcn] = (self.filename, self.date, self.rxlevel, self.lon, self.lat)
            else:
                self.pocket[self.arfcn] = (self.filename, self.date, self.rxlevel, self.lon, self.lat)
'''

class BestGSM(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0

        self.filename = None
        self.date = None
        self.arfcn = None
        self.rx_level = None
        self.lon = None
        self.lat = None

        self.pocket = {}
        self.channel_pocket = set()
        self.allowed_channels = []
        self.last_log = None
        self.last_date = None
        self.last_arfcn = -1
        self.pocket_value_max_age = 2

    def pocketAging(self):
        for key in sorted(list(self.pocket.keys())):
            age = self.pocket[key][4]
            age -= 1
            if age == 0:
                self.pocket.pop(key, None)
            else:
                self.pocket[key] = self.pocket[key][0], self.pocket[key][1], \
                                   self.pocket[key][2], self.pocket[key][3], age

    def pocketDump(self):
        maxValue = -float("inf")
        maxValueIdx = None
        # перебираем весь карман в поисках максимума
        for key in sorted(list(self.pocket.keys())):
            rx_level = self.pocket[key][3]
            if rx_level > maxValue:
                maxValue = rx_level
                maxValueIdx = key

        if len(self.pocket.keys()) != 0 and maxValueIdx != None:
            arfcn = maxValueIdx
            value_to_send = [
                    self.pocket[maxValueIdx][0],
                    self.last_date, #берем последнее известное время в качестве времени максимума
                    arfcn,
                    self.pocket[maxValueIdx][2],
                    self.lon, # обновляем координаты
                    self.lat
            ]
            return value_to_send


    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        shortChannelList = '' if 'ch_list' not in config else config['ch_list']
        self.allowed_channels = self.allowed_channels if 'ch_list' not in config else channelListParser(shortChannelList)
        self.pocket_value_max_age = self.pocket_value_max_age if 'max_age' not in config else config['max_age']

        AbstractProcessor.Configure(self, config)
        # вычисляем положение ключевых колонок
        self.cols = [self.FindColumn('filename'),
                     self.FindColumn('date_time'),
                     self.FindColumn('arfcn'),
                     self.FindColumn('rx_level'),
                     self.FindColumn('longitude'),
                     self.FindColumn('latitude')]

        self.columns_reduced = ['filename', 'date_time', 'arfcn', 'rx_level', 'longitude', 'latitude']
        config['columns'] = self.columns_reduced
        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if not record:
            best_result = self.pocketDump()
            if self.nextnode:
                self.nextnode.Execute(best_result)
                self.nextnode.Execute(None)
        else:
            AbstractProcessor.Execute(self)
            # получаем ключевые данные
            self.filename, self.date, self.arfcn = record[self.cols[0]], record[self.cols[1]], record[self.cols[2]]
            self.rx_level, self.lon, self.lat = record[self.cols[3]], record[self.cols[4]], record[self.cols[5]]

            if self.arfcn == None or self.rx_level == None:  # проверка полноты данных
                pass
            else:
                if not self.last_log:                                   # выставляем на обработку первый лог
                    self.last_log = self.filename

                if self.filename != self.last_log:                      # смена лога?
                    self.last_log = self.filename
                    best_result = self.pocketDump()                     # извлекаем текущий максимум
                    if self.nextnode and best_result:
                        self.nextnode.Execute(best_result)              # передаем на следующий этап
                    self.pocket.clear()                                 # чистим контекст
                    self.channel_pocket.clear()
                    if self.arfcn in self.allowed_channels:                # заполняем текущей строкой контекст
                        self.pocket[self.arfcn] = (self.filename, self.date, self.rx_level,
                                                self.lon, self.lat, self.pocket_value_max_age)
                        self.channel_pocket.add(self.arfcn)
                        self.last_ch = self.arfcn
                        self.last_date = self.date
                    else:
                        self.last_date = None
                        self.last_arfcn = -1
                else:
                    conditions = [
                       (self.arfcn != self.last_arfcn and self.arfcn in self.channel_pocket),
                       #(self.ch == self.last_ch and self.date != self.last_date),
                       #(self.date != self.last_date),
                       # (self.arfcn != self.last_arfcn)
                        #(self.lon != self.last_lon or self.lat != self.last_lat)
                    ]
                    new_meas_cycle_detected = any(conditions)

                    if new_meas_cycle_detected:
                        self.pocketAging()
                        best_result = self.pocketDump()  # извлекаем текущий максимум
                        if self.nextnode and best_result:
                            self.nextnode.Execute(best_result)  # передаем на следующий этап
                        self.channel_pocket.clear()

                    if self.arfcn in self.allowed_channels:
                        self.pocket[self.arfcn] = (
                            self.filename, self.date, self.rx_level,
                            self.lon, self.lat, self.pocket_value_max_age)
                    self.channel_pocket.add(self.arfcn)

                    self.last_arfcn = self.arfcn  # готовимся к следующей смене канала
                    self.last_date = self.date  # готовимся к следующей смене времени

'''
class BestUMTS(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0

        self.filename = None
        self.date = None
        self.ch = None
        self.sc = None
        self.rscp = None
        self.lon = None
        self.lat = None

        self.pocket = {}
        self.allowed_channels = []
        self.last_log = None
        self.last_ch = -1

    def pocketDump(self):
        maxValue = -float("inf")
        maxValueIdx = (None, None)
        # перебираем весь карман в поисках максимума, уменьшая счетчик на каждом ключе и удаляя ключи по достижении 0
        for key in sorted(list(self.pocket.keys())):
            channel = key[0]

            fn, dt, rscp, lon, lat, cnt = self.pocket[key]
            cnt -= 1
            if cnt == 0:
                self.pocket.pop (key, None)
            else:
                self.pocket[key]=(fn, dt, rscp, lon, lat, cnt)
                if channel in self.allowed_channels:
                    rscp = self.pocket[key][2]
                    if rscp > maxValue:
                        maxValue = rscp
                        maxValueIdx = key

        if len(self.pocket.keys()) != 0 and maxValueIdx != (None, None):
            ch, sc = maxValueIdx
            return [self.pocket[maxValueIdx][0],
                    self.pocket[maxValueIdx][1],
                    ch,
                    sc,
                    self.pocket[maxValueIdx][2],
                    self.pocket[maxValueIdx][3],
                    self.pocket[maxValueIdx][4]]

    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        shortChannelList = '' if 'ch_list' not in config else config['ch_list']
        self.allowed_channels = self.allowed_channels if 'ch_list' not in config else channelListParser(shortChannelList)

        AbstractProcessor.Configure(self, config)
        # вычисляем положение ключевых колонок
        self.cols = [self.FindColumn('filename'),
                     self.FindColumn('date_time'),
                     self.FindColumn('ch'),
                     self.FindColumn('sc'),
                     self.FindColumn('rscp'),
                     self.FindColumn('longitude'),
                     self.FindColumn('latitude')]

        self.columns_reduced = ['filename', 'date_time', 'ch', 'sc', 'rscp', 'longitude', 'latitude']
        config['columns'] = self.columns_reduced
        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if not record:
            best_result = self.pocketDump()
            if self.nextnode:
                self.nextnode.Execute(best_result)
                self.nextnode.Execute(None)
        else:
            AbstractProcessor.Execute(self)
            # получаем ключевые данные
            self.filename, self.date, self.ch, self.sc, self.rscp, self.lon, self.lat = \
                record[self.cols[0]], record[self.cols[1]], record[self.cols[2]], record[self.cols[3]], \
                record[self.cols[4]], record[self.cols[5]], record[self.cols[6]]

            if not self.last_log:                # выставляем на обработку первый лог
                self.last_log = self.filename

            if self.ch and self.sc != None and self.rscp != None:
                if self.ch != self.last_ch:
                    if self.ch in self.allowed_channels:
                        self.last_ch = self.ch
                        #print(f"вывод из pocket при текущем ch={self.ch} и предыдущем {self.last_ch}")
                        best_result = self.pocketDump()
                        if self.nextnode and best_result:
                            self.nextnode.Execute(best_result)
                            #print('\t', best_result)
                if self.ch in self.allowed_channels:
                    self.pocket[(self.ch, self.sc)] = (
                    self.filename, self.date, self.rscp, self.lon, self.lat, 6)
'''
class BestUMTS(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0

        self.filename = None
        self.date = None
        self.ch = None
        self.sc = None
        self.rscp = None
        self.lon = None
        self.lat = None

        self.pocket = {}
        self.channel_pocket = set()
        self.allowed_channels = []
        self.last_log = None
        self.last_date = None
        self.last_ch = -1
        self.pocket_value_max_age = 1000

    def pocketAging(self):
        for key in sorted(list(self.pocket.keys())):
            age = self.pocket[key][5]
            age -= 1
            if age == 0:
                self.pocket.pop(key, None)
            else:
                self.pocket[key] = self.pocket[key][0], self.pocket[key][1], self.pocket[key][2],\
                                   self.pocket[key][3], self.pocket[key][4], age

    def pocketDump(self):
        maxValue = -float("inf")
        maxValueIdx = (None, None)
        # перебираем весь карман в поисках максимума
        for key in sorted(list(self.pocket.keys())):
            rscp = self.pocket[key][3]
            if rscp > maxValue:
                maxValue = rscp
                maxValueIdx = key

        if len(self.pocket.keys()) != 0 and maxValueIdx != (None, None):
            ch, sc = maxValueIdx
            value_to_send = [
                    self.pocket[maxValueIdx][0],
                    self.last_date, #берем последнее известное время в качестве времени максимума
                    ch,
                    sc,
                    self.pocket[maxValueIdx][2],
                    self.lon, # обновляем координаты
                    self.lat
            ]
            return value_to_send


    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        shortChannelList = '' if 'ch_list' not in config else config['ch_list']
        self.allowed_channels = self.allowed_channels if 'ch_list' not in config else channelListParser(shortChannelList)
        self.pocket_value_max_age = self.pocket_value_max_age if 'max_age' not in config else config['max_age']

        AbstractProcessor.Configure(self, config)
        # вычисляем положение ключевых колонок
        self.cols = [self.FindColumn('filename'),
                     self.FindColumn('date_time'),
                     self.FindColumn('ch'),
                     self.FindColumn('sc'),
                     self.FindColumn('rscp'),
                     self.FindColumn('longitude'),
                     self.FindColumn('latitude')]

        self.columns_reduced = ['filename', 'date_time', 'ch', 'sc', 'rscp', 'longitude', 'latitude']
        config['columns'] = self.columns_reduced
        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if not record:
            best_result = self.pocketDump()
            if self.nextnode:
                self.nextnode.Execute(best_result)
                self.nextnode.Execute(None)
        else:
            AbstractProcessor.Execute(self)
            # получаем ключевые данные
            self.filename, self.date, self.ch, self.sc, self.rscp, self.lon, self.lat = \
                record[self.cols[0]], record[self.cols[1]], record[self.cols[2]], record[self.cols[3]], \
                record[self.cols[4]], record[self.cols[5]], record[self.cols[6]]

            if self.ch == None or self.sc == None or self.rscp == None:  # проверка полноты данных
                pass
            else:
                if not self.last_log:                                   # выставляем на обработку первый лог
                    self.last_log = self.filename

                if self.filename != self.last_log:                      # смена лога?
                    self.last_log = self.filename
                    best_result = self.pocketDump()                     # извлекаем текущий максимум
                    if self.nextnode and best_result:
                        self.nextnode.Execute(best_result)              # передаем на следующий этап
                    self.pocket.clear()                                 # чистим контекст
                    self.channel_pocket.clear()
                    if self.ch in self.allowed_channels:                # заполняем текущей строкой контекст
                        self.pocket[(self.ch, self.sc)] = (self.filename, self.date, self.rscp,
                                                            self.lon, self.lat, self.pocket_value_max_age)
                        self.channel_pocket.add(self.ch)
                        self.last_ch = self.ch
                        self.last_date = self.date
                    else:
                        self.last_date = None
                        self.last_ch = -1
                else:
                    conditions = [
                       #(self.ch != self.last_ch and self.ch in self.channel_pocket),
                       #(self.ch == self.last_ch and self.date != self.last_date),
                        (self.date != self.last_date)
                    ]
                    new_meas_cycle_detected = any(conditions)

                    if new_meas_cycle_detected:
                        self.pocketAging()
                        best_result = self.pocketDump()  # извлекаем текущий максимум
                        if self.nextnode and best_result:
                            self.nextnode.Execute(best_result)  # передаем на следующий этап
                        self.channel_pocket.clear()

                    if self.ch in self.allowed_channels:
                        self.pocket[(self.ch, self.sc)] = (
                            self.filename, self.date, self.rscp,
                            self.lon, self.lat, self.pocket_value_max_age)
                    self.channel_pocket.add(self.ch)

                    self.last_ch = self.ch  # готовимся к следующей смене канала
                    self.last_date = self.date  # готовимся к следующей смене времени


class BestLTE(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0

        self.filename = None
        self.date = None
        self.dl_bw = None
        self.ch = None
        self.pci = None
        self.rsrp = None
        self.lon = None
        self.lat = None

        self.pocket = {}
        self.pocket_age = 20
        self.channel_pocket = set()
        self.allowed_channels = []
        self.last_log = None
        self.last_date = None
        self.last_ch = -1
        self.pocket_value_max_age = 1000

    def pocketAging(self):
        for key in sorted(list(self.pocket.keys())):
            age = self.pocket[key][6]
            age -= 1
            if age == 0:
                self.pocket.pop(key, None)
            else:
                self.pocket[key] = self.pocket[key][0], self.pocket[key][1], \
                                   self.pocket[key][2],self.pocket[key][3],\
                                   self.pocket[key][4], self.pocket[key][5], age

    def pocketDump(self):
        maxValue = -float("inf")
        maxValueIdx = (None, None)
        # перебираем весь карман в поисках максимума
        for key in sorted(list(self.pocket.keys())):
            rsrp = self.pocket[key][3]
            if rsrp > maxValue:
                maxValue = rsrp
                maxValueIdx = key

        if len(self.pocket.keys()) != 0 and maxValueIdx != (None, None):
            ch, pci = maxValueIdx
            value_to_send = [
                    self.pocket[maxValueIdx][0],
                    #self.pocket[maxValueIdx][1],
                    self.last_date, #берем последнее известное время в качестве времени максимума
                    ch,
                    pci,
                    self.pocket[maxValueIdx][2],
                    self.pocket[maxValueIdx][3],
                    #self.pocket[maxValueIdx][4],
                    #self.pocket[maxValueIdx][5]
                    self.lon, # обновляем координаты
                    self.lat]
            #if self.last_date != self.last_sent_date:
            #    self.last_sent_date = self.last_date
            return value_to_send


    def Configure(self, config):
        self.cond = self.cond if "cond" not in config else config["cond"]
        shortChannelList = '' if 'ch_list' not in config else config['ch_list']
        self.pocket_value_max_age = self.pocket_value_max_age if 'max_age' not in config else config['max_age']
        self.allowed_channels = self.allowed_channels if 'ch_list' not in config else channelListParser(shortChannelList)

        AbstractProcessor.Configure(self, config)
        # вычисляем положение ключевых колонок
        self.cols = [self.FindColumn('filename'),
                     self.FindColumn('date_time'),
                     self.FindColumn('ch'),
                     self.FindColumn('dl_bw'),
                     self.FindColumn('pci'),
                     self.FindColumn('rsrp'),
                     self.FindColumn('longitude'),
                     self.FindColumn('latitude')]

        self.columns_reduced = ['filename', 'date_time', 'ch', 'dl_bw', 'pci', 'rsrp', 'longitude', 'latitude']
        config['columns'] = self.columns_reduced
        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if not record:
            best_result = self.pocketDump()
            if self.nextnode:
                self.nextnode.Execute(best_result)
                self.nextnode.Execute(None)
        else:
            AbstractProcessor.Execute(self)
            # получаем ключевые данные
            self.filename, self.date, self.ch, self.dl_bw, self.pci, self.rsrp, self.lon, self.lat = \
                record[self.cols[0]], record[self.cols[1]], record[self.cols[2]], record[self.cols[3]], \
                record[self.cols[4]], record[self.cols[5]], record[self.cols[6]], record[self.cols[7]]


            if self.ch == None or self.pci == None or self.rsrp == None:  # проверка полноты данных
                pass
            else:
                if not self.last_log:                                   # выставляем на обработку первый лог
                    self.last_log = self.filename

                if self.filename != self.last_log:                      # смена лога?
                    self.last_log = self.filename
                    best_result = self.pocketDump()                     # извлекаем текущий максимум
                    if self.nextnode and best_result:
                        self.nextnode.Execute(best_result)              # передаем на следующий этап
                    self.pocket.clear()                                 # чистим контекст
                    self.channel_pocket.clear()
                    if self.ch in self.allowed_channels:                # заполняем текущей строкой контекст
                        self.pocket[(self.ch, self.pci)] = (self.filename, self.date, self.dl_bw, self.rsrp,
                                                            self.lon, self.lat, self.pocket_value_max_age)
                        self.channel_pocket.add(self.ch)
                        self.last_ch = self.ch
                        self.last_date = self.date
                    else:
                        self.last_date = None
                        self.last_ch = -1
                else:
                    conditions = [
                       #(self.ch != self.last_ch and self.ch in self.channel_pocket),
                       #(self.ch == self.last_ch and self.date != self.last_date),
                       #(self.date != self.last_date and self.ch in self.allowed_channels),
                        (self.date != self.last_date)
                    ]
                    new_meas_cycle_detected = any(conditions)

                    if new_meas_cycle_detected:
                        self.pocketAging()
                        best_result = self.pocketDump()  # извлекаем текущий максимум
                        if self.nextnode and best_result:
                            self.nextnode.Execute(best_result)  # передаем на следующий этап

                        #self.pocket_age -= 1
                        #if self.pocket_age == 0:
                        #"    self.pocket_age = 20
                        #"    self.pocket.clear()

                        self.channel_pocket.clear()

                    if self.ch in self.allowed_channels:
                        self.pocket[(self.ch, self.pci)] = (
                            self.filename, self.date, self.dl_bw, self.rsrp,
                            self.lon, self.lat, self.pocket_value_max_age)
                    self.channel_pocket.add(self.ch)

                    self.last_ch = self.ch  # готовимся к следующей смене канала
                    self.last_date = self.date  # готовимся к следующей смене времени


class StateMachine(Filter):
    def __init__(self):
        super().__init__()
        self.usage = 0
        self.keyfield = None
        self.keyfieldIdx = -1
        self.timefield = None
        self.timefieldIdx = -1
        self.bypass = False
        self.state = None
        self.enabledstates = []
        self.transitions = []
        self.firstexec = True

    def Configure(self, config):
        self.columns = self.columns if "columns" not in config else config["columns"]
        self.keyfield = self.keyfield if 'keyfield' not in config else config['keyfield']
        self.keyfieldIdx = super().FindColumn(self.keyfield)
        self.timefield = 'date_time'
        self.timefieldIdx = super().FindColumn(self.timefield)

        self.state = self.state if 'InitialState' not in config else config['InitialState']
        self.enabledstates = self.enabledstates if 'EnabledStates' not in config else config['EnabledStates']
        self.transitions = self.transitions if 'Transitions' not in config else config['Transitions']

        if self.keyfieldIdx == -1 or self.timefieldIdx == -1:
            self.bypass = True

        if self.nextnode:
            self.nextnode.Configure(config)

    def Execute(self, record):
        if record:
            AbstractProcessor.Execute(self)
            if self.bypass:
                self.nextnode.Execute(record)
            else:
                # сбрасываем все триггеры переходов и выставляем начальное время
                if self.firstexec:
                    for t in self.transitions:
                        t['triggered'] = False
                        t['baseTime'] = record[self.timefieldIdx]
                        self.firstexec = False

                currentTime = record[self.timefieldIdx]

                for col in range(len(self.columns)):
                    field_name = self.columns[col]
                    locals()[field_name] = record[col]

                value = record[self.keyfieldIdx]

                if self.state in self.enabledstates:
                    self.nextnode.Execute(record)

                for t in self.transitions:
                    if self.state == t['from']:
                        timeToTrigger = None if 'ttt' not in t else int(t['ttt'])
                        try:
                            if not eval(t['trigger'].strip()):
                                t['triggered'] = False
                            else:
                                if not timeToTrigger:
                                    self.state = t['to']
                                    baseTime = currentTime
                                else:
                                    if not t['triggered']:
                                        t['triggered'] = True
                                        t['baseTime'] = record[self.timefieldIdx]
                                    else:
                                        currentTime = record[self.timefieldIdx]
                                        timeDifference = (currentTime - t['baseTime']).total_seconds()
                                        if timeDifference >= timeToTrigger / 1000:
                                            self.state = t['to']
                                            t['triggered'] = False
                                            t['baseTime'] = currentTime
                        except:
                            pass

