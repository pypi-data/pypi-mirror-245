import math
import copy
import time
import sys
import os
from typing import Union
from openpyxl.styles.cell_style import StyleArray
from tpltable._utils import *
# from filefinder import *

# Handle Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
# openpyxl.worksheet.worksheet.Worksheet
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


def load_workbook(fpath: str):
    """
    加载工作簿
    :param fpath:
    :return:
    """
    wb = openpyxl.load_workbook(fpath)
    fnametype = os.path.basename(fpath)
    wb.name = fnametype
    return wb


class TableFB:  # Table Formatter Backend
    @staticmethod
    def _get_bookinfo(book: Workbook, justify=True) -> str:
        """
        获取工作簿的信息
        :time: 2023-12-06
        """
        _ = getattr(book, 'name', f"Unknown-WorkBook")
        _ = f"[{_}]"
        if justify:
            _ = _.ljust(20)
        return _

    @staticmethod
    def _get_sheetinfo(sheet: Worksheet, justify=True) -> str:
        """
        获取工作表的信息
        :time: 2023-12-06
        """
        parent = getattr(sheet, "parent", None)
        _p = getattr(parent, 'name', f"Unknown-WorkBook")
        _t = getattr(sheet, 'title', f"Unknown-Sheet")
        _ = f"[{_p}.{_t}]"
        if justify:
            _ = _.ljust(25)
        return _

    @staticmethod
    def _calculate_cell_count(sheet:Worksheet):
        """
        统计目标sheet中的单元格数量
        一个合并单元格视作1个单元格
        :time: 2023-12-06
        """
        count = 0
        # 统计所有单元格的数量
        count = sheet.max_row * sheet.max_column
        # 遍历每个合并单元格, 统计其包含的单元格数量
        for merged_cell in sheet.merged_cells:
            # 对count总数减去该合并单元格包含的基本单元格的数量
            count -= (merged_cell.max_row - merged_cell.min_row + 1) * (merged_cell.max_col - merged_cell.min_col + 1) - 1
            # 对count总数加上1, 表示该合并单元格本身
            count += 1

        return count

    @staticmethod
    def _remove_offset_area(ws: Worksheet, offset: list) -> Worksheet:
        """
        移除目标sheet的前offset列和行
        :time: 2023-12-06
        """
        if offset[1] > 0:
            ws.delete_rows(1, offset[1])
        if offset[0] > 0:
            ws.delete_cols(1, offset[0])
        return ws

    @staticmethod
    def _try_show_progress(progress:float, _next_time:float):
        """
        显示进度条
        :time: 2023-12-06
        """
        # print(time.time(), _next_time)
        if time.time() >= _next_time:
            progress *= 100

            # 像这样: [=====>       ] 30%
            # 一共20个等号, 100%的时候是20个等号
            equal_str = '=' * int(progress // 5)
            blank_str = ' ' * (20 - int(progress // 5))
            print(f"[{equal_str + '>' + blank_str}] {progress:.1f}%", end='\r')
            # sys.stdout.flush()
            return True
        return False

    @staticmethod
    def _scan_table(sheet: Worksheet):
        """
        扫描单个sheet, 记录以$开头的单元格(包括合并格)的位置, 以及对应的值
        :param sheet: openpyxl.worksheet.worksheet.Worksheet
        :return:
        :time: 2023-12-06
        """
        tdata = {}

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('$') and cell.coordinate not in tdata:
                    tdata[cell.coordinate] = cell.value

        # 检查values有无重复, 若重复, 报错并指出两个重复值的key
        vk_dict = {}
        for key, val in tdata.items():
            if val not in vk_dict:
                vk_dict[val] = [key]
            elif len(vk_dict[val]) == 1 and (':' in key) == (':' in vk_dict[val][0]):
                raise ValueError(f"values have duplicates: {val} in {key} and {vk_dict[val][0]}")
            elif len(vk_dict[val]) >= 2:
                raise ValueError(f"values have duplicates: {val} in {key} and {vk_dict[val][0]}, {vk_dict[val][1]}")

        return tdata

    @staticmethod
    def _bGET_MERGED_RANGES_LT(sheet: Worksheet):
        """
        获取所有合并单元格的范围和所有合并单元格的左上角
        :param sheet:
        :return:
        """
        # 如果已经保存过, 则直接返回
        if hasattr(sheet, '_bMERGED_RANGES') and hasattr(sheet, '_bMERGED_LT'):
            return sheet._bMERGED_RANGES, sheet._bMERGED_LT

        merged_ranges, merged_lt = [], []
        for merged_cell in sheet.merged_cells.ranges:
            merged_ranges.append(merged_cell.bounds)
            merged_lt.append((merged_cell.min_row, merged_cell.min_col))

        # 保存属性到sheet中
        sheet._bMERGED_RANGES = merged_ranges
        sheet._bMERGED_LT = merged_lt

        return merged_ranges, merged_lt

    @staticmethod
    def _match_table(sheet_template: Worksheet, target_sheet: Worksheet, patience_count: int, debug=False):
        """
        比较两个sheet的内容是否一致, 如果不一致, 返回差异
        :param sheet_template:
        :param target_sheet:
        :param patience_count: 容忍的差异个数
        :param debug: debug时返回str而不是bool
        :return:
        :time: 2023-12-06
        """
        differences = []

        # 获取所有template的合并单元格的范围和所有合并单元格的左上角
        merged_ranges, merged_lt = TableFB._bGET_MERGED_RANGES_LT(sheet_template)

        # 比对单元格内容
        for row in sheet_template.iter_rows():
            for cell in row:
                # 忽略位于合并单元格内的单元格(除非该单元格是合并单元格的左上角)
                if cell.coordinate in merged_ranges and cell.coordinate not in merged_lt:
                    continue
                if not str(cell.value).startswith("$"):
                    target_cell = target_sheet[cell.coordinate]
                    if cell.value != target_cell.value:
                        differences.append(f"单元格 {cell.coordinate} 的内容不一致. 模板: {cell.value}, 目标: {target_cell.value}")

        if len(differences) > patience_count:
            return (f"差异过大: {len(differences)} > {patience_count}:\n" + '\n'.join(differences)) if debug else False
        return True

    @staticmethod
    def _is_empty_row(sheet, irow):
        """
        判断一行是否为空
        :param sheet:
        :param irow:
        :return:
        :time: 2023-12-06
        """
        for cell in sheet[irow]:
            if cell.value:
                return False
        return True

    @staticmethod
    def _is_empty_col(sheet, icol):
        """
        判断一列是否为空
        :param sheet:
        :param icol:
        :return:
        :time: 2023-12-06
        """
        for cell in sheet[icol]:
            if cell.value:
                return False
        return True

    @staticmethod
    def _convert_fDATAr(fDATA):
        """
        将fDATA中的key转换为range4: (min_row, max_row, min_col, max_col)
        返回一个新的fDATAr
        :time: 2023-12-06
        """
        fDATAr = {}
        for key, value in fDATA.items():
            _min_col, _min_row, _max_col, _max_row = range_boundaries(key)
            new_key = (_min_row, _max_row, _min_col, _max_col)
            fDATAr[new_key] = value
        return fDATAr
    @staticmethod
    def _calculate_tplrange4(sheet, fDATA):
        """
        计算fDATA的range4
        :time: 2023-12-06
        """
        min_row, max_row, min_col, max_col = float('inf'), 0, float('inf'), 0
        for key in fDATA.keys():
            _min_col, _min_row, _max_col, _max_row = range_boundaries(key)
            min_row, max_row = min(min_row, _min_row), max(max_row, _max_row)
            min_col, max_col = min(min_col, _min_col), max(max_col, _max_col)

        # 逼近min_row和min_col
        for i in range(min_row - 1, 0, -1):
            if not TableFB._is_empty_row(sheet, i):
                min_row = i
                break
        for i in range(min_col - 1, 0, -1):
            if not TableFB._is_empty_col(sheet, i):
                min_col = i
                break

        return min_row, max_row, min_col, max_col

    @staticmethod
    def _calculate_tplshape(sheet, fDATA):
        """
        计算fDATA的长和宽
        :time: 2023-12-06
        """
        min_row, max_row, min_col, max_col = TableFB._calculate_tplrange4(sheet, fDATA)

        height = max_row - min_row + 1
        width = max_col - min_col + 1
        return width, height


    @staticmethod
    def _move_template(fDATAr, n, m):
        """
        "移动"模板fDATAr
        x方向移动n个长度,y方向移动m个宽度
        :time: 2023-12-06
        """
        new_fDATAr = {}
        for key, value in fDATAr.items():
            _min_row, _max_row, _min_col, _max_col = key
            new_key = (_min_row + m, _max_row + m, _min_col + n, _max_col + n)
            new_fDATAr[new_key] = value
        return new_fDATAr


    # styles = ('_fonts', '_fills', '_borders', '_protections', '_alignments', '_number_formats', '_named_styles', '_table_styles', '_colors', '_differential_styles')
    styles_attrs = {
        '_fonts': 'fontId',
        '_fills': 'fillId',
        '_borders': 'borderId',
        '_protections': 'protectionId',
        '_alignments': 'alignmentId',
        '_number_formats': 'numFmtId',
        '_named_styles': 'xfId',
        # pivotButton?
        # quotePrefix?
    }
    @staticmethod
    def _bSHARE_STYLES(book_source, book_target):
        """
        共享样式数据到目标book
        :time: 2023-12-06
        """
        if getattr(book_target, '_SHARE_WITH', None) == id(book_source):
            return
        for style in TableFB.styles_attrs:
            coll_source = getattr(book_source, style)
            setattr(book_target, style, coll_source)
        book_target._SHARE_WITH = id(book_source)
    @staticmethod
    def _copy_style(src_cell, dest_cell):
        """
        复制源单元格的样式到目标单元格。
        :param src_cell: 源单元格
        :param dest_cell: 目标单元格
        :time: 2023-12-06
        """
        # if src_cell.has_style:
        #     dest_cell.font = copy.copy(src_cell.font)
        if src_cell.has_style:
            TableFB._bSHARE_STYLES(src_cell.parent.parent, dest_cell.parent.parent)
            if not getattr(dest_cell, "_style"):
                dest_cell._style = StyleArray()
            for style, attr in TableFB.styles_attrs.items():
                _value = getattr(src_cell._style, attr, None)
                if _value is not None:
                    setattr(dest_cell._style, attr, _value)

    @staticmethod
    def _iterate_over_target(target: Worksheet, fDATA, force_shape):
        """
        迭代器，输入target:WorkSheet, fDATA, force_shape:tuple (col, row)
        每次yield一个新的WorkSheet, 这个新的ws只有模板的长宽, 按照先行后列的顺序在target上"移动"模板
        :time: 2023-12-06
        """
        # fDATAr = TableFB._convert_fDATAr(fDATA)
        template_width, template_height = TableFB._calculate_tplshape(target, fDATA)
        rows, cols = target.max_row, target.max_column
        template_ix = 0
        template_iy = 0

        if force_shape is None:
            # Auto Scan Mode
            range_row = range(1, rows + 1, template_height)
            range_col = range(1, cols + 1, template_width)
        else:
            # Force Shape Mode
            range_row = range(1, template_height * force_shape[0] - 1, template_height)
            range_col = range(1, template_width * force_shape[1] - 1, template_width)
            # 补全目标:
            # 如果目标的长宽不足, 则在尾部新增空白行/列
            if rows < template_height * force_shape[0]:
                target.insert_rows(rows + 1, template_height * force_shape[0] - rows)
            if cols < template_width * force_shape[1]:
                target.insert_cols(cols + 1, template_width * force_shape[1] - cols)
            rows, cols = max(rows, template_height * force_shape[0]), max(cols, template_width * force_shape[1])

        merge_cells = target.merged_cells
        for i in range_row:  # 第一个维度是行(对应y轴)
            for j in range_col:  # 第二个维度是列(对应x轴)
                if i + template_height - 1 > rows or j + template_width - 1 > cols:
                    continue
                wb = Workbook()
                ws = wb.active
                for row in range(i, i + template_height):
                    for col in range(j, j + template_width):
                        # value font border fill
                        cell = target.cell(row=row, column=col)
                        new_cell = ws.cell(row=row - i + 1, column=col - j + 1)
                        new_cell.value = cell.value
                        TableFB._copy_style(cell, new_cell)

                # 获取目标区域内的所有合并单元格
                for merged in merge_cells.ranges:
                    if merged.min_row >= i and merged.max_row <= i + template_height - 1 and merged.min_col >= j and merged.max_col <= j + template_width - 1:
                        ws.merge_cells(start_row=merged.min_row - i + 1, start_column=merged.min_col - j + 1,
                                       end_row=merged.max_row - i + 1, end_column=merged.max_col - j + 1)

                if template_ix >= cols // template_width:
                    template_ix = 0
                    template_iy += 1

                yield (template_ix, template_iy), (i, i + template_height - 1, j, j + template_width - 1), ws
                template_ix += 1

    @staticmethod
    def _copy_to(source: Worksheet, target: Worksheet, source_range4: tuple, target_range4: tuple):
        """
        复制source特定区域中的数据到target的特定区域中
        *区域大小必须相同*
        :param source: 源sheet
        :param target: 目标sheet
        :param source_range4: (min_row, max_row, min_col, max_col)
        :param target_range4: (min_row, max_row, min_col, max_col)
        :return:
        :time: 2023-12-06
        """
        assert len(source_range4) == 4, f"source_range4 must be a tuple of 4 elements. But got {len(source_range4)} elements."
        assert len(target_range4) == 4, f"target_range4 must be a tuple of 4 elements. But got {len(target_range4)} elements."
        assert source_range4[1] - source_range4[0] == target_range4[1] - target_range4[0], \
            f"source_range4 and target_range4 must have the same height. But got {source_range4[1] - source_range4[0]} and {target_range4[1] - target_range4[0]}"

        for i in range(source_range4[0], source_range4[1] + 1):
            for j in range(source_range4[2], source_range4[3] + 1):
                src_cell = source.cell(row=i, column=j)
                dest_cell = target.cell(row=i + target_range4[0] - source_range4[0], column=j + target_range4[2] - source_range4[2])
                dest_cell.value = src_cell.value
                TableFB._copy_style(src_cell, dest_cell)

        # Copy merged cells
        for merged_cell_range in source.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_cell_range.bounds
            if (min_row >= source_range4[0] and max_row <= source_range4[1] and
                    min_col >= source_range4[2] and max_col <= source_range4[3]):
                target.merge_cells(
                    start_row=min_row + target_range4[0] - source_range4[0],
                    start_column=min_col + target_range4[2] - source_range4[2],
                    end_row=max_row + target_range4[0] - source_range4[0],
                    end_column=max_col + target_range4[2] - source_range4[2]
                )

    @staticmethod
    def _paste_to(target: Worksheet, range4: tuple, lib: dict):
        """
        将lib中的数据写入到target中
        :param target: 目标sheet
        :param range4: (min_row, max_row, min_col, max_col)
        :param lib: dict, key为以$开头的字符串, value为值. 当对应单元格的数据为$xxx时, 将其替换为value
        :return:
        :time: 2023-12-06
        """
        assert len(range4) == 4, f"range4 must be a tuple of 4 elements. But got {len(range4)} elements."

        for i in range(range4[0], range4[1] + 1):
            for j in range(range4[2], range4[3] + 1):
                cell = target.cell(row=i, column=j)
                if isinstance(cell.value, str) and cell.value.startswith('$'):
                    if cell.value in lib:
                        cell.value = lib[cell.value]
                    else:
                        wPrint("TableFB._paste_to", f"{cell.value} not found in lib.")


class BookTableFB:

    @staticmethod
    def _check_excel_fpath(excel_fpath: str, error=True):
        """
        检查excel文件路径是否正确
        :time: 2023-12-06
        """
        if not os.path.exists(excel_fpath):
            if error:
                raise FileNotFoundError(f"File not found: {excel_fpath}")
            else:
                return False
        if not os.path.isfile(excel_fpath):
            if error:
                raise TypeError(f"{excel_fpath} is not a file.")
            else:
                return False
        if excel_fpath[-4:].lower() not in ('.xls', 'xlsx'):
            if error:
                raise TypeError(f"{excel_fpath} is not a excel file.")
            else:
                return False
        return True

    @staticmethod
    def _bIS_TARGET_MIXED(target: list) -> bool:
        """
        检查目标对象是否是混合对象列表
        :param target: list of Workbook|Worksheet
        :return: bool
        :time: 2023-12-06
        """
        types = set([type(t) for t in target])
        return Workbook in types and Worksheet in types

    @staticmethod
    def _bGET_LISTARGET_TYPE(target: list) -> type:
        """
        返回目标列表中元素的类型
        :param target: list of Workbook|Worksheet
        :return: type
        :time: 2023-12-06
        """
        if not BookTableFB._bIS_TARGET_MIXED(target):
            return type(target[0])
        else:
            raise ValueError(f"目标对象类型不正确，应为Workbook, Worksheet, 或 list of Worksheet|Workbook. But got list of mixed type.")

    @staticmethod
    def _bEXPAND_TARGET_TO_SHEETS(target: Union[Workbook, Worksheet, list]) -> list:
        """
        将目标对象转换为list of Worksheet
        :param target: 可以为Workbook, Worksheet, list of Worksheet|Workbook
        :return: list
        :time: 2023-12-06
        """
        if isinstance(target, Workbook):
            return target.worksheets
        elif isinstance(target, Worksheet):
            return [target]
        elif isinstance(target, (list, tuple)):
            sheets = []
            for t in target:
                if isinstance(t, Workbook):
                    sheets.extend(t.worksheets)
                elif isinstance(t, Worksheet):
                    sheets.append(t)
            return sheets
        else:
            raise ValueError("目标对象类型不正确，应为Workbook, Worksheet, 或 list of Worksheet|Workbook")

    @staticmethod
    def _get_workbook_name(wb: Workbook) -> str:
        """
        获取Workbook的名称
        :param wb:
        :return:
        :time: 2023-12-06
        """
        name = getattr(wb, 'name', None)
        if name is None:
            name = wb.name = f"WorkBook<{id(wb)}>"
        return name
