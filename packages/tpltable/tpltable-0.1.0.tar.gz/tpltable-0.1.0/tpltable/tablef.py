from typing import Tuple, Union, Sized

import numpy as np
import openpyxl
from tpltable._tablefb import *
from tpltable.areasdata import AData


"""
加载目标文件的模板, 记录以$开头的单元格的位置, 以及对应的值
之后需要创建新的文件时, 将新的值填入对应的位置
"""

MAX_TIMEDELTA = 1  # 最大显示时间差, 单位: s

class TableF(TableFB):
    """
    计划用于单个2dTable的处理
    不考虑多个sheet的情况
    """

    def __init__(self, tSHEET: Worksheet):
        """

        :param tSHEET:
        """
        self.tSHEET = tSHEET  # Template Sheet
        self.fDATA = self._scan_table(self.tSHEET)  # formatter data

        # ----
        self._pCOUNT = None
        self._offset = [0, 0]   # (row, col)

    def __len__(self):
        if self._pCOUNT is None:
            self._pCOUNT = self._calculate_cell_count(self.tSHEET)
        return self._pCOUNT


    @property
    def offset(self):
        """
        offset: col, row
        设置后:
            readf时可以忽略目标的前几列和前几行
                *一般readf时, 目标的offset范围内是固定的表头. 不需要且不能参与模板匹配环节*
            writef时可以只拷贝一次模板的前几列和前几行
                *一般writef时, 模板的offset范围内是固定的表头. 不需要多次拷贝*
        :time: 2023-12-06
        """
        return self._offset

    @offset.setter
    def offset(self, value):
        assert isinstance(value, Sized), f"offset must be a Sized. But got {type(value)}"
        assert len(value) == 2, f"offset must be a Sized with len 2. But got len {len(value)}"
        self._offset[0] = value[0]
        self._offset[1] = value[1]

    def match(self, target: Worksheet, tolerance_rate=0.02, force_shape=None, debug=False):
        """
        比较模板是否匹配目标
        :param target:
        :param tolerance_rate: 容忍的比例, 比如0.02表示容忍2%的差异
        :param force_shape: 强制指定读取时模板的扫描shape, 为None时, 会自动计算
            *比如模板大小为4*5, 该参数设定为(2, 3) 表示强制读取目标的(A1:H20)区域为2x3个模板的数据*
            *比如模板大小为4*5, 该参数设定为(1, 1) 表示强制读取目标的(A1:D5)区域为1x1个模板的数据*
        :param debug:
        :return:
        :time: 2023-12-06
        """
        # 读取数据
        results, layer_result = [], []

        # crop
        r_target = target
        target = self._remove_offset_area(target, self.offset)

        # 以模板fDATA为基准, 扫描target. 迭代器返回的是: (tpos, range4, ws), 其中ws是target的一个可以被fDATA匹配的子区域
        for tpos, range4, ws in self._iterate_over_target(target, self.fDATA):
            tx, ty = tpos
            if tx == 0:
                layer_result = []
                results.append(layer_result)

            # 检查匹配性
            _match = self._match_table(self.tSHEET, ws, int(tolerance_rate * len(self)), debug)
            if debug:
                if isinstance(_match, str):
                    wPrint("Sheet.Match", self._get_sheetinfo(r_target), f"TemplatePos: {tpos}|Range4: {range4} -> does not match the target. Reason: \n\t{_match}")
                    layer_result.append(False)
                    continue
            if _match is False:
                wPrint("Sheet.Match", self._get_sheetinfo(r_target), f"TemplatePos: {tpos}|Range4: {range4} -> does not match the target.")
                layer_result.append(False)
                continue
            #  --- 检查结束 ---
            layer_result.append(True)
        return np.array(results)

    def readf(self, target: Worksheet, tolerance_rate=0.02, force_shape=None, debug=False):
        """
        格式化读取单个sheet的数据
        例如:
        self.fDATA = {"A1": "$name", "B1:D1": "$age"}
        return: {"$name": "xxx", "$age": "xxx"}

        :param target: 目标sheet
        :param tolerance_rate: 容忍的差异比例, 比如0.02表示容忍2%的差异
        :param force_shape: 强制指定读取时模板的扫描shape, 为None时, 会自动计算
            *比如模板大小为4*5, 该参数设定为(2, 3) 表示强制读取目标的(A1:H20)区域为2x3个模板的数据*
            *比如模板大小为4*5, 该参数设定为(1, 1) 表示强制读取目标的(A1:D5)区域为1x1个模板的数据*
        :param debug: debug模式
        :return: ndarray of dict
        """
        # 读取数据
        results, layer_result, flag = [], [], 0
        fDATAr = self._convert_fDATAr(self.fDATA)

        # crop
        r_target = target
        target = self._remove_offset_area(target, self.offset)

        for tpos, range4, ws in self._iterate_over_target(target, self.fDATA, force_shape):
            flag = 1
            tx, ty = tpos
            if tx == 0:
                layer_result = []
                results.append(layer_result)

            # 检查匹配性
            _match = self._match_table(self.tSHEET, ws, int(tolerance_rate * len(self)), debug)
            if debug:
                if isinstance(_match, str):
                    wPrint("Sheet.ReadF", self._get_sheetinfo(r_target), f"TemplatePos: {tpos}|Range4: {range4} -> does not match the target. Reason: \n\t{_match}")
                    layer_result.append({})
                    continue
            if _match is False:
                wPrint("Sheet.ReadF", self._get_sheetinfo(r_target), f"TemplatePos: {tpos}|Range4: {range4} -> does not match the target.")
                layer_result.append({})
                continue
            #  --- 检查结束 ---

            data = {}
            for key, value in fDATAr.items():
                _min_row, _max_row, _min_col, _max_col = key
                if _min_row == _max_row and _min_col == _max_col:
                    data[value] = ws.cell(row=_min_row, column=_min_col).value
                else:
                    data[value] = [ws.cell(row=row, column=col).value for row in range(_min_row, _max_row + 1) for col in
                                   range(_min_col, _max_col + 1)]
            layer_result.append(data)

        if flag == 0:
            wPrint("Sheet.ReadF", f"Target is smaller than template. Target-Size: {target.max_row}x{target.max_column} < Template-Size: {self.tSHEET.max_row}x{self.tSHEET.max_column}")
        return np.array(results)

    def writef(self, ndarray_data: np.ndarray, ws: Worksheet = None, append_at: str = 'b', progress=True, debug=False) -> Worksheet:
        """
        格式化写入数据
        :param ndarray_data: ndarray of dict. dict be like: {"$xxx": str}
        :param ws: 待写入的sheet,如果为None, 则创建一个新的sheet
        :param append_at: 写入的位置, 可以是r b rb中的一种
        :param progress: 是否显示进度
        :param debug: 暂时没啥用
        :return:
        """
        assert isinstance(ndarray_data, np.ndarray), f"ndarray_data must be a ndarray. But got {type(ndarray_data)}"
        assert ndarray_data.ndim == 2, f"ndarray_data must be a 2d array. But got {ndarray_data.ndim}d array."
        assert isinstance(ndarray_data[0, 0], dict), f"ndarray_data must be a ndarray of dict. But got {type(ndarray_data[0, 0])}"
        assert append_at in ('r', 'b', 'rb'), f"append_at must be 'r', 'b' or 'rb'. But got {append_at}"
        dshape = ndarray_data.shape
        if ws is None:
            # 创建全新的sheet
            ws = Workbook().active

        # 检查ws最大的行列数, 从下一个位置开始写入
        _max_row, _max_col = ws.max_row, ws.max_column

        if append_at == 'r':
            _start_row, _start_col = _max_row + 1, 1
        elif append_at == 'b':
            _start_row, _start_col = 1, _max_col + 1
        elif append_at == 'rb':
            _start_row, _start_col = _max_row + 1, _max_col + 1
        else:
            raise ValueError(f"Unknown append_at: {append_at}")

        # 检查empty
        if self._is_empty_row(ws, _start_row):
            _start_row -= 1
            _start_row = max(_start_row, 1)
        if self._is_empty_col(ws, _start_col):
            _start_col -= 1
            _start_col = max(_start_col, 1)

        range4 = list(self._calculate_tplrange4(self.tSHEET, self.fDATA))
        range4[0] += self.offset[0]
        range4[2] += self.offset[1]
        tWIDTH, tHEIGHT = range4[3] - range4[2] + 1, range4[1] - range4[0] + 1

        # -- 先拷贝offset内的内容到目标
        # 计算offset的range4
        # 先计算竖着的区域
        oh_range4 = (1, range4[1], 1, self.offset[1])
        self._copy_to(self.tSHEET, ws, oh_range4, oh_range4)
        # 再计算横着的区域
        ov_range4 = (1, self.offset[0], 1, range4[3])
        self._copy_to(self.tSHEET, ws, ov_range4, ov_range4)

        # -- 显示进度(记录时间, 超出预计时间则显示一下进度)
        _next_time = time.time() + MAX_TIMEDELTA
        finish_count, max_count = 0, dshape[0] * dshape[1]
        last_progress = 0
        if progress:
            self._try_show_progress(0, _next_time - MAX_TIMEDELTA - 1)

        # 在目标sheet中创建足够的空白区域 ndarray_data.shape * (tWIDTH, tHEIGHT)
        for i in range(dshape[0]):
            for j in range(dshape[1]):
                _lib = ndarray_data[i, j]
                if not isinstance(_lib, dict):
                    raise TypeError(f"ndarray_data[{i}, {j}] is not a dict. But got {type(_lib)}")
                if not _lib:
                    wPrint("Sheet.WriteF", f"ndarray_data[{i}, {j}] is empty. Skip.")
                    continue
                target_range4 = (
                    self.offset[0] + _start_row + i * tHEIGHT,
                    self.offset[0] + _start_row + (i + 1) * tHEIGHT - 1,
                    self.offset[1] + _start_col + j * tWIDTH,
                    self.offset[1] + _start_col + (j + 1) * tWIDTH - 1
                )
                # 逐个拷贝模板, 使用self._copy_from_to(source: Worksheet, target: Worksheet, source_range4: tuple, target_range4: tuple)
                self._copy_to(self.tSHEET, ws, range4, target_range4)
                # continue  # debug
                # 逐个粘贴数据, 注意要选取ndarray_data中对应位置的dict, 使用self._paste_lib_to(target: Worksheet, range4: tuple, lib: dict)
                if not isinstance(_lib, dict):
                    raise TypeError(f"ndarray_data[{i}, {j}] is not a dict. But got {type(_lib)}")
                self._paste_to(ws, target_range4, _lib)

                # -- 显示进度
                finish_count += 1
                if progress and self._try_show_progress(finish_count / max_count, _next_time):
                    last_progress = finish_count / max_count
                    _next_time = time.time() + MAX_TIMEDELTA

        if progress and abs(last_progress - 1) > 1e-4:
            self._try_show_progress(1, _next_time - MAX_TIMEDELTA - 1)

        # if progress:
        #     print()  # 换行,

        return ws

    @property
    def format(self):
        """
        返回单个sheet的pipe input format
        * 这其实是所有$XXX字段的集合
        :return: list
        """
        return list(self.fDATA.values())



class BookTableF(BookTableFB):
    """
    这是一个名为BookTableF的类，它用于用模板excel表格去匹配目标工作表的数据或使用目标工作表按照模板excel输出。
    BookTableF类：

    tSHEET_EQUIV：表示工作表之间等价的类型。
        @Note: 该模式下, 只会关注模板文件的第一个sheet, 之后的sheet会被忽略. 并试图用模板文件的第一个sheet去匹配目标文件的所有sheet.
    tSHEET_DIFF：表示工作表之间不等价的类型。
        @Note: 该模式下, 会关注模板文件的所有sheet, 并试图用模板文件的每一个sheet去匹配目标文件的对应sheet.
    # ---------------------------------------------------------------------------------->
    > 下面的两个常量仅在tSHEET_DIFF模式下有效
    sITER_INDEX：根据索引来获取工作表。
        @Note: 当stype设为tSHEET_DIFF时, 该模式下, 会按照模板文件的sheet的顺序来匹配目标文件的sheet. 目标文件sheet数量多出模板文件的部分会被忽略.
    sITER_NAME：根据工作表的名称来获取工作表。
        @Note: 当stype设为tSHEET_DIFF时, 该模式下, 会按照模板文件的sheet的名称来匹配目标文件的sheet. 任意匹配不到的sheet都会报错.
    """
    # sheet 间等价的type
    tSHEET_EQUIV = 0
    # sheet 间不等价的type
    tSHEET_DIFF = 1

    # 根据索引来获取sheet
    sITER_INDEX = 0
    # 根据sheet的name来获取sheet
    sITER_NAME = 1

    def __init__(self, fpath: str, sheet_type: int = 0, iter_mode: int = 0):
        assert sheet_type in (self.tSHEET_EQUIV, self.tSHEET_DIFF), \
            f"sheet_type must be tSHEET_EQUIV or tSHEET_DIFF. But got {sheet_type}"
        self.stype = sheet_type  # sheet type
        assert iter_mode in (self.sITER_INDEX, self.sITER_NAME), \
            f"iter_mode must be sITER_INDEX or sITER_NAME. But got {iter_mode}"
        self.siter = iter_mode  # sheet iter mode
        # 检查文件
        self._check_excel_fpath(fpath)
        self.fpath = fpath
        self.wb = load_workbook(fpath)
        self.wss = self.wb.worksheets

        self._offsets = [[0, 0] for _ in range(len(self.wb.worksheets))]  # (row, col)

    @property
    def offsets(self):
        """
        处理数据时的, 忽略目标的前offset[0]行和前offset[1]列
        """
        return self._offsets

    def iterwith(self, target: Union[Workbook, Worksheet, list, str]):
        """
        这个函数会根据stype和siter来遍历模板文件和目标文件的工作表。返回的迭代器一次能返回两个工作表。

        1. 如果stype等于tSHEET_EQUIV，那么函数会遍历目标的每一个工作表，并且每次都返回模板文件的第一个工作表和一个目标工作表。
        2. 如果stype等于tSHEET_DIFF，那么函数的行为会根据siter的值来改变：
            (1). 如果siter等于sITER_INDEX，那么函数会遍历模板文件和目标的每一个工作表，并且每次都返回当前的工作表。(若为list of WorkBook, 则逐个Book进行迭代)
            (2). 如果siter等于sITER_NAME，那么函数会遍历模板文件的每一个工作表，并且每次都返回具有相同名称的工作表。(若为list of WorkBook, 则逐个Book进行迭代)

        :param target: 可以为Workbook, Worksheet, list of Worksheet|Workbook, str(str等同于WorkBook)
            tSHEET_EQUIV 情况下, 支持所有类型的target, 允许输入list of (WorkSheet + Workbook)这种混合数据, 并且会逐sheet进行组合迭代
            tSHEET_DIFF 情况下, 输入单个target会被调整为[target], 不支持list of Workbook和混合数据的输入. 对WorkBook和list of Worksheet进行单组迭代(每次仍然只迭代出两个sheet); 对list of WorkBook进行遍历,逐个Book进行单组迭代
        :return: 返回的迭代器一次能返回两个工作表。
        """
        wbSHEETS = self.wb.worksheets
        if isinstance(target, str):
            self._check_excel_fpath(target)
            target = load_workbook(target)
        if self.stype == self.tSHEET_EQUIV:
            tarSHEETS = self._bEXPAND_TARGET_TO_SHEETS(target)
            for i, target_sheet in enumerate(tarSHEETS):
                yield wbSHEETS[0], target_sheet
        elif self.stype == self.tSHEET_DIFF:
            if not isinstance(target, (list, tuple)):
                target = [target]
            _type = self._bGET_LISTARGET_TYPE(target)
            if _type == Worksheet:
                # 根据siter来获取迭代结果
                tarSHEETS = target
                if len(wbSHEETS) > len(tarSHEETS):
                    raise ValueError(f"目标缺少sheet. 模板sheet数:{len(wbSHEETS)} > 目标sheet数:{len(tarSHEETS)}")

                if self.siter == self.sITER_INDEX:
                    for i, sheet in enumerate(wbSHEETS):
                        yield sheet, tarSHEETS[i]
                elif self.siter == self.sITER_NAME:
                    tarSHEETS_titles = [sheet.title for sheet in tarSHEETS]
                    for sheet in wbSHEETS:
                        if sheet.title not in tarSHEETS_titles:
                            raise ValueError(f"目标缺少特定sheet: {sheet.title}")
                        yield sheet, target[tarSHEETS_titles.index(sheet.title)]
                else:
                    raise ValueError(f"Unknown sheet-iter-mode: {self.siter}")
            elif _type == Workbook:
                if len(target) > 1:
                    raise ValueError(
                        f"目标对象类型不正确，应为Workbook, Worksheet, 或 list of Worksheet  (AT tSHEET_DIFF MODE). But got list of Workbook.")
                for sheet_tpl, sheet_tar in self.iterwith(target[0].worksheets):
                    yield sheet_tpl, sheet_tar
            else:
                raise ValueError(f"Unknown target single type: {_type}")
        else:
            raise ValueError(f"Unknown sheet-type: {self.stype}")

    def match(self, target: Union[Workbook, Worksheet, list, str], tolerance_rate=0.02, simplify=True, debug=False, sheet_match_mode='any') -> Union[
        dict, bool]:
        """
        比较模板是否匹配目标
        :param target: 可以为Workbook, Worksheet, list of Worksheet|Workbook, str(str等同于WorkBook)
            tSHEET_EQUIV 情况下, 支持所有类型的target, 允许输入list of (WorkSheet + Workbook)这种混合数据, 并且会逐sheet进行<单次匹配>
            tSHEET_DIFF 情况下, 输入单个target会被调整为[target], 不支持list of Workbook和混合数据的输入. 对WorkBook和list of Worksheet进行<单组匹配>; 对list of WorkBook进行遍历,逐个Book进行<单组匹配>
                        ------------------------------------------------------------------------------------------------------------------------------------
            @Example:
                假设: ----------------->
                    模板(tpl)文件有两个sheet: s1, s2

                    tar1为单个sheet: tar_st1
                    tar2为list[sheet: s1, sheet: s2]

                    tar3为Workbook: tar_wb1(包含两个sheet: s1, s2)
                    tar4为list[Workbook: tar_wb1(包含两个sheet: s1, s2), Workbook: tar_wb2(包含两个sheet: s1, s2)]

                    tar5为list[sheet: s1, Workbook: tar_wb1(包含两个sheet: s1, s2)]
                >>> # --
                1. 当tSHEET_EQUIV时, 先将target转换为list of Worksheet, 然后逐sheet进行<单次读取>:
                    tar? -> list of Worksheet
                    以模板第一个sheet(即tpl.s1)去与tar?中的每一个sheet进行一次TableF.readf()操作
                    返回值可能像这样:
                        bool
                2. 当tSHEET_DIFF时, 按照siter的值来决定如何组合sheet进行比较:
                    @以tar2为例:
                        按sITER_INDEX进行读取时:
                            > 以模板第一个sheet与tar2中的第一个sheet进行一次TableF.readf()操作
                            > 以模板第二个sheet与tar2中的第二个sheet进行一次TableF.readf()操作
                        按sITER_NAME进行读取时:
                            > 以模板的sheet:s1与tar2中的同名sheet进行一次TableF.readf()操作
                            > 以模板的sheet:s2与tar2中的同名sheet进行一次TableF.readf()操作
                        不论siter为何值, 返回值都像这样:
                            {
                                ((tpl.name, tpl.s1.title), (tar2[0].parent.name, tar2[0].title)): bool,
                                ((tpl.name, tpl.s2.title), (tar2[1].parent.name, tar2[1].title)): bool,
                            }
                    @以tar3为例:
                        将tar3转化为一组sheet: [tar3.s1, tar3.s2], 然后按照tar2的方式进行读取
                    !该模式下不支持list of WorkBook和list of Worksheet|WorkBook的混合输入!
        :param tolerance_rate: 容忍的比例, 只能容忍内容不一致!(合并单元格不同时会直接导致匹配失败)
        :param simplify: 是否允许简化单个元素时的输出
        :param debug:
        :param sheet_match_mode: 'any' or 'all'  -- 用于对TableF.match() -> ndarray of bool 的结果进行处理
        :return: bool | dict {( (tpl.name, tpl.title), (tar.name, tar.title) ): value}
            只进行单个匹配时,返回bool
            其他情况返回dict, key为 (sheet_tpl.parent.name, sheet_tpl.title), (sheet_tar.parent.name, sheet_tar.title)), value为bool
        """
        results = {}
        if isinstance(target, str):
            self._check_excel_fpath(target)
            target = load_workbook(target)
        for sheet_tpl, sheet_tar in self.iterwith(target):
            tblf_tpl = TableF(sheet_tpl)
            tblf_tpl.offset = self.offsets[self.wss.index(sheet_tpl)]
            tblf_match = tblf_tpl.match(sheet_tar, tolerance_rate=tolerance_rate, debug=debug)
            results_key = ((self._get_workbook_name(sheet_tpl.parent), sheet_tpl.title), (self._get_workbook_name(sheet_tar.parent), sheet_tar.title))
            result = (sheet_match_mode == 'any' and tblf_match.any()) or (sheet_match_mode == 'all' and tblf_match.all())
            results[results_key] = result
            if debug and not result:
                wPrint("Book.Match", TableFB._get_sheetinfo(sheet_tar), f"Can not match"
                      f" by {self._get_workbook_name(sheet_tpl.parent)}.{sheet_tpl.title}")

        if len(results) == 1 and simplify:
            return list(results.values())[0]
        return results

    def readf(self, target: Union[Workbook, Worksheet, list, str], tolerance_rate=0.02, force_shape=None, simplify=True, debug=False) -> Union[dict, np.ndarray]:
        """
        格式化读取数据
        :param target: 可以为Workbook, Worksheet, list of Worksheet|Workbook, str(str等同于WorkBook)
            tSHEET_EQUIV 情况下, 支持所有类型的target, 允许输入list of (WorkSheet + Workbook)这种混合数据, 并且会逐sheet进行<单次读取>
            tSHEET_DIFF 情况下, 输入单个target会被调整为[target], 不支持list of Workbook和混合数据的输入. 对WorkBook和list of Worksheet进行<单组读取>; 对list of WorkBook进行遍历,逐个Book进行<单组读取>
            ------------------------------------------------------------------------------------------------------------------------------------
            @Example:
                假设: ----------------->
                    模板(tpl)文件有两个sheet: s1, s2

                    tar1为单个sheet: tar_st1
                    tar2为list[sheet: s1, sheet: s2]

                    tar3为Workbook: tar_wb1(包含两个sheet: s1, s2)
                    tar4为list[Workbook: tar_wb1(包含两个sheet: s1, s2), Workbook: tar_wb2(包含两个sheet: s1, s2)]

                    tar5为list[sheet: s1, Workbook: tar_wb1(包含两个sheet: s1, s2)]
                >>> # --
                1. 当tSHEET_EQUIV时, 先将target转换为list of Worksheet, 然后逐sheet进行<单次读取>:
                    tar? -> list of Worksheet
                    以模板第一个sheet(即tpl.s1)去与tar?中的每一个sheet进行一次TableF.readf()操作
                    返回值可能像这样:
                        ndarray of {"$XXX": str, ...}
                2. 当tSHEET_DIFF时, 按照siter的值来决定如何组合sheet进行比较:
                    @以tar2为例:
                        按sITER_INDEX进行读取时:
                            > 以模板第一个sheet与tar2中的第一个sheet进行一次TableF.readf()操作
                            > 以模板第二个sheet与tar2中的第二个sheet进行一次TableF.readf()操作
                        按sITER_NAME进行读取时:
                            > 以模板的sheet:s1与tar2中的同名sheet进行一次TableF.readf()操作
                            > 以模板的sheet:s2与tar2中的同名sheet进行一次TableF.readf()操作
                        不论siter为何值, 返回值都像这样:
                            {
                                (tar2[0].parent.name, tar2[0].title): ndarray of {"$XXX": str, ...},
                                (tar2[1].parent.name, tar2[1].title): ndarray of {"$XXX": str, ...},
                            }
                    @以tar3为例:
                        将tar3转化为一组sheet: [tar3.s1, tar3.s2], 然后按照tar2的方式进行读取

                    !该模式下不支持list of WorkBook和list of Worksheet|WorkBook的混合输入!
        >>> # --
        :param tolerance_rate: 容忍的比例, 只能容忍内容不一致!(合并单元格不同时会直接导致匹配失败)
        :param force_shape: 强制指定读取时模板的扫描shape, 为None时, 会自动计算
        :param simplify: 是否允许简化单个元素时的输出
        :param debug:
        :return: ndarray of dict {(tar.name, tar.title): value}
            当ndarray数量超过1时, 返回ndarray of dict
        """
        results = {}
        if isinstance(target, str):
            self._check_excel_fpath(target)
            target = load_workbook(target)
        for sheet_tpl, sheet_tar in self.iterwith(target):
            tblf_tpl = TableF(sheet_tpl)
            tblf_tpl.offset = self.offsets[self.wss.index(sheet_tpl)]
            tblf_read = tblf_tpl.readf(sheet_tar, tolerance_rate=tolerance_rate, force_shape=force_shape, debug=debug)
            results_key = (self._get_workbook_name(sheet_tar.parent), sheet_tar.title)
            results[results_key] = tblf_read
            if debug:
                if not tblf_read:
                    wPrint("Book.ReadF", TableFB._get_bookinfo(target), f"Can not readF by {self._get_workbook_name(sheet_tpl.parent)}.{sheet_tpl.title}")
                else:
                    iPrint("Book.ReadF", TableFB._get_bookinfo(target), f"ReadF by {self._get_workbook_name(sheet_tpl.parent)}.{sheet_tpl.title}")
        if len(results) == 1 and simplify:
            return list(results.values())[0]
        return results

    def writef(self, ndarray_data: np.ndarray, wb: Workbook = None, append_at: str = 'b', progress=True, debug=False) -> Workbook:
        """
        格式化写入数据
        只会写在模板文件的第一个sheet中
        :param ndarray_data: ndarray of dict. dict be like: {"$xxx": str}
        :param wb: 待写入的Workbook,如果为None, 则创建一个新的Workbook
        :param append_at: 写入的位置, 可以是r b rb中的一种
        :param debug: 暂时没啥用
        :return:
        """
        if wb is None:
            wb = Workbook()
        ws = wb.active
        tblf_tpl = TableF(self.wb.worksheets[0])
        tblf_tpl.offset = self.offsets[0]
        tblf_tpl.writef(ndarray_data, ws, append_at, progress, debug)
        return wb

    @property
    def format(self) -> Union[dict, list]:
        """
        返回模板文件的pipe input format

        :return: dict {key: ["$XXX", ...]}|list ["$XXX", ...]
            key与stype和siter有关
            当stype为tSHEET_EQUIV时或DIFF且siter为sITER_INDEX时, key为int > 0
            当stype为tSHEET_DIFF且siter为sITER_NAME时, key为str 即 sheet.title
            ! 当返回的dict的长度为1时, 返回的是list !
        """
        if self.stype == self.tSHEET_EQUIV or (self.stype == self.tSHEET_DIFF and self.siter == self.sITER_INDEX):
            _ = {i: TableF(sht).format for i, sht in enumerate(self.wb.worksheets)}
        elif self.stype == self.tSHEET_DIFF and self.siter == self.sITER_NAME:
            _ = {sht.title: TableF(sht).format for sht in self.wb.worksheets}
        else:
            raise ValueError(f"Unknown sheet-type: {self.stype} and sheet-iter-mode: {self.siter}")
        if len(_) == 1:
            return list(_.values())[0]
        return _


if __name__ == '__main__':
    tpl = BookTableF('tpl.xlsx')
    _ = tpl.readf('a.xlsx', debug=True)
    print(_)

