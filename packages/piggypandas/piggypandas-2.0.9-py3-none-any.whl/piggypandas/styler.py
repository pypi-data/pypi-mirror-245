import pandas as pd
from pandas.io.formats.style import Styler
import re
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet
from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
from typing import Sequence, Callable, Tuple, List, Union, Optional, Set


Mask = Union[str, re.Pattern]
MaskList = Sequence[Mask]
Mask2Style = Tuple[Mask, Union[str, Callable]]
Mask2StyleMap = Sequence[Mask2Style]
Mask2Size = Tuple[Mask, int]
Mask2SizeMap = Sequence[Mask2Size]


class MassColumnStyler:

    def __init__(self, style_map: Optional[Mask2StyleMap] = None):
        self._map: List[Mask2Style] = list()
        if style_map is not None:
            self.extend(style_map)

    def clear(self):
        self._map.clear()

    def append(self, style: Mask2Style):
        col_mask, col_style = style
        if not isinstance(col_mask, re.Pattern):
            col_mask = re.compile(col_mask, re.I)
        if isinstance(col_style, str):
            self._map.append((col_mask, lambda _: col_style))
        else:
            self._map.append((col_mask, col_style))

    def extend(self, style_map: Mask2StyleMap):
        for x in style_map:
            self.append(x)

    def apply_column_styles(self, styler: Styler, first_match: bool = True) -> Styler:
        for col_name in styler.data.columns:
            for col_mask, col_style in self._map:
                if col_mask.search(col_name):
                    styler = styler.applymap(lambda x: col_style(x), subset=col_name)
                    if first_match:
                        break
        return styler


class ExcelColumnSizer:

    def __init__(self, size_map: Optional[Mask2SizeMap]):
        self._map: List[Mask2Size] = list()
        if size_map is not None:
            self.extend(size_map)

    def clear(self):
        self._map.clear()

    def append(self, size: Mask2Size):
        col_mask, col_size = size
        if not isinstance(col_mask, re.Pattern):
            col_mask = re.compile(col_mask, re.I)
        self._map.append((col_mask, col_size))

    def extend(self, size_map: Mask2SizeMap):
        for x in size_map:
            self.append(x)

    @staticmethod
    def _build_sheet_name_list(writer: pd.ExcelWriter, sheets: Optional[MaskList] = None) -> Sequence[str]:
        if sheets is not None:
            sheet_set: Set[str] = set()
            for sheet_mask in sheets:
                if not isinstance(sheet_mask, re.Pattern):
                    sheet_mask = re.compile(sheet_mask, re.I)
                sheet_set = sheet_set.union([s for s in writer.sheets.keys() if sheet_mask.search(s)])
                return [s for s in sheet_set]
        else:
            return [s for s in writer.sheets.keys()]

    def set_column_widths(self, writer: pd.ExcelWriter, sheets: Optional[MaskList] = None, first_match: bool = True):
        sheet_names = ExcelColumnSizer._build_sheet_name_list(writer, sheets)
        for sheet_name in sheet_names:
            if writer.engine == 'openpyxl':
                ws: OpenpyxlWorksheet = writer.sheets[sheet_name]
                for col_index in range(ws.min_column, ws.max_column + 1):
                    cell = ws.cell(row=ws.min_row, column=col_index)
                    col_header = str(cell.value).strip()
                    for col_mask, col_width in self._map:
                        if col_mask.search(col_header):
                            ws.column_dimensions[openpyxl_get_column_letter(col_index)].width = col_width
                            if first_match:
                                break
