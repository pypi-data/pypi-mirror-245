############################################################################
# opsanalyze version 2.1
# Created by :  Okhtai Alizadeh Arasi
# LinkedIn URL: linkedin.com/in/oktai-alizade-94aa5538
# Mobile - whatsapp: +989144011724
# Telegram channel: https://t.me/OKprograms
# Telegram: @Oktai_Arasi
# Instagram: opensees_apps
# December 4, 2023
############################################################################
'''
A python package for performing different kinds of analysis on openseespy models.
'''

import os
from datetime import datetime
import numpy as np
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl import workbook
import webbrowser

def find_nodes(ops, xlim=[], ylim=[], zlim=[]):
    '''
    function to find nodes inside a region defined by xlim, ylim and zlim

    ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
    xlim: A python list containing minimum and maximun x of the region (optional, default value is an empty list)
    ylim: A python list containing minimum and maximun y of the region (optional, default value is an empty list)
    zlim: A python list containing minimum and maximun z of the region (optional, default value is an empty list)

    return: A python list containing the tags of the nodes in the region
    '''

    if len(xlim) not in [0, 2]:
        print('xlim must be an empty list(default) or a list with two float number. default will be used ')
        xlim = []

    if len(zlim) not in [0, 2]:
        print('zlim must be an empty list(default) or a list with two float number. default will be used ')
        zlim = []

    if len(ylim) not in [0, 2]:
        print('ylim must be an empty list(default) or a list with two float number. default will be used ')
        ylim = []

    nodeTags = []
    for nod in ops.getNodeTags():
        xyz = ops.nodeCoord(nod)
        if len(xyz) == 2:
            xyz.append(0.0)

        add_node = True
        if len(xlim) != 0:
            for xx in [xyz[0]]:
                if xx < xlim[0] or xx > xlim[1]:
                    add_node = False

        if len(ylim) != 0:
            for yy in [xyz[1]]:
                if yy < ylim[0] or yy > ylim[1]:
                    add_node = False

        if len(zlim) != 0:
            for zz in [xyz[2]]:
                if zz < zlim[0] or zz > zlim[1]:
                    add_node = False

        if add_node != False:
            nodeTags.append(nod)

    if len(nodeTags) == 0:
            return False
    else:
        return nodeTags


def find_elements(ops, xlim=[], ylim=[], zlim=[]):
    '''
    function to find elements inside a region defined by xlim, ylim and zlim

    ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
    xlim: A python list containing minimum and maximun x of the region (optional, default value is an empty list)
    ylim: A python list containing minimum and maximun y of the region (optional, default value is an empty list)
    zlim: A python list containing minimum and maximun z of the region (optional, default value is an empty list)

    return: A python list containing the tags of the elements in the region
    '''
    if len(xlim) not in [0, 2]:
        print('xlim must be an empty list(default) or a list with two float number. default will be used ')
        xlim = []

    if len(zlim) not in [0, 2]:
        print('zlim must be an empty list(default) or a list with two float number. default will be used ')
        zlim = []

    if len(ylim) not in [0, 2]:
        print('ylim must be an empty list(default) or a list with two float number. default will be used ')
        ylim = []

    eleTags = []
    for ele in ops.getEleTags():
        elenodes = ops.eleNodes(ele)
        XX = []
        YY = []
        ZZ = []
        for nod in elenodes:
            xyz = ops.nodeCoord(nod)

            XX.append(xyz[0])
            YY.append(xyz[1])
            if len(xyz) == 2:
                ZZ.append(0.0)
            else:
                ZZ.append(xyz[2])

        add_ele = True
        if len(xlim) != 0:
            for xx in XX:
                if xx < xlim[0] or xx > xlim[1]:
                    add_ele = False
        if len(ylim) != 0:
            for yy in YY:
                if yy < ylim[0] or yy > ylim[1]:
                    add_ele = False
        if len(zlim) != 0:
            for zz in ZZ:
                if zz < zlim[0] or zz > zlim[1]:
                    add_ele = False

        if add_ele != False:
            eleTags.append(ele)

    if len(eleTags) == 0:
        return False
    else:
        return eleTags


def eigen(ops, name_project, name_analysis='eigen', num_Modes=3, solver='-genBandArpack'):
    """
    Function to perform eigen analysis.

    name_project: Project name
    name_analysis: Analysis name
    ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
    num_Modes: Number of eigenvalues required.(Optional, default value is 3)
    solver: String detailing type of solver: '-genBandArpack', '-fullGenLapack', (Optional, default value is '-genBandArpack')
    return: lambda, omega, Tn
    """

    print('######################')
    print('### Eigen Analysis[', name_analysis,']')
    print('######################')


    _name = name_project
    name = name_analysis

    start_time = datetime.now().replace(microsecond=0)
    print('Start Time: {}'.format(start_time))
    print("------------")
    logfilename = _name + '\\opslogfile_' + name + '.txt'
    ops.logFile(logfilename, '-noEcho')

    try:
        if not os.path.exists(_name + "\\"):
            os.mkdir(_name + "\\")

        lambdaN = ops.eigen(solver, num_Modes)
        omega = []
        Tn = []
        for i in range(num_Modes):
            lambdaI = lambdaN[i]
            omega.append(pow(lambdaI, 0.5))
            tt = (2 * np.pi) / pow(lambdaI, 0.5)
            Tn.append(tt)
            print('T' + str(i+1) + ' = ' + str(round(tt, 3)) + '     f' + str(i+1) + ' = ' + str(round(1 / tt, 3)))

        nodes_ = [str(x) for x in ops.getNodeTags()]
        cols = ['mode_number']
        cols.extend(nodes_)
        otputs = pd.DataFrame(columns=cols)

        for i in range(num_Modes):
            vectors = []
            for nod in ops.getNodeTags():
                # magn = np.sqrt(np.sum(np.array(ops.nodeEigenvector(nod, i + 1)) ** 2))
                vectors.append(np.array(ops.nodeEigenvector(nod, i + 1)))

            data_ = [i + 1]
            data_.extend(vectors)
            otputs.loc[i + 1] = data_

        print('Analysis successful')

        print("------------")
        end_time = datetime.now().replace(microsecond=0)
        print('End Time: {}'.format(end_time))
        print('Duration: {}'.format(end_time - start_time))
        print("------------")


        if os.path.exists(_name + "\\output_" + name + "\\"):
            shutil.rmtree(_name + "\\output_" + name + "\\")

        os.mkdir(_name + "\\output_" + name + "\\")
        file_re = _name + "\\output_" + name + "\\" + ".feather"
        db_output = otputs.reset_index()
        db_output.to_feather(file_re, compression='zstd')

        print('Data was stored in ' + _name + "\\output_" + name)
        print("     ")

        return lambdaN, omega, Tn

    except KeyboardInterrupt:
        print("Analysis terminated by user")
        exit()
    finally:
        print("Log file: " + logfilename)

def __eigen(ops,  num_Modes, solver='-genBandArpack'):
    """
    Function to perform eigen analysis.
    ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
    num_Modes: Number of eigenvalues required.(Optional, default value is 3)
    solver: String detailing type of solver: '-genBandArpack', '-fullGenLapack', (Optional, default value is '-genBandArpack')
    return: lambda, omega, Tn
    """

    try:

        lambdaN = ops.eigen(solver, num_Modes)
        omega = []
        Tn = []
        for i in range(num_Modes):
            lambdaI = lambdaN[i]
            omega.append(pow(lambdaI, 0.5))
            tt = (2 * np.pi) / pow(lambdaI, 0.5)
            Tn.append(tt)

        return lambdaN, omega, Tn

    except KeyboardInterrupt:
        print("Analysis terminated by user")
        exit()

def analyze_push_cyclic(ops, name_project, name_analysis, analysis_option, filename_protocole, cnodeTag, cdof, Dy, du_min,
                        du_max, duDiv=2, resp_nodes=['Disp', 'Reaction'], resp_elements=['force', 'stresses', 'strains'],
                        resp_sections=[], nodes=[], elements=[], elements_section=[], everynstep=1, numIter=10,
                        exitrun='y'):

    """
    Function to perform pushover cyclic analysis.

    ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
    name_project: Project name
    name_analysis: Analysis name
    analysis_option: A python dictionary contains analysis options.
        Example:
        analysis_option = {'constraints': ['Plain'],
                           'numberer': ['Plain'],
                           'system': ['BandGeneral'],
                           'test': ['NormUnbalance', 1.0e-5, 1000],
                           'algorithm': ['NewtonLineSearch']}
    filename_protocol: A two-column data file including steps and displacements.
    cnodeTag: node whose response controls solution.
    cdof: degree of freedom at the node.
    Dy: Yield displacement
    du_min: the min stepsize the user will allow.
    du_max: the max stepsize the user will allow.
    du_Div:Refer to "How it works?"
    resp_nodes : A python list containing the type of node responses that are stored (optional, default value is ['Disp', 'Reaction'])
    resp_elements: A python list containing the type of element responses that are stored (optional, default value is ['force', 'stresses', 'strains'])
    resp_sections: See the notes below (optional, default value is an empty list)
    nodes: A python list containing tags of the nodes whose responses are stored (optional, default value is an empty list)
    elements: A python list containing tags of the elements whose responses are stored (optional, default value is an empty list)
    elements_section: A python list containing tags of the elements whose section responses are stored (optional, default value is an empty list)
    everynstep: The frequency of outputs that are stored. (Optional, default value is every 1 increments)
    numIter: the number of iterations the user would like to occur in the solution algorithm(Optional, default value is 10).
    exitrun: whether to exit execution when encounters an unconvergance issue or terminates the current analysis and continue.
             Valid values are 'yes', 'y', 'no' or 'n' (optional, default value is 'y').

    How it works?
    - Analysis is started with du = du_max.
    - If the analysis does not converge at a certain step, du is reduced by du_Div times(du = du / du_Div).
         This continues until the analysis converges at that step or du becomes smaller than du_min in which case the
         analysis is terminated.
    - After 10 successful steps du is increased by du_Div times(du = du * du_Div).
      This continues until du becomes greater that du_max in which case du will be set to du_max.
    - At the end of each cycle program adjusts du so that the location corresponds to the peak displacement in that cycle.

    Note:
        - The program saves the data after the analysis is complete successfully or terminates by unconvergence issue.
        - Valid values for resp_nodes are:
            Disp
            Vel
            Accel
            IncrDisp
            IncrDeltaDisp
            Reaction
            Unbalance
            RayleighForces
        - resp_sections is a 2D array of Nx4 size. each row contains section number, fiber y coordinate, fiber z coordinate and fiber material tag.
          Example:
              resp_sections =[[num_section1, y1, z1, matTag1], [num_section2, y2, z2, matTag2], ...]
              resp_sections = [[1, 0.0, 0.0, 1], [5, 0.3, 0.3, 1]]
        - elements_section list must contains beam-column element tags.
        - Storage files are saved in name_project\output_name_analysis\ respname.feather path.
          Example: for name_project = myproject, name_analysis = ntha, Storage file pathes would be
                 myproject\output_ntha\Disp.feather
                 myproject\output_ntha\Reaction.feather
                 myproject\output_ntha\force.feather
                 ...
    """
    _name = name_project
    name = name_analysis

    print('#################################')
    print('### Push Over Cyclic Analysis[', name_analysis, ']')
    print('#################################')

    if not os.path.exists(_name + "\\"):
        os.mkdir(_name + "\\")

    logfilename = _name + '\\opslogfile_' + name + '.txt'
    ops.logFile(logfilename, '-noEcho')

    nodes_ = nodes
    elements_ = elements
    if len(nodes) == 0:
        nodes_ = ops.getNodeTags()

    if len(elements) == 0:
        elements_ = ops.getEleTags()

    elements_sec = elements_section
    if len(elements_section) == 0:
        elements_sec = ops.getEleTags()

    dofs = list(range(1, len(ops.nodeMass(nodes_[0])) + 1))
    otputs = {}

    resp_node_ID = {'Disp': 1, 'Vel': 2, 'Accel': 3, 'IncrDisp': 4, 'IncrDeltaDisp': 5, 'Reaction': 6, 'Unbalance': 7,
                    'RayleighForces': 8}

    cols_nodes = ['step', 'loc']
    cols_nodes.extend([str(x) for x in nodes_])

    cols_ele = ['step', 'loc']
    cols_ele.extend([str(x) for x in elements_])

    otputs['Disp'] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_nodes:
        if re_n != 'Disp':
            otputs[re_n] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_elements:
        otputs[re_n] = pd.DataFrame(columns=cols_ele)

    ###### Section
    cols_stressstrain = ['step', 'loc', 'eletag', 'section', 'y', 'z', 'strain', 'stress']
    cols_force = ['step', 'loc', 'eletag', 'section', 'axial strain', 'axial force', 'curvature', 'moment']
    otputs_stressstrain = pd.DataFrame(columns=cols_stressstrain)
    otputs_force = pd.DataFrame(columns=cols_force)
    secs_force = []
    if len(resp_sections) != 0:
        secs_force = list(dict.fromkeys(np.array(resp_sections)[:, 0]))

    start_time = datetime.now().replace(microsecond=0)
    print('Start Time: {}'.format(start_time))
    print("------------")

    step = []
    disp = []
    with open(filename_protocole) as f:
        lines = f.readlines()
    for line in lines:
        line2 = line.split("\t")

        step.append(float(line2[0]))
        disp.append(float(line2[1]))

    f.close()

    if disp[0] == 0.0 and step[0] == 0:
        disp.pop(0)

    TargetDisp = [x * Dy for x in disp]
    print('# Analysis Option:')
    ops.wipeAnalysis()
    print('ops.wipeAnalysis()')
    for key, vals in analysis_option.items():
        str_opt = 'ops.' + key + '('
        for i in range(len(vals)-1):
            val = vals[i]
            if isinstance(val, str):
                str_opt = str_opt + "\'" + val + "\'" + ', '
            else:
                str_opt = str_opt + str(val) + ', '
        val = vals[-1]
        if isinstance(val, str):
            str_opt = str_opt + "\'" + val + "\'" + ')'
        else:
            str_opt = str_opt + str(val) + ')'

        print(str_opt)
        eval(str_opt)
    try:
        print("------------")
        print('# Start Analysis: ')
        node_location = 0
        Nstep = 1
        savedstep = 1
        _count = 0
        _count_force = 0
        for i in range(len(TargetDisp)):
            print('Step ' + str(i + 1) + '/' + str(len(TargetDisp)) + ':')
            if TargetDisp[i] > node_location:
                du = du_max
            else:
                du = du_max * -1

            if abs(TargetDisp[i] - node_location) < abs(du):
                du = TargetDisp[i] - node_location

            print('    Try du = ', str(du))
            ops.integrator('DisplacementControl', cnodeTag, cdof, du, numIter)
            ops.analysis('Static')

            num_suc = 0
            while round(abs(node_location - TargetDisp[i]), 7) > 0:
                du_end = TargetDisp[i] - node_location
                if abs(du_end) < abs(du):
                    du = du_end
                    print('    Try du = ', str(du))
                    ops.integrator('DisplacementControl', cnodeTag, cdof, du, numIter)
                    num_suc = 0

                if num_suc == 10:
                    if abs(du * duDiv) <= du_max:
                        du = du * duDiv
                        if abs(du_end) < abs(du):
                            du = du_end

                        ops.integrator('DisplacementControl', cnodeTag, cdof, du, numIter)
                        print('    Try du = ', str(du))
                        num_suc = 0

                ok = ops.analyze(1)
                if ok != 0:
                    num_suc = 0
                    print('    Analysis failed at step ', str(Nstep))
                    du = du / duDiv
                    if abs(du) < du_min:
                        print('  Analysis failed: du < dumin =  ', str (du_min), '     Loc = ', str(round(node_location, 7))
                              , '    Target = ', str(TargetDisp[i]))

                        if exitrun.lower() in ['y', 'yes']:
                            exit()
                        else:
                            return()

                    print('    Try du = ', str(du))
                    ops.integrator('DisplacementControl', cnodeTag, cdof, du, numIter)
                else:
                    if (Nstep == savedstep) or (node_location + du == TargetDisp[-1]):
                        for n_sec in elements_sec:
                            for sec_props in resp_sections:
                                sec, y, z, mattag = sec_props
                                stress, strain = ops.eleResponse(n_sec, 'section', str(sec), 'fiber', str(y),
                                                                 str(z), str(mattag), 'stressStrain')

                                data_ele_st = [Nstep, node_location + du, n_sec, sec, y, z, strain, stress]
                                otputs_stressstrain.loc[_count] = data_ele_st
                                _count += 1

                            for sec in secs_force:
                                Fx, Mx = ops.eleResponse(n_sec, 'section', str(sec), 'force')
                                axialstrain, curvature = ops.eleResponse(n_sec, 'section', str(sec), 'deformation')
                                data_ele_force = [Nstep, node_location + du, n_sec, sec, axialstrain, Fx, curvature, Mx]
                                otputs_force.loc[_count_force] = data_ele_force
                                _count_force += 1

                        for re_n in resp_elements:
                            data_ele = [Nstep, node_location + du]
                            data_ele.extend([ops.eleResponse(n, re_n) for n in elements_])
                            otputs[re_n].loc[Nstep] = data_ele

                        for re_n in resp_nodes:
                            resp_s = []
                            for n in nodes_:
                                resp = []
                                for dof in dofs:
                                    if re_n == 'Reaction':
                                        ops.reactions()

                                    resp.append(ops.nodeResponse(n, dof, resp_node_ID[re_n]))

                                resp_s.append(resp)

                            data_node = [Nstep, node_location + du]
                            data_node.extend(resp_s)
                            otputs[re_n].loc[Nstep] = data_node

                        savedstep += everynstep
                    node_location += du
                    print('    Analysis successful at step ', str(Nstep), '     Loc = ', str(round(node_location, 7)),
                              '    Target = ', str(TargetDisp[i]))
                    Nstep += 1
                    num_suc += 1

        print('Analysis successful')

    except KeyboardInterrupt:
        print("Analysis terminated by user")
        exit()

    finally:
        print("------------")
        end_time = datetime.now().replace(microsecond=0)
        print('End Time: {}'.format(end_time))
        print('Duration: {}'.format(end_time - start_time))
        print("------------")

        print('Saving Data ...')
        if os.path.exists(_name + "\\output_" + name + "\\"):
            shutil.rmtree(_name + "\\output_" + name + "\\")
        os.mkdir(_name + "\\output_" + name + "\\")

        savedfiles = []
        for ou in otputs.keys():
            file_re = _name + "\\output_" + name + "\\" + ou + ".feather"
            db_output = otputs[ou].reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

        if len(resp_sections) != 0 and len(elements_section) != 0:
            file_re = _name + "\\output_" + name + "\\section_stress" + ".feather"
            db_output = otputs_stressstrain.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

            file_re = _name + "\\output_" + name + "\\section_force" + ".feather"
            db_output = otputs_force.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

        for sfile in savedfiles:
            print('Data was stored in ' + sfile)

        print("     ")
        print("Log file: " + logfilename)


def analyze_static(ops, name_project, name_analysis, analysis_option, num_steps=10, resp_nodes=['Disp', 'Reaction'],
                   resp_elements=['force', 'stresses', 'strains'], resp_sections=[], nodes=[], elements=[],
                   elements_section=[], everynstep=1, loadConst='yes', time=0.0, exitrun='y'):

    """
    Function to perform static analysis.

    ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
    name_project: Project name
    name_analysis: Analysis name
    analysis_option: A python dictionary contains analysis options.
         Example:
         analysis_option = {'constraints': ['Plain'],
                            'numberer': ['Plain'],
                            'system': ['BandGeneral'],
                            'test': ['NormUnbalance', 1.0e-5, 1000],
                            'algorithm': ['NewtonLineSearch']}
    num_steps: Number of analysis steps to perform(Optional, default value is 10).
    resp_nodes : A python list containing the type of node responses that are stored (optional, default value is ['Disp', 'Reaction'])
    resp_elements: A python list containing the type of element responses that are stored (optional, default value is ['force', 'stresses', 'strains'])
    resp_sections: See the notes below (optional, default value is an empty list)
    nodes: A python list containing tags of the nodes whose responses are stored (optional, default value is an empty list)
    elements: A python list containing tags of the elements whose responses are stored (optional, default value is an empty list)
    elements_section: A python list containing tags of the elements whose section responses are stored (optional, default value is an empty list)
    everynstep: The frequency of outputs that are stored. (Optional, default value is every 1 increments)
    loadConst: Is used to set the loads constant in the domain. Valid values are 'yes', 'y', 'no' or 'n'.
        (Optional, default value is 'yes').
    time: Time domain is to be set to(Optional, default value is 0.0).
    exitrun: whether to exit execution when encounters an unconvergance issue or terminates the current analysis and continue.
             Valid values are 'yes', 'y', 'no' or 'n' (optional, default value is 'y').

    Note:
        - The program stores the data after the analysis is complete successfully or terminates by unconvergence issues.
        - Valid values for resp_nodes are:
            Disp
            Vel
            Accel
            IncrDisp
            IncrDeltaDisp
            Reaction
            Unbalance
            RayleighForces
        - resp_sections is a 2D array of Nx4 size. each row contains section number, fiber y coordinate, fiber z coordinate and fiber material tag.
          Example:
              resp_sections =[[num_section1, y1, z1, matTag1], [num_section2, y2, z2, matTag2], ...]
              resp_sections = [[1, 0.0, 0.0, 1], [5, 0.3, 0.3, 1]]
        - elements_section list must contains beam-column element tags.
        - Storage files are saved in name_project\output_name_analysis\ respname.feather  path.
            Example: for name_project = myproject, name_analysis = ntha, Storage file pathes would be
                     myproject\output_ntha\Disp.feather
                     myproject\output_ntha\Reaction.feather
                     myproject\output_ntha\force.feather
                     ...

    """

    print('##########################')
    print('### Static Analysis[', name_analysis, ']')
    print('##########################')

    _name = name_project
    name = name_analysis

    if not os.path.exists(_name + "\\"):
        os.mkdir(_name + "\\")

    logfilename = _name + '\\opslogfile_' + name + '.txt'
    ops.logFile(logfilename, '-noEcho')

    nodes_ = nodes
    elements_ = elements

    if len(nodes) == 0:
        nodes_ = ops.getNodeTags()

    if len(elements) == 0:
        elements_ = ops.getEleTags()

    elements_sec = elements_section
    if len(elements_section) == 0:
        elements_sec = ops.getEleTags()

    dofs = list(range(1, len(ops.nodeMass(nodes_[0])) + 1))
    otputs = {}

    resp_node_ID = {'Disp': 1, 'Vel': 2, 'Accel': 3, 'IncrDisp': 4, 'IncrDeltaDisp': 5, 'Reaction': 6, 'Unbalance': 7,
                    'RayleighForces': 8}

    cols_nodes = ['step', 'time']
    cols_nodes.extend([str(x) for x in nodes_])

    cols_ele = ['step', 'time']
    cols_ele.extend([str(x) for x in elements_])

    otputs['Disp'] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_nodes:
        if re_n != 'Disp':
            otputs[re_n] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_elements:
        otputs[re_n] = pd.DataFrame(columns=cols_ele)

    ###### Section
    cols_stressstrain = ['step', 'time', 'eletag', 'section', 'y', 'z', 'strain', 'stress']
    cols_force = ['step', 'time', 'eletag', 'section', 'axial strain', 'axial force', 'curvature', 'moment']
    otputs_stressstrain = pd.DataFrame(columns=cols_stressstrain)
    otputs_force = pd.DataFrame(columns=cols_force)
    secs_force = []
    if len(resp_sections) != 0:
        secs_force = list(dict.fromkeys(np.array(resp_sections)[:, 0]))

    secs = []

    ##
    start_time = datetime.now().replace(microsecond=0)
    print('Start Time: {}'.format(start_time))
    print("------------")

    ops.record()

    print('# Analysis Option:')
    ops.wipeAnalysis()
    print('ops.wipeAnalysis()')
    for key, vals in analysis_option.items():
        str_opt = 'ops.' + key + '('
        for i in range(len(vals) - 1):
            val = vals[i]
            if isinstance(val, str):
                str_opt = str_opt + "\'" + val + "\'" + ', '
            else:
                str_opt = str_opt + str(val) + ', '
        val = vals[-1]
        if isinstance(val, str):
            str_opt = str_opt + "\'" + val + "\'" + ')'
        else:
            str_opt = str_opt + str(val) + ')'

        print(str_opt)
        eval(str_opt)
    print("ops.integrator(\'LoadControl\', " + str(1 / num_steps) + ")")
    print("ops.analysis(\'Static\')")

    ops.integrator('LoadControl', 1 / num_steps)
    ops.analysis('Static')

    try:
        print("------------")
        print('# Start Analysis: ')
        savedstep = 1
        _count = 0
        _count_force = 0
        num_steps = int(num_steps)
        for step in range(num_steps):
            ok = ops.analyze(1)
            if ok != 0:
                print('    Analysis failed at step ' + str(step + 1) + '/' + str(num_steps))

                if exitrun.lower() in ['y', 'yes']:
                    exit()
                else:
                    return ()

            else:
                if step + 1 == savedstep or (step + 1 == num_steps):
                    for n_sec in elements_sec:
                        for sec_props in resp_sections:
                            sec, y, z, mattag = sec_props
                            stress, strain = ops.eleResponse(n_sec, 'section', str(sec), 'fiber', str(y),
                                                             str(z), str(mattag), 'stressStrain')

                            data_ele_st = [step + 1, step + 1, n_sec, sec, y, z, strain, stress]
                            otputs_stressstrain.loc[_count] = data_ele_st
                            _count += 1

                        for sec in secs_force:
                            Fx, Mx = ops.eleResponse(n_sec, 'section', str(sec), 'force')
                            axialstrain, curvature = ops.eleResponse(n_sec, 'section', str(sec), 'deformation')
                            data_ele_force = [step + 1, step + 1, n_sec, sec, axialstrain, Fx, curvature, Mx]
                            otputs_force.loc[_count_force] = data_ele_force
                            _count_force += 1

                    for re_n in resp_elements:
                        data_ele = [step + 1, step + 1]
                        data_ele.extend([ops.eleResponse(n, re_n) for n in elements_])
                        otputs[re_n].loc[step + 1] = data_ele

                    for re_n in resp_nodes:
                        resp_s = []
                        for n in nodes_:
                            resp = []
                            for dof in dofs:
                                if re_n == 'Reaction':
                                    ops.reactions()

                                resp.append(ops.nodeResponse(n, dof, resp_node_ID[re_n]))

                            resp_s.append(resp)

                        data_node = [step + 1, step + 1]
                        data_node.extend(resp_s)
                        otputs[re_n].loc[step + 1] = data_node

                    savedstep += everynstep

            print('    Analysis successful at step ' + str(step + 1) + '/' + str(num_steps))

        if loadConst.lower() in ['y', 'yes']:
             ops.loadConst('-time', time)

        ops.remove('recorders')

        print('Analysis successful')

    except KeyboardInterrupt:
        print("Analysis terminated by user")
        exit()

    finally:
        print("------------")
        end_time = datetime.now().replace(microsecond=0)
        print('End Time: {}'.format(end_time))
        print('Duration: {}'.format(end_time - start_time))
        print("------------")

        print('Saving Data ...')
        if os.path.exists(_name + "\\output_" + name + "\\"):
            shutil.rmtree(_name + "\\output_" + name + "\\")
        os.mkdir(_name + "\\output_" + name + "\\")

        savedfiles = []
        for ou in otputs.keys():
            file_re = _name + "\\output_" + name + "\\" + ou + ".feather"
            db_output = otputs[ou].reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

        if len(resp_sections) != 0 and len(elements_section) != 0:
            file_re = _name + "\\output_" + name + "\\section_stress" + ".feather"
            db_output = otputs_stressstrain.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

            file_re = _name + "\\output_" + name + "\\section_force" + ".feather"
            db_output = otputs_force.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

        for sfile in savedfiles:
            print('Data was stored in ' + sfile)
        print("     ")
        print("Log file: " + logfilename)


def analyze_push_mono(ops, name_project, name_analysis, analysis_option, TargetDisp, cnodeTag, cdof, du_min, du_max,
                      duDiv=2, resp_nodes=['Disp', 'Reaction'], resp_elements=['force', 'stresses', 'strains'],
                      resp_sections=[], nodes=[], elements=[], elements_section=[], everynstep=1,  numIter=10,
                      exitrun='y'):

    """
    Function to perform pushover monotonic analysis.

    ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
    name_project: Project name
    name_analysis: Analysis name
    analysis_option: A python dictionary contains analysis options.
        Example:
        analysis_option = {'constraints': ['Plain'],
                           'numberer': ['Plain'],
                           'system': ['BandGeneral'],
                           'test': ['NormUnbalance', 1.0e-5, 1000],
                           'algorithm': ['NewtonLineSearch']}
    TargetDisp: pushover analysis is carried out on the structure until the displacement of the cnode equals to
        the TargetDisp(target displacement).
    cnodeTag: node whose response controls solution.
    cdof: degree of freedom at the node.
    du_min: the min stepsize the user will allow.
    du_max: the max stepsize the user will allow.
    du_Div: Refer to "How it works?"
    resp_nodes : A python list containing the type of node responses that are stored (optional, default value is ['Disp', 'Reaction'])
    resp_elements: A python list containing the type of element responses that are stored (optional, default value is ['force', 'stresses', 'strains'])
    resp_sections: See the notes below (optional, default value is an empty list)
    nodes: A python list containing tags of the nodes whose responses are stored (optional, default value is an empty list)
    elements: A python list containing tags of the elements whose responses are stored (optional, default value is an empty list)
    elements_section: A python list containing tags of the elements whose section responses are stored (optional, default value is an empty list)
    everynstep: The frequency of outputs that are stored. (Optional, default value is every 1 increments)
    numIter: the number of iterations the user would like to occur in the solution algorithm(Optional, default value is 10).
    exitrun: whether to exit execution when encounters an unconvergance issue or terminates the current analysis and continue.
         Valid values are 'yes', 'y', 'no' or 'n' (optional, default value is 'y').

    How it works?
    - Analysis is started with du = du_max.
    - If the analysis does not converge at a certain step, du is reduced by du_Div times(du = du / du_Div).
         This continues until the analysis converges at that step or du becomes smaller than du_min in which case the
         analysis is terminated.
    - After 10 successful steps du is increased by du_Div times(du = du * du_Div).
      This continues until du becomes greater that du_max in which case du will be set to du_max.
    - At the end, program adjusts du so that the location corresponds to the target displacement.

    Note:
        - The program saves the data after the analysis is complete successfully or terminates by unconvergence issue.
        - Valid values for resp_nodes are:
            Disp
            Vel
            Accel
            IncrDisp
            IncrDeltaDisp
            Reaction
            Unbalance
            RayleighForces
        - resp_sections is a 2D array of Nx4 size. each row contains section number, fiber y coordinate, fiber z coordinate and fiber material tag.
          Example:
              resp_sections =[[num_section1, y1, z1, matTag1], [num_section2, y2, z2, matTag2], ...]
              resp_sections = [[1, 0.0, 0.0, 1], [5, 0.3, 0.3, 1]]
        - elements_section list must contains beam-column element tags.
        -Storage files are saved in name_project\output_name_analysis\ respname.feather  path.
            Example: for name_project = myproject, name_analysis = ntha, Storage file pathes would be
                     myproject\output_ntha\Disp.feather
                     myproject\output_ntha\Reaction.feather
                     myproject\output_ntha\force.feather
                     ...
    """

    print('####################################')
    print('### Push Over Monotonic Analysis[', name_analysis, ']')
    print('####################################')

    _name = name_project
    name = name_analysis

    if not os.path.exists(_name + "\\"):
        os.mkdir(_name + "\\")

    logfilename = _name + '\\opslogfile_' + name + '.txt'
    ops.logFile(logfilename, '-noEcho')

    nodes_ = nodes
    elements_ = elements
    if len(nodes) == 0:
        nodes_ = ops.getNodeTags()

    if len(elements) == 0:
        elements_ = ops.getEleTags()

    elements_sec = elements_section
    if len(elements_section) == 0:
        elements_sec = ops.getEleTags()

    dofs = list(range(1, len(ops.nodeMass(nodes_[0])) + 1))
    otputs = {}

    resp_node_ID = {'Disp': 1, 'Vel': 2, 'Accel': 3, 'IncrDisp': 4, 'IncrDeltaDisp': 5, 'Reaction': 6, 'Unbalance': 7,
                    'RayleighForces': 8}

    cols_nodes = ['step', 'loc']
    cols_nodes.extend([str(x) for x in nodes_])

    cols_ele = ['step', 'loc']
    cols_ele.extend([str(x) for x in elements_])

    otputs['Disp'] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_nodes:
        if re_n != 'Disp':
            otputs[re_n] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_elements:
        otputs[re_n] = pd.DataFrame(columns=cols_ele)

    ###### Section
    cols_stressstrain = ['step', 'loc', 'eletag', 'section', 'y', 'z', 'strain', 'stress']
    cols_force = ['step', 'loc', 'eletag', 'section', 'axial strain', 'axial force', 'curvature', 'moment']
    otputs_stressstrain = pd.DataFrame(columns=cols_stressstrain)
    otputs_force = pd.DataFrame(columns=cols_force)
    secs_force = []
    if len(resp_sections) != 0:
        secs_force = list(dict.fromkeys(np.array(resp_sections)[:, 0]))

    start_time = datetime.now().replace(microsecond=0)
    print('Start Time: {}'.format(start_time))
    print("------------")

    print('# Analysis Option:')
    ops.wipeAnalysis()
    print('ops.wipeAnalysis()')
    for key, vals in analysis_option.items():
        str_opt = 'ops.' + key + '('
        for i in range(len(vals) - 1):
            val = vals[i]
            if isinstance(val, str):
                str_opt = str_opt + "\'" + val + "\'" + ', '
            else:
                str_opt = str_opt + str(val) + ', '
        val = vals[-1]
        if isinstance(val, str):
            str_opt = str_opt + "\'" + val + "\'" + ')'
        else:
            str_opt = str_opt + str(val) + ')'

        print(str_opt)
        eval(str_opt)

    try:
        print("------------")
        print('# Start Analysis: ')

        Nstep = 1
        savedstep = 1
        _count = 0
        _count_force = 0
        node_location = 0

        if TargetDisp > node_location:
            du = du_max
        else:
            du = du_max * -1
        print('    du = ', str(du))

        ops.integrator('DisplacementControl', cnodeTag, cdof, du, numIter)
        ops.analysis('Static')

        num_suc = 0
        while round(abs(node_location - TargetDisp), 7) > 0:
            if abs(TargetDisp - node_location) < abs(du):
                du = TargetDisp - node_location
                print('    Try du = ', str(du))
                ops.integrator('DisplacementControl', cnodeTag, cdof, du, numIter)
                num_suc = 0

            if num_suc == 10:
                if abs(du * duDiv) <= du_max:
                    du = du * duDiv
                    ops.integrator('DisplacementControl', cnodeTag, cdof, du, numIter)
                    print('    Try du = ', str(du))
                    num_suc = 0

            ok = ops.analyze(1)
            if ok != 0:
                num_suc = 0
                print('    Analysis failed at step ', str(Nstep))
                du = du / duDiv
                if abs(du) < du_min:
                    print('  Analysis failed: du < dumin ', '     Disp = ', str(round(node_location, 7)),
                          '    Target = ', str(TargetDisp))

                    if exitrun.lower() in ['y', 'yes']:
                        exit()
                    else:
                        return ()

                print('    Try du = ', str(du))
                ops.integrator('DisplacementControl', cnodeTag, cdof, du, numIter)
            else:
                if (Nstep == savedstep) or (node_location + du == TargetDisp):
                    for n_sec in elements_sec:
                        for sec_props in resp_sections:
                            sec, y, z, mattag = sec_props
                            stress, strain = ops.eleResponse(n_sec, 'section', str(sec), 'fiber', str(y),
                                                             str(z), str(mattag), 'stressStrain')

                            data_ele_st = [Nstep, node_location + du, n_sec, sec, y, z, strain, stress]
                            otputs_stressstrain.loc[_count] = data_ele_st
                            _count += 1

                        for sec in secs_force:
                            Fx, Mx = ops.eleResponse(n_sec, 'section', str(sec), 'force')
                            axialstrain, curvature = ops.eleResponse(n_sec, 'section', str(sec), 'deformation')
                            data_ele_force = [Nstep, node_location + du, n_sec, sec, axialstrain, Fx, curvature, Mx]
                            otputs_force.loc[_count_force] = data_ele_force
                            _count_force += 1

                    for re_n in resp_elements:
                        data_ele = [Nstep, node_location + du]
                        data_ele.extend([ops.eleResponse(n, re_n) for n in elements_])
                        otputs[re_n].loc[Nstep] = data_ele

                    for re_n in resp_nodes:
                        resp_s = []
                        for n in nodes_:
                            resp = []
                            for dof in dofs:
                                if re_n == 'Reaction':
                                    ops.reactions()

                                resp.append(ops.nodeResponse(n, dof, resp_node_ID[re_n]))

                            resp_s.append(resp)

                        data_node = [Nstep, node_location + du]
                        data_node.extend(resp_s)
                        otputs[re_n].loc[Nstep] = data_node

                    savedstep += everynstep

                node_location += du
                print('    Analysis successful at step ', str(Nstep), '     Disp = ',
                      str(round(node_location, 7)), '    Target = ', str(TargetDisp))
                Nstep += 1
                num_suc += 1

        print('Analysis successful')

    except KeyboardInterrupt:
        print("Analysis terminated by user")
        exit()

    finally:
        print("------------")
        end_time = datetime.now().replace(microsecond=0)
        print('End Time: {}'.format(end_time))
        print('Duration: {}'.format(end_time - start_time))
        print("------------")
        print('Saving Data ...')
        if os.path.exists(_name + "\\output_" + name + "\\"):
            shutil.rmtree(_name + "\\output_" + name + "\\")
        os.mkdir(_name + "\\output_" + name + "\\")

        savedfiles = []
        for ou in otputs.keys():
            file_re = _name + "\\output_" + name + "\\" + ou + ".feather"
            db_output = otputs[ou].reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

        if len(resp_sections) != 0 and len(elements_section) != 0:
            file_re = _name + "\\output_" + name + "\\section_stress" + ".feather"
            db_output = otputs_stressstrain.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

            file_re = _name + "\\output_" + name + "\\section_force" + ".feather"
            db_output = otputs_force.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)


        for sfile in savedfiles:
            print('Data was stored in ' + sfile)
        print("     ")
        print("Log file: " + logfilename)


def analyze_multiplesupport(ops, name_project, name_analysis, analysis_option, grmotions,
                            tag_pattern, tag_timeseries, tag_groundmotion, integration='Trapezoidal',
                            time_analysis=0.0, dt_min=0.0, dt_max=0.0, dt_Div=2, resp_nodes=['Disp', 'Reaction'],
                            resp_elements=['force', 'stresses', 'strains'], resp_sections=[], nodes=[], elements=[],
                            elements_section=[], everynstep=1, exitrun='n'):

    """

        Function to perform multiple support analysis.

        ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
        name_project: Project name
        name_analysis: Analysis name
        analysis_option: A python dictionary contains analysis options.
            Example:
            analysis_option = {'constraints': ['Plain'],
                               'numberer': ['RCM'],
                               'system': ['BandGeneral'],
                               'test': ['NormDispIncr', 1.0e-8, 100],
                               'algorithm': ['Newton'],
                               'integrator': ['Newmark',  0.5,  0.25]}
        grmotions:A python dictionary contains ground motion properties.
            key-value pairs are an integer value and a python list containing:
                 filename_record: A two-column data file including times and values.
                 scfactor: A factor to multiply load factors by.
                 type: Type of the history. Valid values are '-disp', '-vel' or '-accel'.
                 direction: Dof of enforced response. Valid range is from 1 through ndf at node.
                 node: Tag of node on which constraint is to be placed.
            Example:
            grmotions = {
                         1: [filename_record_1, scfactor_1, type_1, direction_1, node_1],
                         2: [filename_record_2, scfactor_2, type_2, direction_2, node_2],
                         3: [filename_record_3, scfactor_3, type_3, direction_3, node_3],
                         4: [filename_record_4, scfactor_4, type_4, direction_4, node_4],
                         5: [filename_record_5, scfactor_5, type_5, direction_5, node_5],
                         ....
                         n: [filename_record_n, scfactor_n, type_n, direction_n, node_n]
                         }
        tag_pattern: Program creates a pattern object of type MultipleSupport with tag_pattern as its unique tag.
        tag_timeseries: Program creates n TimeSeries objects of type path with tag_timeseries + i as their unique tags(i = 1 through n).
        tag_groundmotion: Program creates n GroundMotion objects with tag_groundmotion + i as their unique tags(i = 1 through n).
        integration: numerical integration method. Valid values are 'Trapezoidal' or 'Simpson' (Optional, default value is 'Trapezoidal').
        time_analysis: analysis time(Optional, default value is 0.0).
        dt_min: minimum time steps (Optional, default value is 0.0).
        dt_max: maximum time steps(Optional, default value is 0.0).
        dt_Div: Refer to "How it works?"
        resp_nodes : A python list containing the type of node responses that are stored (optional, default value is ['Disp', 'Reaction'])
        resp_elements: A python list containing the type of element responses that are stored (optional, default value is ['force', 'stresses', 'strains'])
        resp_sections: See the notes below (optional, default value is an empty list)
        nodes: A python list containing tags of the nodes whose responses are stored (optional, default value is an empty list)
        elements: A python list containing tags of the elements whose responses are stored (optional, default value is an empty list)
        elements_section: A python list containing tags of the elements whose section responses are stored (optional, default value is an empty list)
        everynstep: The frequency of outputs that are stored. (Optional, default value is every 1 increments)

        exitrun: whether to exit execution when encounters an unconvergance issue or terminates the current analysis and continue.
            Valid values are 'yes', 'y', 'no' or 'n' (optional, default value is 'n').

        How it works?
        - Program calculates dt using data provided by user.
        - If dt_max=0 or dt_max > dt, program will set dt_max to dt.
        - If dt_min=0 or dt_min > dt  program will set dt_min to dt_max / 10.

        - Analysis is started with dt = dt_max.
        - If the analysis does not converge at a certain step, dt is reduced by dt_Div times(dt = dt / dt_Div).
             This continues until the analysis converges at that step or dt becomes smaller than dt_min in which case the
             analysis is terminated.
        - After 20 successful steps dt is increased by dt_Div times(dt = dt * dt_Div).
          This continues until dt becomes greater that dt_max in which case dt will be set to dt_max.

        Note:
            - if time_analysis = 0.0 (default value), program uses maximum ground motion time, among all records, as analysis time.
            - if dt_max = 0.0 (default value), program uses minimum time increment, among all records, as dt_max.
            - if dt_min = 0.0 (default value), program uses 1/10 of dt_max  as dt_min.
            - The program saves the data after the analysis is complete successfully or terminates by unconvergence issue.
            - Valid values for resp_nodes are:
                Disp
                Vel
                Accel
                IncrDisp
                IncrDeltaDisp
                Reaction
                Unbalance
                RayleighForces
            - resp_sections is a 2D array of Nx4 size. each row contains section number, fiber y coordinate, fiber z coordinate and fiber material tag.
              Example:
                  resp_sections =[[num_section1, y1, z1, matTag1], [num_section2, y2, z2, matTag2], ...]
                  resp_sections = [[1, 0.0, 0.0, 1], [5, 0.3, 0.3, 1]]
            - elements_section list must contains beam-column element tags.
            -Storage files are saved in name_project\output_name_analysis\ respname.feather path.
                Example: for name_project = myproject, name_analysis = ntha, Storage file pathes would be
                         myproject\output_ntha\Disp.feather
                         myproject\output_ntha\Reaction.feather
                         myproject\output_ntha\force.feather
                         ...
        """

    _name = name_project
    name = name_analysis

    otputs = {}
    # try:
    print('##########################')
    print('### Multiple Support Analysis[', name_analysis,']')
    print('##########################')

    if not os.path.exists(_name + "\\"):
        os.mkdir(_name + "\\")

    logfilename = _name + '\\opslogfile_' + name + '.txt'
    ops.logFile(logfilename, '-noEcho')

    nodes_ = nodes
    elements_ = elements
    if len(nodes) == 0:
        nodes_ = ops.getNodeTags()

    if len(elements) == 0:
        elements_ = ops.getEleTags()

    elements_sec = elements_section
    if len(elements_section) == 0:
        elements_sec = ops.getEleTags()

    dofs = list(range(1, len(ops.nodeMass(nodes_[0])) + 1))

    resp_node_ID = {'Disp': 1, 'Vel': 2, 'Accel': 3, 'IncrDisp': 4, 'IncrDeltaDisp': 5, 'Reaction': 6, 'Unbalance': 7,
                    'RayleighForces': 8}

    cols_nodes = ['step', 'time']
    cols_nodes.extend([str(x) for x in nodes_])

    cols_ele = ['step', 'time']
    cols_ele.extend([str(x) for x in elements_])

    otputs['Disp'] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_nodes:
        if re_n != 'Disp':
            otputs[re_n] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_elements:
        otputs[re_n] = pd.DataFrame(columns=cols_ele)

    ###### Section
    cols_stressstrain = ['step', 'time', 'eletag', 'section', 'y', 'z', 'strain', 'stress']
    cols_force = ['step', 'time', 'eletag', 'section', 'axial strain', 'axial force', 'curvature', 'moment']
    otputs_stressstrain = pd.DataFrame(columns=cols_stressstrain)
    otputs_force = pd.DataFrame(columns=cols_force)
    secs_force = []
    if len(resp_sections) != 0:
        secs_force = list(dict.fromkeys(np.array(resp_sections)[:, 0]))

    start_time = datetime.now().replace(microsecond=0)
    print('Start Time: {}'.format(start_time))
    print("------------")

    str_pattern = "ops.pattern(\'MultipleSupport\'" + ", " + str(tag_pattern) + ")"
    print(str_pattern)
    eval(str_pattern)

    dt_records = {}
    filenames_temp = []
    t_maxes = {}
    accelerations = {}
    for key, grmotion in grmotions.items():
        time = []
        acceleration = []
        filename_record, scfactor,  type_record, direction, tag_node = grmotion

        with open(filename_record) as f:
            lines = f.readlines()
        for line in lines:
            line2 = line.split("\t")
            time.append(float(line2[0]))
            acceleration.append(float(line2[1]))

        f.close()
        accelerations[key] = acceleration

        _dt = time[1] - time[0]
        dt_records[key] = _dt
        t_maxes[key] = time[-1]

    time_final = time_analysis
    if time_analysis <= 0.0:
        time_final = max(t_maxes.values())

    _dt = min(dt_records.values())
    dt_message = ''
    if dt_max <= 0:
        dt_message = '### dt_max <= 0.0, dt_max was set to dt = ' + str(_dt)
        dt_max = _dt

    if dt_max > _dt:
        dt_message = '### Warning:  dt_max = ' + str(dt_max) + '   dt = ' + str(_dt) + '  dt_max > dt'

    if dt_min <= 0:
        dt_message = '### dt_min <= 0.0, dt_min was set to dt_max / 10'
        dt_min = dt_max / 10

    if dt_min > _dt:
        dt_message = '### dt_min >  dt, dt_min was set to dt_max / 10'
        dt_min = dt_max / 10

    if dt_min > dt_max:
        dt_message = '### dt_min >  dt_max, dt_min was set to dt_max / 10'
        dt_min = dt_max / 10

    dt = dt_max

    tag_count = 0
    for key, grmotion in grmotions.items():
        filename_record, scfactor, type_record, direction, tag_node = grmotion
        filename_temp = '__tempacc_' + str(tag_count + 1) + '.txt'
        filenames_temp.append(filename_temp)
        if os.path.exists(filename_temp):
            os.remove(filename_temp)

        with open(filename_temp, 'w') as f:
            for acc in accelerations[key]:
                f.write(f"{acc}\n")

        # Set time series to be passed to uniform excitation
        str_timeSeries = "ops.timeSeries(\'Path\'," + ' ' + str(tag_timeseries + tag_count) + ", \'-filePath\' ,\'" + filename_temp + \
                         "', \'-dt\', " + str(_dt) + ", \'-factor\', " + str(scfactor) + ")"

        print(str_timeSeries)
        eval(str_timeSeries)

        str_groundMotion = "ops.groundMotion(" + str(tag_groundmotion + tag_count) + ', \'Plain\', \'' + type_record + '\', ' + \
                           str(tag_timeseries + tag_count) + ", \'-int\', \'" + integration + "\')"
        print(str_groundMotion)
        eval(str_groundMotion)

        str_imposedMotion = "ops.imposedMotion(" + str(tag_node) + ", " + str(direction) + ", " + str(tag_groundmotion + tag_count) + ")"

        print(str_imposedMotion)
        eval(str_imposedMotion)
        tag_count += 1

    try:
        print("------------")
        print('# Analysis Option:')
        ops.wipeAnalysis()
        print('ops.wipeAnalysis()')
        for key, vals in analysis_option.items():
            str_opt = 'ops.' + key + '('
            for i in range(len(vals)-1):
                val = vals[i]
                if isinstance(val, str):
                    str_opt = str_opt + "\'" + val + "\'" + ', '
                else:
                    str_opt = str_opt + str(val) + ', '
            val = vals[-1]
            if isinstance(val, str):
                str_opt = str_opt + "\'" + val + "\'" + ')'
            else:
                str_opt = str_opt + str(val) + ')'

            print(str_opt)
            eval(str_opt)
        print("------------")
        print('# Start Analysis: ')

        ops.analysis('Transient')

        Nstep = 1
        savedstep = 1
        _count = 0
        _count_force = 0
        time_cur = 0
        num_suc = 0
        print(dt_message)
        print('    dt = ', str(dt_max))
        while round(time_final - time_cur, 7) > 0:
            dt_end = time_final - time_cur
            if dt_end < dt:
                dt = dt_end
                print('    Try dt = ', str(dt))
                num_suc = 0

            if num_suc == 20:
                if dt * dt_Div <= dt_max:
                    dt = dt * dt_Div
                    if dt_end < dt:
                        dt = dt_end

                    print('    Try dt = ', str(dt))
                    num_suc = 0

            ok = ops.analyze(1, dt)
            if ok != 0:
                print('    Analysis failed at step ', str(Nstep), '   time = ', str(round(time_cur, 4)) , '/',
                      str(round(time_final, 4)), '   ', str(round(100 * time_cur/ time_final, 2)),'%')
                dt = dt / dt_Div
                if abs(dt) < dt_min:
                    print('  Analysis failed: dt < dtmin ', '   time = ', str(round(time_cur, 4)) , '/',
                      str(round(time_final, 4)), '   ', str(round(100 * time_cur/ time_final, 2)),'%')

                    if exitrun.lower() in ['y', 'yes']:
                        exit()
                    else:
                        return ()

                print('    Try dt = ', str(dt))
                num_suc = 0
            else:
                if (Nstep == savedstep) or (time_cur + dt == time_final):
                    for n_sec in elements_sec:
                        for sec_props in resp_sections:
                            sec, y, z, mattag = sec_props
                            stress, strain = ops.eleResponse(n_sec, 'section', str(sec), 'fiber', str(y),
                                                             str(z), str(mattag), 'stressStrain')

                            data_ele_st = [Nstep, time_cur + dt, n_sec, sec, y, z, strain, stress]
                            otputs_stressstrain.loc[_count] = data_ele_st
                            _count += 1

                        for sec in secs_force:
                            Fx, Mx = ops.eleResponse(n_sec, 'section', str(sec), 'force')
                            axialstrain, curvature = ops.eleResponse(n_sec, 'section', str(sec), 'deformation')
                            data_ele_force = [Nstep, time_cur + dt, n_sec, sec, axialstrain, Fx, curvature, Mx]
                            otputs_force.loc[_count_force] = data_ele_force
                            _count_force += 1

                    for re_n in resp_elements:
                        data_ele = [Nstep, time_cur + dt]
                        data_ele.extend([ops.eleResponse(n, re_n) for n in elements_])
                        otputs[re_n].loc[Nstep] = data_ele

                    for re_n in resp_nodes:
                        resp_s = []
                        for n in nodes_:
                            resp = []
                            for dof in dofs:
                                if re_n == 'Reaction':
                                    ops.reactions()

                                resp.append(ops.nodeResponse(n, dof, resp_node_ID[re_n]))

                            resp_s.append(resp)

                        data_node = [Nstep, time_cur + dt]
                        data_node.extend(resp_s)
                        otputs[re_n].loc[Nstep] = data_node

                    savedstep += everynstep


                time_cur += dt
                print('    Analysis successful at step ', str(Nstep), '   time = ', str(round(time_cur, 4)) , '/',
                      str(round(time_final, 4)), '   ', str(round(100 * time_cur/ time_final, 2)),'%')
                Nstep += 1
                num_suc += 1

        print('Analysis successful')

    except KeyboardInterrupt:
        print("Analysis terminated by user")
        exit()

    finally:
        print("------------")
        end_time = datetime.now().replace(microsecond=0)
        print('End Time: {}'.format(end_time))
        print('Duration: {}'.format(end_time - start_time))
        print("------------")

        print('Saving Data ...')
        if os.path.exists(_name + "\\output_" + name + "\\"):
            shutil.rmtree(_name + "\\output_" + name + "\\")
        os.mkdir(_name + "\\output_" + name + "\\")

        savedfiles = []
        for ou in otputs.keys():
            file_re = _name + "\\output_" + name + "\\" + ou + ".feather"
            db_output = otputs[ou].reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

        if len(resp_sections) != 0 and len(elements_section) != 0:
            file_re = _name + "\\output_" + name + "\\section_stress" + ".feather"
            db_output = otputs_stressstrain.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

            file_re = _name + "\\output_" + name + "\\section_force" + ".feather"
            db_output = otputs_force.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

        for sfile in savedfiles:
            print('Data was stored in ' + sfile)
        print("     ")
        print("Log file: " + logfilename)
        for filename_temp in filenames_temp:
            os.remove(filename_temp)


def analyze_transient(ops, name_project, name_analysis, analysis_option, filename_record, tag_timeseries, scfactor,
                      tag_pattern, direction=1, time_analysis=0.0, dt_min=0.0, dt_max=0.0, dt_Div=2, resp_nodes=['Disp', 'Reaction'],
                      resp_elements=['force', 'stresses', 'strains'], resp_sections=[], nodes=[], elements=[],
                   elements_section=[], everynstep=1, type='-accel', exitrun='n'):
    """
    Function to perform transient analysis.

    ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
    name_project: Project name
    name_analysis: Analysis name
    analysis_option: A python dictionary contains analysis options.
        Example:
        analysis_option = {'constraints': ['Plain'],
                           'numberer': ['RCM'],
                           'system': ['BandGeneral'],
                           'test': ['NormDispIncr', 1.0e-8, 100],
                           'algorithm': ['Newton'],
                           'integrator': ['Newmark',  0.5,  0.25]}
    filename_record: A two-column data file including times and values.
    tag_timeseries: Program creates a TimeSeries object of type path with tag_timeseries as its unique tag.
    tag_pattern: Program creates a pattern object of type UniformExcitation with tag_pattern as its unique tag.
    scfactor: A factor to multiply load factors by(Optional, default value is 1.0).
    direction: 	direction in which ground motion acts(Optional, default value is 1).
        1 - corresponds to translation along the global X axis
        2 - corresponds to translation along the global Y axis
        3 - corresponds to translation along the global Z axis
        4 - corresponds to rotation about the global X axis
        5 - corresponds to rotation about the global Y axis
        6 - corresponds to rotation about the global Z axis
    time_analysis: The analysis time(Optional, default value is 0.0).
    dt_min: minimum time steps (Optional, default value is 0.0).
    dt_max: maximum time steps(Optional, default value is 0.0).
    dt_Div: Refer to "How it works?"
    resp_nodes : A python list containing the type of node responses that are stored (optional, default value is ['Disp', 'Reaction'])
    resp_elements: A python list containing the type of element responses that are stored (optional, default value is ['force', 'stresses', 'strains'])
    resp_sections: See the notes below (optional, default value is an empty list)
    nodes: A python list containing tags of the nodes whose responses are stored (optional, default value is an empty list)
    elements: A python list containing tags of the elements whose responses are stored (optional, default value is an empty list)
    elements_section: A python list containing tags of the elements whose section responses are stored (optional, default value is an empty list)
    everynstep: The frequency of outputs that are stored. (Optional, default value is every 1 increments)
    type: Type of the history. Valid values are '-disp', '-vel' or '-accel' (Optional, default value is '-accel').
    exitrun: whether to exit execution when encounters an unconvergance issue or terminates the current analysis and continue.
        Valid values are 'yes', 'y', 'no' or 'n' (optional, default value is 'n').

    How it works?
    - Program calculates dt using data provided by user.
    - If dt_max=0 or dt_max > dt, program will set dt_max to dt.
    - If dt_min=0 or dt_min > dt  program will set dt_min to dt_max / 10.

    - Analysis is started with dt = dt_max.
    - If the analysis does not converge at a certain step, dt is reduced by dt_Div times(dt = dt / dt_Div).
         This continues until the analysis converges at that step or dt becomes smaller than dt_min in which case the
         analysis is terminated.
    - After 20 successful steps dt is increased by dt_Div times(dt = dt * dt_Div).
      This continues until dt becomes greater that dt_max in which case dt will be set to dt_max.

    Note:
        - if time_analysis = 0.0 (default value), program uses ground motion time as analysis time.
        - if dt_max = 0.0 (default value), program uses time increment of the record, as dt_max.
        - if dt_min = 0.0 (default value), program uses 1/10 of dt_max as dt_min.
        - The program saves the data after the analysis is complete successfully or terminates by unconvergence issue.
        - Valid values for resp_nodes are:
            Disp
            Vel
            Accel
            IncrDisp
            IncrDeltaDisp
            Reaction
            Unbalance
            RayleighForces
        - resp_sections is a 2D array of Nx4 size. each row contains section number, fiber y coordinate, fiber z coordinate and fiber material tag.
          Example:
              resp_sections =[[num_section1, y1, z1, matTag1], [num_section2, y2, z2, matTag2], ...]
              resp_sections = [[1, 0.0, 0.0, 1], [5, 0.3, 0.3, 1]]
        - elements_section list must contains beam-column element tags.
        -Storage files are saved in name_project\output_name_analysis\ respname.feather path.
            Example: for name_project = myproject, name_analysis = ntha, Storage file pathes would be
                     myproject\output_ntha\Disp.feather
                     myproject\output_ntha\Reaction.feather
                     myproject\output_ntha\force.feather
                     ...
    """
    _name = name_project
    name = name_analysis

    otputs = {}
    # try:
    print('##########################')
    print('### Transient Analysis[', name_analysis ,']')
    print('##########################')

    if not os.path.exists(_name + "\\"):
        os.mkdir(_name + "\\")

    logfilename = _name + '\\opslogfile_' + name + '.txt'
    ops.logFile(logfilename, '-noEcho')

    nodes_ = nodes
    elements_ = elements
    if len(nodes) == 0:
        nodes_ = ops.getNodeTags()

    if len(elements) == 0:
        elements_ = ops.getEleTags()

    elements_sec = elements_section
    if len(elements_section) == 0:
        elements_sec = ops.getEleTags()

    dofs = list(range(1, len(ops.nodeMass(nodes_[0])) + 1))

    resp_node_ID = {'Disp': 1, 'Vel': 2, 'Accel': 3, 'IncrDisp': 4, 'IncrDeltaDisp': 5, 'Reaction': 6, 'Unbalance': 7,
                    'RayleighForces': 8}

    cols_nodes = ['step', 'time']
    cols_nodes.extend([str(x) for x in nodes_])

    cols_ele = ['step', 'time']
    cols_ele.extend([str(x) for x in elements_])

    otputs['Disp'] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_nodes:
        if re_n != 'Disp':
            otputs[re_n] = pd.DataFrame(columns=cols_nodes)

    for re_n in resp_elements:
        otputs[re_n] = pd.DataFrame(columns=cols_ele)

    ###### Section
    cols_stressstrain = ['step', 'time', 'eletag', 'section', 'y', 'z', 'strain', 'stress']
    cols_force = ['step', 'time', 'eletag', 'section', 'axial strain', 'axial force', 'curvature', 'moment']
    otputs_stressstrain = pd.DataFrame(columns=cols_stressstrain)
    otputs_force = pd.DataFrame(columns=cols_force)
    secs_force = []
    if len(resp_sections) != 0:
        secs_force = list(dict.fromkeys(np.array(resp_sections)[:, 0]))

    start_time = datetime.now().replace(microsecond=0)
    print('Start Time: {}'.format(start_time))
    print("------------")

    time = []
    acceleration = []
    with open(filename_record) as f:
        lines = f.readlines()
    for line in lines:
        line2 = line.split("\t")
        time.append(float(line2[0]))
        acceleration.append(float(line2[1]))

    f.close()

    _dt = time[1] - time[0]

    filename_temp = '__tempacc.txt'
    if os.path.exists(filename_temp):
        os.remove(filename_temp)

    with open(filename_temp, 'w') as f:
        for acc in acceleration:
            f.write(f"{acc}\n")

    # Set time series to be passed to uniform excitation
    str_timeSeries = "ops.timeSeries(\'Path\'," + ' ' + str(tag_timeseries) + ", \'-filePath\' ,\'" + filename_temp + \
                     "', \'-dt\', " + str(_dt) + ", \'-factor\', " + str(scfactor) + ")"

    print(str_timeSeries)
    eval(str_timeSeries)
    # ops.timeSeries('Path', tag_timeseries, '-filePath', filename_temp, '-dt', _dt, '-factor', scfactor)

    # Create UniformExcitation load pattern
    #                         tag dir
    # ops.pattern('UniformExcitation',  tag_pattern,  direction,  type, tag_timeseries)
    str_pattern = "ops.pattern(\'UniformExcitation\', " + str(tag_pattern) + ', ' + str(direction) + ', \'' + type + \
                  '\', ' + str(tag_timeseries) + ')'
    print(str_pattern)
    eval(str_pattern)

    print('# Analysis Option:')
    ops.wipeAnalysis()
    print('ops.wipeAnalysis()')
    for key, vals in analysis_option.items():
        str_opt = 'ops.' + key + '('
        for i in range(len(vals)-1):
            val = vals[i]
            if isinstance(val, str):
                str_opt = str_opt + "\'" + val + "\'" + ', '
            else:
                str_opt = str_opt + str(val) + ', '
        val = vals[-1]
        if isinstance(val, str):
            str_opt = str_opt + "\'" + val + "\'" + ')'
        else:
            str_opt = str_opt + str(val) + ')'

        print(str_opt)
        eval(str_opt)

    try:
        print("------------")
        print('# Start Analysis: ')

        if dt_max <= 0:
            print('### dt_max <= 0.0, dt_max was set to dt = ' + str(_dt))
            dt_max = _dt

        if dt_max > _dt:
            print("Warning: ")
            print('### dt_max = ' + str(dt_max) + '   dt = ' + str(_dt) + '  dt_max > dt')

        if dt_min <= 0:
            print('### dt_min <= 0.0, dt_min was set to dt_max / 10')
            dt_min = dt_max / 10

        if dt_min > _dt:
            print('### dt_min >  dt, dt_min was set to dt_max / 10')
            dt_min = dt_max / 10

        if dt_min > dt_max:
            print('### dt_min >  dt_max, dt_min was set to dt_max / 10')
            dt_min = dt_max / 10

        print('    dt = ', str(dt_max))
        dt = dt_max
        ops.analysis('Transient')

        Nstep = 1
        savedstep = 1
        _count = 0
        _count_force = 0
        time_final = time_analysis
        if time_analysis <= 0.0:
            time_final = time[-1]
        time_cur = 0
        num_suc = 0

        while round(time_final - time_cur, 7) > 0:
            dt_end = time_final - time_cur
            if dt_end < dt:
                dt = dt_end
                print('    Try dt = ', str(dt))
                num_suc = 0

            if num_suc == 20:
                if dt * dt_Div <= dt_max:
                    dt = dt * dt_Div
                    if dt_end < dt:
                        dt = dt_end

                    print('    Try dt = ', str(dt))
                    num_suc = 0

            ok = ops.analyze(1, dt)
            if ok != 0:
                print('    Analysis failed at step ', str(Nstep), '   time = ', str(round(time_cur, 4)) , '/',
                      str(round(time_final, 4)), '   ', str(round(100 * time_cur/ time_final, 2)),'%')
                dt = dt / dt_Div
                if abs(dt) < dt_min:
                    print('  Analysis failed: dt < dtmin ', '   time = ', str(round(time_cur, 4)) , '/',
                      str(round(time_final, 4)), '   ', str(round(100 * time_cur/ time_final, 2)),'%')

                    if exitrun.lower() in ['y', 'yes']:
                        exit()
                    else:
                        return ()

                print('    Try dt = ', str(dt))
                num_suc = 0
            else:
                if (Nstep == savedstep) or (time_cur + dt == time_final):
                    for n_sec in elements_sec:
                        for sec_props in resp_sections:
                            sec, y, z, mattag = sec_props
                            stress, strain = ops.eleResponse(n_sec, 'section', str(sec), 'fiber', str(y),
                                                             str(z), str(mattag), 'stressStrain')

                            data_ele_st = [Nstep, time_cur + dt, n_sec, sec, y, z, strain, stress]
                            otputs_stressstrain.loc[_count] = data_ele_st
                            _count += 1

                        for sec in secs_force:
                            Fx, Mx = ops.eleResponse(n_sec, 'section', str(sec), 'force')
                            axialstrain, curvature = ops.eleResponse(n_sec, 'section', str(sec), 'deformation')
                            data_ele_force = [Nstep, time_cur + dt, n_sec, sec, axialstrain, Fx, curvature, Mx]
                            otputs_force.loc[_count_force] = data_ele_force
                            _count_force += 1

                    for re_n in resp_elements:
                        data_ele = [Nstep, time_cur + dt]
                        data_ele.extend([ops.eleResponse(n, re_n) for n in elements_])
                        otputs[re_n].loc[Nstep] = data_ele

                    for re_n in resp_nodes:
                        resp_s = []
                        for n in nodes_:
                            resp = []
                            for dof in dofs:
                                if re_n == 'Reaction':
                                    ops.reactions()

                                resp.append(ops.nodeResponse(n, dof, resp_node_ID[re_n]))

                            resp_s.append(resp)

                        data_node = [Nstep, time_cur + dt]
                        data_node.extend(resp_s)
                        otputs[re_n].loc[Nstep] = data_node

                    savedstep += everynstep

                time_cur += dt
                print('    Analysis successful at step ', str(Nstep), '   time = ', str(round(time_cur, 4)) , '/',
                      str(round(time_final, 4)), '   ', str(round(100 * time_cur/ time_final, 2)),'%')
                Nstep += 1
                num_suc += 1

        print('Analysis successful')


    except KeyboardInterrupt:
        print("Analysis terminated by user")
        exit()

    finally:
        print("------------")
        end_time = datetime.now().replace(microsecond=0)
        print('End Time: {}'.format(end_time))
        print('Duration: {}'.format(end_time - start_time))
        print("------------")

        print('Saving Data ...')
        if os.path.exists(_name + "\\output_" + name + "\\"):
            shutil.rmtree(_name + "\\output_" + name + "\\")
        os.mkdir(_name + "\\output_" + name + "\\")

        savedfiles = []
        for ou in otputs.keys():
            file_re = _name + "\\output_" + name + "\\" + ou + ".feather"
            db_output = otputs[ou].reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

        if len(resp_sections) != 0 and len(elements_section) != 0:
            file_re = _name + "\\output_" + name + "\\section_stress" + ".feather"
            db_output = otputs_stressstrain.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

            file_re = _name + "\\output_" + name + "\\section_force" + ".feather"
            db_output = otputs_force.reset_index()
            db_output.to_feather(file_re, compression='zstd')
            savedfiles.append(file_re)

        for sfile in savedfiles:
            print('Data was stored in ' + sfile)
        print("     ")
        print("Log file: " + logfilename)
        os.remove(filename_temp)


def damping(ops, name_project,  xDamp1=0.05, xDamp2=0.0, T1=0.0, T2=0.0, mode1=1, mode2=3, factor_betaK=0.0, factor_betaKinit=0.0, factor_betaKcomm=1.0,
            elements=[], nodes=[], xlim=[], ylim=[], zlim=[], solver='-genBandArpack'):

    """
    A function for applying damping to the structure. Program calculates mass and stiffness coefficients(alphaM and betaK)
    based on first and second periods. Rayleigh damping parameters are assigned to the nodes and elements in the region
    defined by elements and nodes or xlim, ylim and zlim.

    ops: openseespy object. Alias for this package should be ops(import openseespy.opensees as ops).
    xDamp1: Damping ratio for first mode.(Optional, default value is 0.05)
    xDamp2: Damping ratio for second mode.(Optional, default value is 0.0, xDamp2 = xDamp1)
    T1: Period at first mode (Optional, default value is 0. If T1 = 0, program will perform eigen analysis to
        determine T1).
    T2: Period at second mode (Optional, default value is 0. If T2 = 0, program will perform eigen analysis to
        determine T2).
    mode1 = the first mode number used to determine T1(Optional, default value is 1. If T1 = 0, program will perform eigen analysis
              and the period of the mode1 is considered for T1).
    mode2 = the first mode number used to determine T2(Optional, default value is 3. If T2 = 0, program will perform eigen analysis
              and the period of the mode2 is considered for T2).
    factor_betaK: Factor applied to elements current stiffness matrix = factor_betaK * betaK(Optional, default value is 0.0)
    factor_betaKinit: Factor applied to elements initial stiffness matrix = factor_betaKinit * betaK(Optional, default value is 0.0)
    factor_betaKcomm: Factor applied to elements committed stiffness matrix = factor_betaKcomm * betaK(Optional, default value is 1.0)
    elements: tags of selected elements in domain to be included in region.(Optional, default value is an empty list)
    nodes: tags of selected nodes in domain to be included in region.(Optional, default value is an empty list)
    xlim: An empty list or a list containing xmin and xmax of the region.(Optional, default value is an empty list)
    ylim: An empty list or a list containing ymin and ymax of the region.(Optional, default value is an empty list)
    zlim: An empty list or a list containing zmin and zmax of the region.(Optional, default value is an empty list)
    solver: String detailing type of solver: '-genBandArpack', '-fullGenLapack', (Optional, default value is '-genBandArpack')

    return: alphaM, betaK, lambda, omega, Tn

    Note:
        - If elements, nodes, xlim, ylim and zlim are empty lists program will use rayleigh command to apply damping.
          Otherwise region command will be used by the program.

        - If elements is an empty list program uses xlim, ylim and zlim to find elements in the region.
          If xlim, ylim and zlim are empty lists too, program will use all elements in the domain

        - If nodes is an empty list program uses xlim, ylim and zlim to find nodes in the region.
          If xlim, ylim and zlim are empty lists too, program will use all nodes in the domain

    """
    print('##########################')
    print('### Damping')
    print('##########################')
    logfilename = name_project + '\\opslogfile_Damping' + '.txt'
    ops.logFile(logfilename, '-noEcho')

    try:
        Tn = [T1, T2]
        if T1 == 0 or T2 == 0:
            lambdaN_, omega_, Tn_ = __eigen(ops, int(max([mode1, mode2])), solver=solver)
            if T1 <= 0:
                T1 = Tn_[mode1 - 1]
            if T2 <= 0:
                T2 = Tn_[mode2 - 1]

            print("T1 and T2 were set to: T1 = " + str(round(T1, 3)) + " sec[mode" + str(mode1) + "]      and      T2 = " +
                  str(round(T2, 3)) + " sec[mode" + str(mode2) + "]")

        omega1, omega2 = (2 * np.pi) / T1, (2 * np.pi) / T2
        lambdaN = [omega1 ** 2, omega2 ** 2]
        omega = [omega1, omega2]

        if xDamp2 <= 0.0:
            xDamp2 = xDamp1

        alphaM = (2 * (omega2*xDamp1 - omega1*xDamp2) * omega1 * omega2) / ((omega2 - omega1) * (omega2 + omega1))
        betaSt = (2 * (omega2*xDamp2 - omega1*xDamp1)) / ((omega2 - omega1) * (omega2 + omega1))

        if len(elements) == 0 and len(nodes) == 0 and len(xlim) == 0 and len(ylim) == 0 and len(zlim) == 0:
            ops.rayleigh(alphaM, factor_betaK * betaSt, factor_betaKinit * betaSt, factor_betaKcomm * betaSt)

        else:
            delements = elements
            if len(elements) == 0:
                delements = find_elements(ops, xlim, ylim, zlim)

            if delements is False:
                print('Warning: No elements have been found in the region for damping')
            else:
                ops.region(1, '-ele', *delements, '-rayleigh', 0.0, factor_betaK * betaSt,
                           factor_betaKinit * betaSt, factor_betaKcomm * betaSt)

            dnodes = nodes
            if len(nodes) == 0:
                dnodes = find_nodes(ops, xlim, ylim, zlim)

            if dnodes is False:
                print('Warning: No nodes have been found in the region for damping')
            else:
                ops.region(2, '-node', *dnodes, '-rayleigh', alphaM, 0.0, 0.0, 0.0)

        return alphaM, betaSt, lambdaN, omega, Tn
    finally:
        print("Log file: " + logfilename)
        print("    ")

def plot_protocol(ax, filename_protocol, Dy, linewidth=1.0, color='k', xlabel='Step', ylabel='Displacement', title='', grid=True):
    """
    Function to plot loading protocol used in pushover cyclic analysis.

    ax: The axes on which the protocol is drawn.
    filename_protocol: A two-column data file containing steps and displacements.
    Dy: Yield displacement.
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')
    xtitle: Label for the x-axis (optional, default value is 'Step')
    ytitle: Label for the y-axis (optional, default value is 'Displacement')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is True)
    """
    step = []
    disp = []
    with open(filename_protocol) as f:
        lines = f.readlines()
    for line in lines:
        line2 = line.split()

        step.append(float(line2[0]))
        disp.append(float(line2[1]))

    f.close()

    target = [x * Dy for x in disp]

    ax.plot(step, target, color=color, lw=linewidth)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.grid(grid)


def plot_record(ax, filename_record, scfactor=1.0, linewidth=0.5, color='k', xlabel='Time(s)', ylabel='Acceleration(g)', title='', grid=True):

    """
    Function to plot ground motion record used in transient analysis.

    ax: The axes on which the protocol is drawn.
    filename_record: A two-column data file including times and values.
    scfactor: All values are multiplied by scfactor (optional, default value is 1)
    linewidth: Line width of the graph plot (optional, default value is 0.5)
    color: Color of the graph plot (optional, default value is 'b')
    xlabel: Label for the x-axis (optional, default value is 'Time(s)')
    ylabel: Label for the y-axis (optional, default value is 'Acceleration(g)')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is True)
    """

    time = []
    acceleration = []
    with open(filename_record) as f:
        lines = f.readlines()
    for line in lines:
        line2 = line.split()

        time.append(float(line2[0]))
        acceleration.append(float(line2[1]) * scfactor)

    f.close()

    ax.plot(time, acceleration, color=color, lw=linewidth)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.grid(grid)


def retrieve_data(file_response, tags=[], steps=[], dofs=[], scfactors=[], sendtoexcel='y',  file_path='book1.xlsx'):
    '''
    Function to retrieve data from a saved response file.

    file_response: saved response file path
    tags: Tags of nodes or elements whose responses are retrieved (optional, default value is an empty list)
    steps: A python list containing the steps in which the responses are retrieved (optional, default value is an empty list)
    dofs: A python list containing the DOFs in which the responses are retrieved (optional, default value is an empty list)
    scfactors: A python list containing scale factors multiplied by the response at each DOF (optional, default value is an empty list)
    sendtoexcel: Whether to send the retrieved data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the retrieved data is stored (optional, default value is 'book1.xlsx')

    Notes:
        1- Number of items in dofs must be equal to number of items in scfactors.
        2- If tags is an empty list, responses for all nodes/elements will be retrieved
        3- If steps is an empty list, responses at all steps will be retrieved
        4- If dofs is an empty list, responses in all degrees of freedom will be retrieved
        5- If scfactors is an empty list, responses will be multiplied by 1.
        6- Data related to each degree of freedom are stored in separate sheets in excel file with name DOFn.

    return: A python list containing data for each DOF as a pandas DataFrame.
            The column labels of the DataFrames are ['step', 'time', tags as python string].
            Example: df.columns = ['step', 'time', '1', '2', ... 'n']

    Example:
                                   5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---
        file_Disp = "myproject\output_ntha\Disp.feather"
        df_s = opa.retrieve_data(file_Disp, tags=[1,3,5], steps=[], dofs=[1, 3], scfactors=[], sendtoexcel='y',
                         file_path='disps.xlsx')

         df_dof_1, df_dof_3 = df_s

         df_dof_1 :

         |'step'| 'time' |  '1' |  '3' |  '5'  |
         --------------------------------------
         |  0   |   0.0  | 0.0  | 0.0  |  0.0  |
         |  1   |   0.5  | 0.15 | 0.21 | -0.12 |
         |  2   |   1.0  | 0.23 | 0.41 | -0.28 |
           ...
         |  100 |   50   |  1.5 | 0.07 | -0.89 |

    '''

    df = pd.read_feather(file_response)

    if len(tags) == 0:
        tags = [int(x) for x in list(df.columns[3:])]

    if len(steps) == 0:
        steps = df['step']

    if len(dofs) == 0:
        val = list(df[str(tags[0])])[0]
        dofs = list(range(1, len(val) + 1))

    if len(scfactors) == 0:
        scfactors = [1] * len(dofs)

    tim = df.columns[2]

    cols_ = ['step', tim]
    cols_.extend([str(x) for x in tags])
    zerovals = [0, 0.0]
    zerovals.extend([0.0 for x in tags])

    df_s = []

    for i in range(len(dofs)):
        dof = dofs[i]
        scfactor = scfactors[i]
        df_dof = pd.DataFrame(columns=cols_)
        df_dof.loc[0] = zerovals
        ste_count = 1
        for step in steps:
            dff = df.loc[df['step'] == step]
            time = list(dff[tim])[0]
            vals = [step, time]
            for tag in tags:
                tag_str = str(tag)
                data_ = list(dff[tag_str])[0][dof - 1] * scfactor
                vals.append(data_)

            df_dof.loc[ste_count] = vals
            ste_count += 1

        df_s.append(df_dof)

    if sendtoexcel.lower() in ['y', 'yes']:
        for i in range(len(df_s)):
            sheet_name = 'DOF ' + str(dofs[i])
            __sendtoexcel(file_path, sheet_name, df_s[i])

    return df_s


#### Resp
def response(file_response, tags, dof=1, scfactor=1.0, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    Function to retrieve data from a saved response file in specific degree of freedom at all steps.

    file_response: saved response file path
    tags: A python list containing the Tags of nodes whose responses are retrieved
    dof: The DOF in which the responses are retrieved (optional, default value is 1)
    scfactor: The scale factor multiplied by the response (optional, default value is 1.0)
    sendtoexcel: Whether to send the retrieved data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the retrieved data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['step', 'time', tags as python string].
            Example: df.columns = ['step', 'time', '1', '2', ... 'n']

    Example:
                                     5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---
        file_Disp = "myproject\output_ntha\Disp.feather"
        df_disp = opa.response(file_Disp, tags=[1,3,5], dof=1, scfactor=1, sendtoexcel='y', file_path='disps.xlsx',
               sheet_name='disp')

        df_disp :

         |'step'| 'time' |  '1' |  '3' |  '5'  |
         --------------------------------------
         |  0   |   0.0  | 0.0  | 0.0  |  0.0  |
         |  1   |   0.5  | 0.15 | 0.21 | -0.12 |
         |  2   |   1.0  | 0.23 | 0.41 | -0.28 |
           ...
         |  100 |   50   |  1.5 | 0.07 | -0.89 |

    '''


    df_s = retrieve_data(file_response, tags=tags, scfactors=[scfactor], dofs=[dof], sendtoexcel='n')
    df_resp = df_s[0]

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_resp)

    return df_resp


def plot_response(ax, df_resp, xlabel='', ylabel='', title='', linewidth=1.0, legend=[], legend_font=0.0,
             legend_loc='best', grid=True):

    """
    Function to plot data obtained by response function.

    ax: The axes on which the plot is drawn.
    df_resp: A DataFrame is returned by response function.
    xlabel: Label for the x-axis (optional, default value is '')
    ytitle: Label for the y-axis (optional, default value is '')
    ylabel:  Title for the axes (optional, default value is '')
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    legend: A python list containing the elements to be added to the legend (optional, default value is an empty list)
    legend_font: Legend font size (optional, default value is 0.0)
    legend_loc: Legend location (optional, default value is 'best')
    grid: Whether to show the grid lines (optional, default value is True)

    Notes:
        1- If legend is an empty list, elements to be added to the legend will be node/element tags
        2- Valid values for legend_loc are:
            best
            upper right
            upper left
            lower left
            lower right
            right
            center left
            center right
            lower center
            upper center
            center

    """

    tags = [int(x) for x in list(df_resp.columns[2:])]
    tim = df_resp.columns[1]
    times = list(df_resp[tim])
    mylegend = []
    for tag in tags:
        mylegend.append(str(tag))
        values = df_resp[str(tag)]
        ax.plot(times, values, linewidth=linewidth)

    if len(legend) == 0:
        legend = mylegend

    if legend_font <= 0:
        ax.legend(legend, loc=legend_loc)
    else:
        ax.legend(legend, fontsize=legend_font, loc=legend_loc)

    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


###################################################################
#### IDR

def IDR(file_disp, st_tags, st_heights, dof=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='IDR'):
    '''
    A function to extract the drift history of floors.

    file_disp: Nodal displacements storage file path (Disp.feather).
    st_tags: A python list containing the tags of nodes whose displacements are used to calculate the story drifts.
    st_heights: Story heights as a python list.
    dof: The DOF in which the displacements are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'IDR')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['step', 'time', stories as python string].
            Example: df.columns = ['step', 'time', 'story 1', 'story 2', ... 'story n']

    Example:

                                    5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---

        file_Disp = "myproject\output_ntha\Disp.feather"
        df_idr = opa.IDR(file_Disp, st_tags=[1,3,5], st_heights=[3.2, 3.2], dof=1, sendtoexcel='y', file_path='drifts.xlsx',
               sheet_name='drift')

        df_idr :

         |'step'| 'time' | 'story 1'|'story 2'|
         -------------------------------------
         |  1   |   0.0  |    0.0   |   0.0   |
         |  1   |   0.5  |    0.8   |   1.2   |
         |  2   |   1.0  |    1.1   |   1.0   |
           ...
         |  100 |   50   |    0.1   |   0.0   |

    '''
    df_s = retrieve_data(file_disp, tags=st_tags, scfactors=[], dofs=[dof], sendtoexcel='n')
    data_resp = df_s[0]
    tim = data_resp.columns[1]
    cols_ = ['step', tim]
    zerovals = [0, 0.0]
    for i in range(len(st_heights)):
        cols_.append('story ' + str(i+1))
        zerovals.append(0.0)

    times = list(data_resp[tim])
    steps = list(data_resp['step'])

    df_idr = pd.DataFrame(columns=cols_)
    df_idr.loc[0] = zerovals

    drifts = {}
    for i in range(len(st_tags)-1):
        output_time_1 = data_resp[str(st_tags[i])]
        output_time_2 = data_resp[str(st_tags[i + 1])]
        drifts[i+1] = [(x - y) * 100 / st_heights[i] for (x, y) in zip(list(output_time_1), list(output_time_2))]

    for i in range(len(times)):
        vals = [steps[i], times[i]]
        for j in range(len(st_heights)):
            vals.append(drifts[j+1][i])

        df_idr.loc[i+1] = vals

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_idr)

    return df_idr


def plot_IDR(ax, df_idr, xlabel='time', ylabel='IDR[%]', title='', linewidth=1.0, grid=True, legend_font=0.0,
             legend_loc='best'):
    """
    Function to plot data obtained by IDR function.

    ax: The axes on which the plot is drawn.
    df_idr: A DataFrame is returned by IDR function.
    xlabel: Label for the x-axis (optional, default value is 'time')
    ylabel: Label for the y-axis (optional, default value is 'IDR[%]')
    title:  Title for the axes (optional, default value is '')
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    legend_font: Legend font size (optional, default value is 0.0)
    legend_loc: Legend location (optional, default value is 'best')
    grid: Whether to show the grid lines (optional, default value is True)

    Notes:
         1- Valid values for legend_loc are:
            best
            upper right
            upper left
            lower left
            lower right
            right
            center left
            center right
            lower center
            upper center
            center

    """
    stories = list(df_idr.columns[2:])
    tim = df_idr.columns[1]
    times = df_idr[tim]
    legend = []
    for story in stories:
        legend.append(story)
        values = df_idr[story]
        ax.plot(times, values, linewidth=linewidth)

    ax.grid(grid)
    if legend_font <= 0:
        ax.legend(legend, loc=legend_loc)
    else:
        ax.legend(legend, fontsize=legend_font, loc=legend_loc)

    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


#### RDR
def RDR(file_disp, st_tags, st_heights, dof=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='RDR'):
    '''
       A function to extract the residual drifts of floors.

        file_disp: Nodal displacements storage file path (Disp.feather).
        st_tags: A python list containing the tags of nodes whose displacements are used to calculate the story drifts.
        st_heights: Story heights as a python list.
        dof: The DOF in which the displacements are retrieved (optional, default value is 1)
        sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
        file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
        sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

        return: Data as a pandas DataFrame.
                The column labels of the DataFrame are ['story', 'RDR'].
                df.columns = ['story', 'RDR']

        Example:

                                    5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---

            file_Disp = "myproject\output_ntha\Disp.feather"
            df_rdr = opa.RDR(file_Disp, st_tags=[1,3,5], st_heights=[3.2, 3.2], dof=1, sendtoexcel='y', file_path='drifts.xlsx',
                   sheet_name='rdr')

            df_rdr :

             |'story'| 'RDR' |
             ----------------
             |  0   |   0.0  |
             |  1   |   0.1  |
             |  2   |   0.2  |

    '''


    df_idr = IDR(file_disp, st_tags, st_heights, dof=dof, sendtoexcel='n')
    cols_ = ['story', 'RDR']
    zerovals = [0, 0.0]
    df_rdr = pd.DataFrame(columns=cols_)
    df_rdr.loc[0] = zerovals

    count = 1
    for i in range(len(st_heights)):
        absdrift = [abs(x) for x in list(df_idr['story ' + str(i+1)])]
        df_rdr.loc[count] = [str(i+1), absdrift[-1]]
        count += 1

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_rdr)

    return df_rdr


def plot_RDR(ax, df_rdr, xlabel='RDR[%]', ylabel='Story', title='', linewidth=1.0, color='k', marker='o', grid=False):
    """
    Function to plot data obtained by RDR function.

    ax: The axes on which the plot is drawn.
    df_rdr: A DataFrame is returned by RDR function.
    xlabel: Label for the x-axis (optional, default value is 'RDR[%]')
    ylabel: Label for the y-axis (optional, default value is 'Story')
    title:  Title for the axes (optional, default value is '')
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')
    marker: Marker type of the graph plot (optional, default value is 'o')
    grid: Whether to show the grid lines (optional, default value is False)

    """
    st_rdr = [int(x) for x in list(df_rdr['story'])]
    val_rdr = [x for x in list(df_rdr['RDR'])]
    ax.plot(val_rdr, st_rdr, color=color, lw=linewidth, marker=marker, markerfacecolor='w')

    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)

    yticks = st_rdr
    ax.set_yticks(yticks)


#### MaxIDR
def MaxIDR(file_disp, st_tags, st_heights, dof=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='MaxIDR'):
    '''
    A function to extract the maximum interstory drift ratios of the floors.

    file_disp: Nodal displacements storage file path (Disp.feather).
    st_tags: A python list containing the tags of nodes whose displacements are used to calculate the story drifts.
    st_heights: Story heights as a python list.
    dof: The DOF in which the displacements are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'MaxIDR')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['story', 'MaxIDR'].
            df.columns = ['story', 'MaxIDR']

    Example:

                                    5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---

        file_Disp = "myproject\output_ntha\Disp.feather"
        df_maxidr = opa.IDR(file_Disp, st_tags=[1,3,5], st_heights=[3.2, 3.2], dof=1, sendtoexcel='y', file_path='drifts.xlsx',
               sheet_name='maxidr')

        df_maxidr :

             |'story'| 'MaxIDR' |
             --------------------
             |  0    |    0.0   |
             |  1    |    1.5   |
             |  2    |    2.2   |

    '''

    df_idr = IDR(file_disp, st_tags, st_heights, dof=dof, sendtoexcel='n')

    cols_ = ['story', 'MaxIDR']
    zerovals = ['0', 0.0]
    df_maxidr = pd.DataFrame(columns=cols_)
    df_maxidr.loc[0] = zerovals

    count = 1
    for i in range(len(st_heights)):
        absdrift = [abs(x) for x in list(df_idr['story ' + str(i+1)])]
        df_maxidr.loc[count] = [str(i+1), max(absdrift)]
        count += 1

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_maxidr)

    return df_maxidr


def plot_MaxIDR(ax, df_maxidr, xlabel='IDR_Max[%]', ylabel='Story', title='', linewidth=1.0, color='k', marker='o',
                grid=False):
    """
    Function to plot data obtained by MaxIDR function.

    ax: The axes on which the plot is drawn.
    df_maxidr: A DataFrame is returned by MaxIDR function.
    xlabel: Label for the x-axis (optional, default value is 'IDR_Max[%]')
    ylabel: Label for the y-axis (optional, default value is 'Story')
    title:  Title for the axes (optional, default value is '')
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')
    marker: Marker type of the graph plot (optional, default value is 'o')
    grid: Whether to show the grid lines (optional, default value is True)

    """
    st_rdr = [int(x) for x in list(df_maxidr['story'])]
    val_rdr = [x for x in list(df_maxidr['MaxIDR'])]
    ax.plot(val_rdr, st_rdr, color=color, lw=linewidth, marker=marker, markerfacecolor='w')

    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)

    yticks = st_rdr
    ax.set_yticks(yticks)


#### MaxIDR
def MaxIDR_Mean(files_disp, st_tags, st_heights, dof=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1',
                rec_names=[]):
    '''
    A function to extract the maximum interstory drift ratios of the floors for a suite of ground motion records along with mean
    value and standard deviation at each floor.

    files_disp: A python list containing the Nodal displacements storage file pathes corresponding to each ground motion (Disp.feather).
    st_tags: A python list containing the tags of nodes whose displacements are used to calculate the story drifts.
    st_heights: Story heights as a python list.
    dof: The DOF in which the displacements are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')
    rec_names: A pyhon list containing the names specified for each ground motion by the user(optional, default value is an empty list)

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['story', record names, 'mean', 'std'].
            Example: df.columns = ['story', 'gm_1', 'gm_2', 'gm_3', 'mean', 'std']

    Example:

                                    5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---

        files_Disp = ["myproject\output_ntha_1\Disp.feather",
                      "myproject\output_ntha_2\Disp.feather",
                      "myproject\output_ntha_3\Disp.feather"]
        df_maxidr = opa.MaxIDR_Mean(files_Disp, st_tags=[1,3,5], st_heights=[3.2, 3.2], dof=1, sendtoexcel='y', file_path='drifts.xlsx',
               sheet_name='maxidr_mean')

        df_maxidr :

             |'story'| 'gm_1' | 'gm_2' | 'gm_3' | 'mean' | 'std'  |
             ------------------------------------------------------
             |  0    |  0.0   |  0.0   |  0.0   |  0.0   |   0.0  |
             |  1    |  1.5   |  0.8   |  0.9   |  1.066 | 0.3786 |
             |  2    |  2.2   |  1.3   |  0.7   |  1.4   | 0.7550 |

    '''

    cols_ = ['story']
    zerovals = ['0']
    for i in range(len(files_disp)):
        zerovals.append(0.0)

    if len(rec_names) == 0:
        for i in range(len(files_disp)):
            cols_.append('gm_' + str(i + 1))

    else:
        cols_.extend(rec_names)

    cols_.append('mean')
    cols_.append('std')
    zerovals.append(0.0)
    zerovals.append(0.0)

    df_maxdir_mean = pd.DataFrame(columns=cols_)

    for i in range(len(st_heights)+1):
        df_maxdir_mean.loc[i] = zerovals

    for i in range(len(files_disp)):
        file_disp = files_disp[i]
        df_maxidr = MaxIDR(file_disp, st_tags, st_heights, dof=dof, sendtoexcel='n')
        df_maxdir_mean['story'] = list(df_maxidr['story'])
        df_maxdir_mean[cols_[i+1]] = list(df_maxidr['MaxIDR'])

    meanvals = [0.0]
    stdvals = [0.0]
    for i in range(len(st_heights)):
        vals = list(df_maxdir_mean.loc[i+1])
        mean = sum(vals[1:-2]) / len(files_disp)
        val_mean = [(x - mean) ** 2 for x in vals[1:-2]]
        stdvals.append((sum(val_mean) / (len(files_disp) - 1)) ** 0.5)
        meanvals.append(mean)

    df_maxdir_mean['mean'] = meanvals
    df_maxdir_mean['std'] = stdvals

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_maxdir_mean)

    return df_maxdir_mean


def plot_MaxIDR_Mean(ax, df_maxdir_mean, plot_mean='y', plot_std='y', xlabel='IDR_Max[%]', ylabel='Story', title='', rec_names=[],
                     legend_font=0.0, legend_loc='best', grid=False):
    """
    Function to plot data obtained by MaxIDR_Mean function.

    ax: The axes on which the plot is drawn.
    df_maxdir_mean: A DataFrame is returned by MaxIDR_Mean function.
    plot_mean: Whether to plot mean values
    plot_std:  Whether to plot mean + std  and mean - std
    xlabel: Label for the x-axis (optional, default value is 'IDR_Max[%]')
    ylabel: Label for the y-axis (optional, default value is 'Story')
    title:  Title for the axes (optional, default value is '')
    rec_names: A pyhon list containing the names specified for each ground motion by the user(optional, default value is an empty list)
    legend_font: Legend font size (optional, default value is 0.0)
    legend_loc: Legend location (optional, default value is 'best')
    grid: Whether to show the grid lines (optional, default value is True)

    Notes:
         1- Valid values for legend_loc are:
            best
            upper right
            upper left
            lower left
            lower right
            right
            center left
            center right
            lower center
            upper center
            center

    """

    cols_ = list(df_maxdir_mean.columns)
    names_recs = cols_[1:-2]
    if len(rec_names) == 0:
        rec_names = [x for x in names_recs]

    if plot_mean.lower() in ['y', 'yes']:
        rec_names.append('Mean')

    if plot_std.lower() in ['y', 'yes']:
        rec_names.append('Mean + STD')
        rec_names.append('Mean - STD')

    for rec in names_recs:
        st_idr = list(df_maxdir_mean[rec])
        X = []
        Y = []
        for i in range(len(st_idr)):
            X.append(st_idr[i])
            Y.append(i)
        ax.plot(X, Y, lw=0.8, marker='o', markerfacecolor='w')

    st_idr = list(df_maxdir_mean['mean'])
    st_idr_std = list(df_maxdir_mean['std'])

    X = []
    X_std_a = []
    X_std_s = []
    Y = []
    for i in range(len(st_idr)):
        X.append(st_idr[i])
        X_std_a.append(st_idr[i] + st_idr_std[i])
        X_std_s.append(st_idr[i] - st_idr_std[i])
        Y.append(i)

    if plot_mean.lower() in ['y', 'yes']:
         ax.plot(X, Y, color='k',  lw=3.0, marker='s', markerfacecolor='w')

    if plot_std.lower() in ['y', 'yes']:
        ax.plot(X_std_a, Y, color='k', lw=1.5, marker='s', markerfacecolor='w', linestyle='dotted')
        ax.plot(X_std_s, Y, color='k', lw=1.5, marker='s', markerfacecolor='w', linestyle='dashed')


    yticks = [int(x) for x in list(df_maxdir_mean['story'])]

    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)

    if len(rec_names) != 0:
        if legend_font <= 0:
            ax.legend(rec_names, loc=legend_loc)
        else:
            ax.legend(rec_names, fontsize=legend_font, loc=legend_loc)

    ax.set_yticks(yticks)


#### RDR Mean
def RDR_Mean(files_disp, st_tags, st_heights, dof=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1',
                rec_names=[]):
    '''
    A function to extract the residual interstory drift ratios of the floors for a suite of ground motion records along with mean
    value and standard deviation at each floor.

    files_disp: A python list containing the Nodal displacements storage file pathes corresponding to each ground motion (Disp.feather).
    st_tags: A python list containing the tags of nodes whose displacements are used to calculate the story drifts.
    st_heights: Story heights as a python list.
    dof: The DOF in which the displacements are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')
    rec_names: A pyhon list containing the names specified for each ground motion by the user(optional, default value is an empty list)

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['story', record names, 'mean', 'std'].
            Example: df.columns = ['story', 'gm_1', 'gm_2', 'gm_3', 'mean', 'std']

    Example:

                                    5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---

        files_Disp = ["myproject\output_ntha_1\Disp.feather",
                      "myproject\output_ntha_2\Disp.feather",
                      "myproject\output_ntha_3\Disp.feather"]
        df_rdr = opa.MaxIDR_Mean(files_Disp, st_tags=[1,3,5], st_heights=[3.2, 3.2], dof=1, sendtoexcel='y', file_path='drifts.xlsx',
               sheet_name='idr_mean')

        df_rdr :

             |'story'| 'gm_1' | 'gm_2' | 'gm_3' | 'mean' | 'std'  |
             ------------------------------------------------------
             |  0    |  0.0   |  0.0   |  0.0   |  0.0   |   0.0  |
             |  1    |  1.5   |  0.8   |  0.9   |  1.066 | 0.3786 |
             |  2    |  2.2   |  1.3   |  0.7   |  1.4   | 0.7550 |

    '''

    cols_ = ['story']
    zerovals = ['0']
    for i in range(len(files_disp)):
        zerovals.append(0.0)

    if len(rec_names) == 0:
        for i in range(len(files_disp)):
            cols_.append('gm_' + str(i + 1))

    else:
        cols_.extend(rec_names)

    cols_.append('mean')
    cols_.append('std')
    zerovals.append(0.0)
    zerovals.append(0.0)

    df_rdr_mean = pd.DataFrame(columns=cols_)

    for i in range(len(st_heights) + 1):
        df_rdr_mean.loc[i] = zerovals

    for i in range(len(files_disp)):
        file_disp = files_disp[i]
        df_maxidr = RDR(file_disp, st_tags, st_heights, dof=dof, sendtoexcel='n')
        df_rdr_mean['story'] = list(df_maxidr['story'])
        df_rdr_mean[cols_[i + 1]] = list(df_maxidr['RDR'])

    meanvals = [0.0]
    stdvals = [0.0]
    for i in range(len(st_heights)):
        vals = list(df_rdr_mean.loc[i + 1])
        mean = sum(vals[1:-2]) / len(files_disp)
        val_mean = [(x - mean) ** 2 for x in vals[1:-2]]
        stdvals.append((sum(val_mean) / (len(files_disp) - 1)) ** 0.5)
        meanvals.append(mean)

    df_rdr_mean['mean'] = meanvals
    df_rdr_mean['std'] = stdvals

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_rdr_mean)

    return df_rdr_mean


def plot_RDR_Mean(ax, df_rdr_mean, plot_mean='y', plot_std='y', xlabel='RDR[%]', ylabel='Story', title='', rec_names=[], legend_font=0.0,
                  legend_loc='best', grid=False):
    """
    Function to plot data obtained by RDR_Mean function.

    ax: The axes on which the plot is drawn.
    df_rdr_mean: A DataFrame is returned by RDR_Mean function.
    plot_mean: Whether to plot mean values
    plot_std:  Whether to plot mean + std  and mean - std
    xlabel: Label for the x-axis (optional, default value is 'RDR[%]')
    ylabel: Label for the y-axis (optional, default value is 'Story')
    title:  Title for the axes (optional, default value is '')
    rec_names: A pyhon list containing the names specified for each ground motion by the user(optional, default value is an empty list)
    legend_font: Legend font size (optional, default value is 0.0)
    legend_loc: Legend location (optional, default value is 'best')
    grid: Whether to show the grid lines (optional, default value is True)

    Notes:
         1- Valid values for legend_loc are:
            best
            upper right
            upper left
            lower left
            lower right
            right
            center left
            center right
            lower center
            upper center
            center

    """

    cols_ = list(df_rdr_mean.columns)
    names_recs = cols_[1:-2]
    if len(rec_names) == 0:
        rec_names = [x for x in names_recs]

    if plot_mean.lower() in ['y', 'yes']:
        rec_names.append('Mean')

    if plot_std.lower() in ['y', 'yes']:
        rec_names.append('Mean + STD')
        rec_names.append('Mean - STD')

    for rec in names_recs:
        st_idr = list(df_rdr_mean[rec])
        X = []
        Y = []
        for i in range(len(st_idr)):
            X.append(st_idr[i])
            Y.append(i)
        ax.plot(X, Y, lw=0.8, marker='o', markerfacecolor='w')

    st_idr = list(df_rdr_mean['mean'])
    st_idr_std = list(df_rdr_mean['std'])

    X = []
    X_std_a = []
    X_std_s = []
    Y = []
    for i in range(len(st_idr)):
        X.append(st_idr[i])
        X_std_a.append(st_idr[i] + st_idr_std[i])
        X_std_s.append(st_idr[i] - st_idr_std[i])
        Y.append(i)

    if plot_mean.lower() in ['y', 'yes']:
         ax.plot(X, Y, color='k',  lw=3.0, marker='s', markerfacecolor='w')

    if plot_std.lower() in ['y', 'yes']:
        ax.plot(X_std_a, Y, color='k', lw=1.5, marker='s', markerfacecolor='w', linestyle='dotted')
        ax.plot(X_std_s, Y, color='k', lw=1.5, marker='s', markerfacecolor='w', linestyle='dashed')

    yticks = [int(x) for x in list(df_rdr_mean['story'])]
    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)

    if len(rec_names) != 0:
        if legend_font <= 0:
            ax.legend(rec_names, loc=legend_loc)
        else:
            ax.legend(rec_names, fontsize=legend_font, loc=legend_loc)

    ax.set_yticks(yticks)


###################################################################
def plot_IDA(ax, df_ida, xlabel='DM', ylabel='IM', title='IDA Curve', linewidth=1.0, color='k',
             marker='o', grid=False):
    """
    Function to plot data obtained by an IDA function.

    ax: The axes on which the plot is drawn.
    df_ida: A DataFrame is returned by an IDA function.
    xlabel: Label for the x-axis (optional, default value is 'DM')
    ylabel: Label for the y-axis (optional, default value is 'IM')
    title:  Title for the axes (optional, default value is 'IDA Curve')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')
    marker: Marker type of the graph plot (optional, default value is 'o')

    """

    values_IM = list(df_ida['IM'])
    values = list(df_ida['value'])
    ax.plot(values, values_IM, lw=linewidth, color=color, marker=marker, markerfacecolor='w')

    yticks = values_IM
    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)
    ax.set_yticks(yticks)


#### IDA
def IDA_Baseshear(files_reaction, IM, tags_base, dof=1, scfactor=1.0, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    A function to extract the absolute maximum base shear values for a suite of ground motion records.

    files_reaction: A python list containing the nodal reactions storage file pathes corresponding to a suite of ground motion records, (Reaction.feather).
    IM: A python list containing IM values.
    tags_base: A python list containing tags of the nodes whose reactions are used to calculate the base shear.
    scfactor: A factor to multiply base shears by (optional, default value is 1.0)
    dof: The DOF in which the reactions are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['IM', 'value'].

    Example:

                                    5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---

        files_reaction = ["myproject\output_ntha_1\Reaction.feather",
                          "myproject\output_ntha_2\Reaction.feather",
                          "myproject\output_ntha_3\Reaction.feather",
                          "myproject\output_ntha_4\Reaction.feather"]
        IM = [0.1, 0.5, 1.0, 1.5]
        df_ida = opa.IDA_Baseshear(files_reaction, IM, tags_base=[1,2], dof=1, scfactor=1, sendtoexcel='y', file_path='IDA.xlsx',
        sheet_name='base shear')

        df_ida :

             | 'IM' | 'value' |
             -------------------
             |  0.0 |   0.0   |
             |  0.1 |   120   |
             |  0.5 |   400   |
             |  1.0 |   800   |
             |  1.5 |   850   |

    '''


    cols_ = ['IM', 'value']
    df_ida_base = pd.DataFrame(columns=cols_)
    df_ida_base.loc[0] = [0.0, 0.0]

    for i in range(len(files_reaction)):
        file_reaction = files_reaction[i]
        df_base = baseshear(file_reaction, tags_base, dof=dof, scfactor=scfactor, sendtoexcel='n')
        values = [abs(x) for x in list(df_base['base shear'])]
        df_ida_base.loc[i+1] = [IM[i], max(values)]

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_ida_base)

    return df_ida_base


def IDA_Resp(files_resp, IM, tag, dof=1, scfactor=1.0, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    A function to extract the absolute maximum response values of a node or an element for a suite of ground motion records.

    files_resp: A python list containing nodal/element response storage file pathes corresponding to a suite of ground motion records
             (Reaction.feather, Disp.feather, Vel.feather, Accel.feather, force.feather, stress.feather, ...).
    IM: A python list containing IM values.
    tag: tag of the specified node/element.
    scfactor: A factor to multiply the response by(optional, default value is 1.0)
    dof: The DOF in which the responses are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['IM', 'value'].

    Example:

                                    5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---

        files_resp = ["myproject\output_ntha_1\Disp.feather",
                      "myproject\output_ntha_2\Disp.feather",
                      "myproject\output_ntha_3\Disp.feather",
                      "myproject\output_ntha_4\Disp.feather"]
        IM = [0.1, 0.5, 1.0, 1.5]
        df_ida = opa.IDA_Resp(files_resp, IM, tag=6, dof=1, scfactor=1, sendtoexcel='y', file_path='IDA.xlsx',
        sheet_name='roof displacement')

        df_ida :

             | 'IM' | 'value' |
             -------------------
             |  0.0 |   0.0   |
             |  0.1 |   1.1   |
             |  0.5 |   2.3   |
             |  1.0 |   5.0   |
             |  1.5 |   10.2  |

    '''
    cols_ = ['IM', 'value']
    df_ida_resp = pd.DataFrame(columns=cols_)
    df_ida_resp.loc[0] = [0.0, 0.0]

    for i in range(len(files_resp)):
        file_ = files_resp[i]
        df_resp = response(file_, [tag], dof=dof, scfactor=scfactor, sendtoexcel='n')
        values = [abs(x) for x in list(df_resp[str(tag)])]
        df_ida_resp.loc[i+1] = [IM[i], max(values)]

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_ida_resp)

    return df_ida_resp


def IDA_MaxResp(files_resp, IM, tags, dof=1, scfactor=1.0, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    A function to extract the absolute maximum response values of a set of nodes or elements for a suite of ground motion records.

    files_resp: A python list containing nodal/element response storage file pathes corresponding to a suite of ground motion records
             (Reaction.feather, Disp.feather, Vel.feather, Accel.feather, force.feather, stress.feather, ...).
    IM: A python list containing IM values.
    tags: Tags of specified nodes/elements.
    scfactor: A factor to multiply the response by (optional, default value is 1.0)
    dof: The DOF in which the reactions are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['IM', 'value'].

    Example:

                                    5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---

        files_resp = ["myproject\output_ntha_1\Disp.feather",
                      "myproject\output_ntha_2\Disp.feather",
                      "myproject\output_ntha_3\Disp.feather",
                      "myproject\output_ntha_4\Disp.feather"]
        IM = [0.1, 0.5, 1.0, 1.5]
        df_ida = opa.IDA_Resp(files_resp, IM, tags=[4, 6], dof=1, scfactor=1, sendtoexcel='y', file_path='IDA.xlsx',
        sheet_name='maximun displacement')

        df_ida :

             | 'IM' | 'value' |
             -------------------
             |  0.0 |   0.0   |
             |  0.1 |   1.1   |
             |  0.5 |   2.5   |
             |  1.0 |   5.0   |
             |  1.5 |   10.2  |

    '''

    cols_ = ['IM', 'value']
    df_ida_maxresp = pd.DataFrame(columns=cols_)
    df_ida_maxresp.loc[0] = [0.0, 0.0]

    maxvals = []
    for i in range(len(files_resp)):
        file_ = files_resp[i]
        df_resp = response(file_, tags, dof=dof, scfactor=scfactor, sendtoexcel='n')
        for tag in tags:
            values = [abs(x) for x in list(df_resp[str(tag)])]
            maxvals.append(max(values))

        df_ida_maxresp.loc[i+1] = [IM[i], max(values)]

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_ida_maxresp)

    return df_ida_maxresp


def IDA_Drift(files_disp, IM, st_tags, st_heights, story=-1, dof=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    A function to extract the absolute maximum drift ratios at a specified story for a suite of ground motion records.

    files_disp: A python list containing nodal displacements storage file pathes corresponding to a suite of ground motion records(Disp.feather).
    IM: A python list containing IM values.
    st_tags: A python list containing the tags of nodes whose displacements are used to calculate the story drifts.
    st_heights: Story heights as a python list.
    story: Specified story (optional, default value is -1)
    dof: The DOF in which the displacements are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['IM', 'value'].

    Example:

                                    5                                 6
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h2 = 3.2
                                   |                                 |
                                   | 3                               | 4
                                   ----------------------------------
                                   |                                 |
                                   |                                 | h1 = 3.2
                                   |                                 |
                                   | 1                               | 2
                                  ---                               ---

        files_disp = ["myproject\output_ntha_1\Disp.feather",
                      "myproject\output_ntha_2\Disp.feather",
                      "myproject\output_ntha_3\Disp.feather",
                      "myproject\output_ntha_4\Disp.feather"]
        IM = [0.1, 0.5, 1.0, 1.5]
        df_ida = opa.IDA_Drift(files_disp, IM, st_tags=[2, 4, 6], st_heights=[3.2, 3.2], story=-1, dof=1, sendtoexcel='y', file_path='IDA.xlsx',
        sheet_name='Roof Drift')

        df_ida :

             | 'IM' | 'value' |
             -------------------
             |  0.0 |   0.0   |
             |  0.1 |   0.1   |
             |  0.5 |   0.3   |
             |  1.0 |   2.1   |
             |  1.5 |   2.5   |

    '''
    cols_ = ['IM', 'value']
    df_ida_drift = pd.DataFrame(columns=cols_)
    df_ida_drift.loc[0] = [0.0, 0.0]

    if story == -1:
        story = len(st_heights)

    for i in range(len(files_disp)):
        file_disp = files_disp[i]
        df_drift = IDR(file_disp, st_tags, st_heights, dof=dof, sendtoexcel='n')
        value = df_drift['story ' + str(story)]
        absdrift = [abs(x) for x in value]
        df_ida_drift.loc[i + 1] = [IM[i], max(absdrift)]

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_ida_drift)

    return df_ida_drift


def IDA_MaxDrift(files_disp, IM, st_tags, st_heights, dof=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    A function to extract the absolute maximum drift ratios for a suite of ground motion records.

    files_disp: A python list containing nodal displacements storage file pathes corresponding to a suite of ground motion records(Disp.feather).
    IM: A python list containing IM values.
    st_tags: A python list containing the tags of nodes whose displacements are used to calculate the story drifts.
    st_heights: Story heights as a python list.
    dof: The DOF in which the reactions are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['IM', 'value'].

    Example:

                                       5                                 6
                                      ----------------------------------
                                      |                                 |
                                      |                                 | h2 = 3.2
                                      |                                 |
                                      | 3                               | 4
                                      ----------------------------------
                                      |                                 |
                                      |                                 | h1 = 3.2
                                      |                                 |
                                      | 1                               | 2
                                     ---                               ---

        files_disp = ["myproject\output_ntha_1\Disp.feather",
                         "myproject\output_ntha_2\Disp.feather",
                         "myproject\output_ntha_3\Disp.feather",
                         "myproject\output_ntha_4\Disp.feather"]
        IM = [0.1, 0.5, 1.0, 1.5]
        df_ida = opa.IDA_MaxDrift(files_disp, IM, st_tags=[4, 6], st_heights=[3.2, 3.2], dof=1, sendtoexcel='y', file_path='IDA.xlsx',
        sheet_name='Maximum Drift')

        df_ida :

                | 'IM' | 'value' |
                -------------------
                |  0.0 |   0.0   |
                |  0.1 |   0.1   |
                |  0.5 |   0.3   |
                |  1.0 |   2.1   |
                |  1.5 |   2.5   |

    '''
    cols_ = ['IM', 'value']
    df_ida_maxdrift = pd.DataFrame(columns=cols_)
    df_ida_maxdrift.loc[0] = [0.0, 0.0]

    for i in range(len(files_disp)):
        file_disp = files_disp[i]
        df_drift = IDR(file_disp, st_tags, st_heights, dof=dof, sendtoexcel='n')

        maxidr = []
        for j in range(len(st_heights)):
            st_name = 'story ' + str(j+1)
            absdrift = [abs(x) for x in df_drift[st_name]]
            maxidr.append(max(absdrift))

        df_ida_maxdrift.loc[i + 1] = [IM[i], max(maxidr)]

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_ida_maxdrift)

    return df_ida_maxdrift


def IDA_RDrift(files_disp, IM, st_tags, st_heights, story=-1, dof=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    A function to extract the absolute maximum residual drift ratios at a specified story for a suite of ground motion records.

    files_disp: A python list containing nodal displacements storage file pathes corresponding to a suite of ground motion records(Disp.feather).
    IM: A python list containing IM values.
    st_tags: A python list containing the tags of nodes whose displacements are used to calculate the story drifts.
    st_heights: Story heights as a python list.
    story: Specified story (optional, default value is -1)
    dof: The DOF in which the reactions are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['IM', 'value'].

    Example:

                                       5                                 6
                                      ----------------------------------
                                      |                                 |
                                      |                                 | h2 = 3.2
                                      |                                 |
                                      | 3                               | 4
                                      ----------------------------------
                                      |                                 |
                                      |                                 | h1 = 3.2
                                      |                                 |
                                      | 1                               | 2
                                     ---                               ---

        files_disp = ["myproject\output_ntha_1\Disp.feather",
                         "myproject\output_ntha_2\Disp.feather",
                         "myproject\output_ntha_3\Disp.feather",
                         "myproject\output_ntha_4\Disp.feather"]
        IM = [0.1, 0.5, 1.0, 1.5]
        df_ida = opa.IDA_RDrift(files_disp, IM, st_tags=[2, 4, 6], st_heights=[3.2, 3.2], story=-1, dof=1, sendtoexcel='y', file_path='IDA.xlsx',
        sheet_name='Roof Residual Drift')

        df_ida :

                | 'IM' | 'value' |
                -------------------
                |  0.0 |   0.0   |
                |  0.1 |   0.0   |
                |  0.5 |   0.1   |
                |  1.0 |   0.5   |
                |  1.5 |   1.0   |

    '''

    cols_ = ['IM', 'value']
    df_ida_rdrift = pd.DataFrame(columns=cols_)
    df_ida_rdrift.loc[0] = [0.0, 0.0]

    if story == -1:
        story = len(st_heights)

    for i in range(len(files_disp)):
        file_disp = files_disp[i]
        df_drift = RDR(file_disp, st_tags, st_heights, dof=dof, sendtoexcel='n')
        value = df_drift.loc[story]['RDR']
        df_ida_rdrift.loc[i + 1] = [IM[i], abs(value)]

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_ida_rdrift)

    return df_ida_rdrift


def IDA_MaxRDrift(files_disp, IM, st_tags, st_heights, dof=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    A function to extract the absolute maximum residual drift ratios for a suite of ground motion records.

    files_disp: A python list containing nodal displacements storage file pathes corresponding to a suite of ground motion records(Disp.feather).
    IM: A python list containing IM values.
    st_tags: A python list containing the tags of nodes whose displacements are used to calculate the story drifts.
    st_heights: Story heights as a python list.
    dof: The DOF in which the reactions are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['IM', 'value'].

    Example:

                                       5                                 6
                                      ----------------------------------
                                      |                                 |
                                      |                                 | h2 = 3.2
                                      |                                 |
                                      | 3                               | 4
                                      ----------------------------------
                                      |                                 |
                                      |                                 | h1 = 3.2
                                      |                                 |
                                      | 1                               | 2
                                     ---                               ---

        files_disp = ["myproject\output_ntha_1\Disp.feather",
                         "myproject\output_ntha_2\Disp.feather",
                         "myproject\output_ntha_3\Disp.feather",
                         "myproject\output_ntha_4\Disp.feather"]
        IM = [0.1, 0.5, 1.0, 1.5]
        df_ida = opa.IDA_MaxRDrift(files_disp, IM, st_tags=[2, 4, 6], st_heights=[3.2, 3.2], dof=1, sendtoexcel='y', file_path='IDA.xlsx',
        sheet_name='Maximum Drift')

        df_ida :

                | 'IM' | 'value' |
                -------------------
                |  0.0 |   0.0   |
                |  0.1 |   0.1   |
                |  0.5 |   0.3   |
                |  1.0 |   2.1   |
                |  1.5 |   2.5   |

    '''
    cols_ = ['IM', 'value']
    df_ida_maxrdrift = pd.DataFrame(columns=cols_)
    df_ida_maxrdrift.loc[0] = [0.0, 0.0]

    for i in range(len(files_disp)):
        file_disp = files_disp[i]
        df_drift = RDR(file_disp, st_tags, st_heights, dof=dof, sendtoexcel='n')
        value = df_drift['RDR']
        absdrift = [abs(x) for x in value]
        df_ida_maxrdrift.loc[i + 1] = [IM[i], max(absdrift)]

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_ida_maxrdrift)

    return df_ida_maxrdrift


###################################################################
#### Base Shear
def baseshear(file_reaction, tags_base, dof=1, scfactor=1.0, sendtoexcel='y', file_path='book1.xlsx', sheet_name='base shear'):
    '''
    A function to extract base shear history.

    file_reaction: Nodal reactions storage file path(Reaction.feather).
    tags_base: A python list containing tags of the nodes whose reactions are used to calculate the base shear.
    scfactor: A factor to multiply base shear by.
    dof: The DOF in which the reactions are retrieved (optional, default value is 1)
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'base shear')

    return: Data as a pandas DataFrame.
            The column labels of the DataFrame are ['step', 'time', 'base shear'].

    Example:

                                       5                                 6
                                      ----------------------------------
                                      |                                 |
                                      |                                 | h2 = 3.2
                                      |                                 |
                                      | 3                               | 4
                                      ----------------------------------
                                      |                                 |
                                      |                                 | h1 = 3.2
                                      |                                 |
                                      | 1                               | 2
                                     ---                               ---

        file_reaction = "myproject\output_ntha_1\Reaction.feather"
        df_base = opa.baseshear(file_reaction, tags_base=[1, 2], dof=1, scfactor=1, sendtoexcel='y', file_path='book1.xlsx', sheet_name='base shear')

        df_base :

                | 'step' | 'time' | 'base shear' |
                ----------------------------------
                |   0    |   0.0  |     0.0      |
                |   1    |   0.5  |      20      |
                |   2    |   1.5  |      30      |
                |   3    |   2.0  |      40      |
                ...
                |  100   |   50   |     -20      |

    '''
    df_s = retrieve_data(file_reaction, tags=tags_base, scfactors=[scfactor], dofs=[dof], sendtoexcel='n')
    df_base = df_s[0]
    tim = df_base.columns[1]
    cols_ = ['step', tim, 'base shear']

    df_baseshear = pd.DataFrame(columns=cols_)
    for i in range(len(list(df_base[tim]))):
        vals = list(df_base.loc[i])
        vals_step = vals[:2]
        vals_base = sum(vals[2:]) * -1
        vals_step.append(vals_base)
        df_baseshear.loc[i] = vals_step

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_baseshear)

    return df_baseshear


def plot_baseshear(ax, df_baseshear, xlabel='Time', ylabel='Base Shear', title='', linewidth=1.0, color='k', grid=True):
    """
    Function to plot data obtained by baseshear function.

    ax: The axes on which the plot is drawn.
    df_baseshear: A DataFrame is returned by baseshear function.
    xlabel: Label for the x-axis (optional, default value is 'Time')
    ylabel: Label for the y-axis (optional, default value is 'Base Shear')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    tim = df_baseshear.columns[1]
    values_time = list(df_baseshear[tim])
    values_force = list(df_baseshear['base shear'])
    ax.plot(values_time, values_force, lw=linewidth, color=color)
    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


###################################################################
#### Resp - Base Shear
def resp_baseshear(file_resp, file_reaction, tag_resp, tags_base, dof_resp=1, dof_force=1, scfactor_resp=1.0, scfactor_force=1.0
                   , sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    A function to extract data as response-base shear pairs.

    file_Resp: Response storage file path (Reaction.feather, Disp.feather, Vel.feather, Accel.feather, force.feather, stress.feather, ...).
    file_reaction: Nodal reactions storage file path(Reaction.feather).
    tag_resp: Tag of the node/element whose responses are retrieved.
    tags_base: A python list containing tags of the nodes whose reactions are used to calculate the base shear.
    dof_resp: The DOF in which the response is retrieved (optional, default value is 1)
    dof_force: The DOF in which the reactions are retrieved (optional, default value is 1)
    scfactor_resp: A factor to multiply response by.
    scfactor_force: A factor to multiply base shears by.
    sendtoexcel: Whether to send the extracted data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the extracted data is stored (optional, default value is 'book1.xlsx')
    sheet_name: Data are stored in a sheet named sheet_name in excel file (optional, default value is 'sheet1')

    return: Data as a pandas DataFrame.
           The column labels of the DataFrame are ['step', 'time', 'response', 'base shear'].

    Example:

                                      5                                 6
                                     ----------------------------------
                                     |                                 |
                                     |                                 | h2 = 3.2
                                     |                                 |
                                     | 3                               | 4
                                     ----------------------------------
                                     |                                 |
                                     |                                 | h1 = 3.2
                                     |                                 |
                                     | 1                               | 2
                                    ---                               ---

        file_resp = "myproject\output_ntha_1\Disp.feather"
        file_reaction = "myproject\output_ntha_1\Reaction.feather"
        df_resp_base = opa.resp_baseshear(file_resp, file_reaction, tag_resp=6, tags_base=[1, 2], dof_resp=1, dof_force=1, scfactor_resp=1, scfactor_force=1
                       , sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1')

        df_resp_base :

                | 'step' | 'time' |  'response' | 'base shear' |
                --------------------------------------------------
                |   0    |   0.0  |     0.0     |     0.0      |
                |   1    |   0.5  |     1.0     |      20      |
                |   2    |   1.5  |     2.5     |      30      |
                |   3    |   2.0  |     3.0     |      40      |
                ...
                |  100   |   50   |    -1.5     |     -20      |

    '''
    df_s = retrieve_data(file_resp, tags=[tag_resp], scfactors=[scfactor_resp], dofs=[dof_resp], sendtoexcel='n')
    df_res = df_s[0]

    df_base = baseshear(file_reaction, tags_base, dof=dof_force, scfactor=scfactor_force, sendtoexcel='n')
    tim = df_base.columns[1]
    cols_ = ['step', tim, 'response', 'base shear']
    df_resp_base = pd.DataFrame(columns=cols_)
    df_resp_base['step'] = df_res['step']
    df_resp_base[tim] = df_res[tim]
    df_resp_base['response'] = df_res[str(tag_resp)]
    df_resp_base['base shear'] = df_base['base shear']

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_resp_base)

    return df_resp_base


def plot_resp_baseshear(ax, df_resp_base, xlabel='', ylabel='Base Shear', title='', linewidth=1.0, color='k', grid=True):
    """
    Function to plot data obtained by resp_baseshear function.

    ax: The axes on which the plot is drawn.
    df_resp_base: A DataFrame is returned by resp_baseshear function.
    xlabel: Label for the x-axis (optional, default value is '')
    ylabel: Label for the y-axis (optional, default value is 'Base Shear')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    values_time = list(df_resp_base['response'])
    values_force = list(df_resp_base['base shear'])
    ax.plot(values_time, values_force, lw=linewidth, color=color)
    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


###################################################################
def section_stress(file_stress, tag, section, y, z, steps=[], scfactor_stress=1.0, scfactor_strain=1.0, sendtoexcel='y',
                file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    Function to retrieve stress-strain data of a beam-column element section from a saved response file.

    file_stress: saved response file path(file name is section_stress.feather)
    tag: Beam-Column element tag
    section: Section number
    y: fibre y-coordinate
    z: fibre z-coordinate
    steps: A python list containing the steps in which the responses are retrieved (optional, default value is an empty list)
    scfactor_stress: Scale factor multiplied by the stresses (optional, default value 1.0)
    scfactor_strain: Scale factor multiplied by the strains (optional, default value 1.0)
    sendtoexcel: Whether to send the retrieved data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the retrieved data is stored (optional, default value is 'book1.xlsx')

    return: Data as a pandas DataFrame.
           The column labels of the DataFrame are ['step', 'time', 'eletag', 'section', 'y', 'z', 'strain', 'stress'].

    Example:

        file_stress = "myproject\output_ntha\section_stress.feather"
        df_s = opa.section_stress(file_stress, tag=1, section=1, y=0.3, z=0.3)
    '''

    df = pd.read_feather(file_stress)

    tim = df.columns[2]
    zerovals = [0, 0.0, tag, section, y, z, 0.0, 0.0]
    filtered_values = np.where((df['eletag'] == tag) & (df['section'] == section) & (df['y'] == y) & (df['z'] == z))
    if len(steps) == 0:
        df_stress = df.loc[filtered_values].reset_index(drop = True)
        df_stress['stress'] = df_stress['stress'] * scfactor_stress
        df_stress['strain'] = df_stress['strain'] * scfactor_strain
        df_stress = df_stress.drop(columns=['index'])
    else:
        steps = list(dict.fromkeys(steps))
        cols_ = ['step', tim, 'eletag', 'section', 'y', 'z', 'strain', 'stress']

        df_stress = pd.DataFrame(columns=cols_)
        df_stress.loc[0] = zerovals
        ste_count = 1
        dff = df.loc[filtered_values]
        for step in steps:
            dff_ = dff.loc[df['step'] == step]
            time = list(dff_[tim])[0]
            vals = [step, time, tag, section, y, z, list(dff_['strain'])[0] * scfactor_strain,
                    list(dff_['stress'])[0]*scfactor_stress]
            df_stress.loc[ste_count] = vals
            ste_count += 1

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_stress)

    return df_stress


def section_force(file_force, tag, section, steps=[], scfactor_axial=1.0, scfactor_moment=1.0, scfactor_axialstrain=1.0,
               scfactor_curvature=1.0, sendtoexcel='y', file_path='book1.xlsx', sheet_name='sheet1'):
    '''
    Function to retrieve forces and deformations of a beam-column element section from a saved response file.

    file_force: saved response file path(file name is section_force.feather)
    tag: Beam-Column element tag
    section: Section number
    steps: A python list containing the steps in which the responses are retrieved (optional, default value is an empty list)
    scfactor_axial: Scale factor multiplied by the axial forces (optional, default value 1)
    scfactor_moment: Scale factor multiplied by the moments (optional, default value 1)
    scfactor_axialstrain: Scale factor multiplied by the axial strains (optional, default value 1)
    scfactor_curvature: Scale factor multiplied by the curvatures (optional, default value 1)
    sendtoexcel: Whether to send the retrieved data to excel spreed sheets (optional, default value is 'y')
    file_path: Excel file in which the retrieved data is stored (optional, default value is 'book1.xlsx')

    return: Data as a pandas DataFrame.
           The column labels of the DataFrame are ['step', 'time', 'eletag', 'section', 'axial strain', 'axial force', 'curvature', 'moment'].

    Example:

        file_force = "myproject\output_ntha\section_force.feather"
        df_s = opa.section_force(file_force, tag=1, section=1)

    '''

    df = pd.read_feather(file_force)
    tim = df.columns[2]
    zerovals = [0, 0.0, tag, section, 0.0, 0.0, 0.0, 0.0]
    filtered_values = np.where((df['eletag'] == tag) & (df['section'] == section))

    if len(steps) == 0:
        df_force = df.loc[filtered_values].reset_index(drop = True)
        df_force['axial strain'] = df_force['axial strain'] * scfactor_axialstrain
        df_force['axial force'] = df_force['axial force'] * scfactor_axial
        df_force['curvature'] = df_force['curvature'] * scfactor_curvature
        df_force['moment'] = df_force['moment'] * scfactor_moment
        df_force = df_force.drop(columns=['index'])

    else:
        steps = list(dict.fromkeys(steps))

        cols_ = ['step', tim, 'eletag', 'section', 'axial strain', 'axial force', 'curvature', 'moment']

        df_force = pd.DataFrame(columns=cols_)
        df_force.loc[0] = zerovals
        ste_count = 1
        dff = df.loc[filtered_values]
        for step in steps:
            dff_ = dff.loc[df['step'] == step]
            time = list(dff_[tim])[0]
            vals = [step, time, tag, section, list(dff_['axial strain'])[0]*scfactor_axialstrain,
                    list(dff_['axial force'])[0]*scfactor_axial, list(dff_['moment'])[0] * scfactor_moment,
                    list(dff_['curvature'])[0]*scfactor_curvature]
            df_force.loc[ste_count] = vals
            ste_count += 1

    if sendtoexcel.lower() in ['y', 'yes']:
        __sendtoexcel(file_path, sheet_name, df_force)

    return df_force

def plot_section_stress(ax, df_stress, xlabel='', ylabel='stress', title='', linewidth=1.0, color='k', grid=True):
    """
    Function to plot stress history for a beam-column element section.

    ax: The axes on which the plot is drawn.
    df_stress: A DataFrame is returned by section_stress function.
    xlabel: Label for the x-axis (optional, default value is '')
    ylabel: Label for the y-axis (optional, default value is 'stress')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    tim = df_stress.columns[1]
    values_x = list(df_stress[tim])
    values_y = list(df_stress['stress'])
    if values_x[0] != 0:
        values_x.insert(0, 0.0)
        values_y.insert(0, 0.0)
    ax.plot(values_x, values_y, lw=linewidth, color=color)
    ax.grid(grid)
    if xlabel == '':
        xlabel=tim
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


def plot_section_strain(ax, df_stress, xlabel='', ylabel='strain', title='', linewidth=1.0, color='k', grid=True):
    """
    Function to plot strain history for a beam-column element section.

    ax: The axes on which the plot is drawn.
    df_stress: A DataFrame is returned by section_stress function.
    xlabel: Label for the x-axis (optional, default value is '')
    ylabel: Label for the y-axis (optional, default value is 'strain')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    tim = df_stress.columns[1]
    values_x = list(df_stress[tim])
    values_y = list(df_stress['strain'])
    if values_x[0] != 0:
        values_x.insert(0, 0.0)
        values_y.insert(0, 0.0)
    ax.plot(values_x, values_y, lw=linewidth, color=color)
    ax.grid(grid)
    if xlabel == '':
        xlabel=tim
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


def plot_section_stressstrain(ax, df_stress, xlabel='strain', ylabel='stress', title='', linewidth=1.0, color='k', grid=True):
    """
    Function to plot stress vs strain curve for a beam-column element section.

    ax: The axes on which the plot is drawn.
    df_stress: A DataFrame is returned by section_stress function.
    xlabel: Label for the x-axis (optional, default value is 'strain')
    ylabel: Label for the y-axis (optional, default value is 'stress')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    values_x = list(df_stress['strain'])
    values_y = list(df_stress['stress'])
    if values_x[0] != 0:
        values_x.insert(0, 0.0)
        values_y.insert(0, 0.0)
    ax.plot(values_x, values_y, lw=linewidth, color=color)
    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)

def plot_section_axialforce(ax, df_force, xlabel='', ylabel='axial force', title='', linewidth=1.0, color='k', grid=True):
    """
    Function to plot axial force history for a beam-column element section.

    ax: The axes on which the plot is drawn.
    df_force: A DataFrame is returned by section_force function.
    xlabel: Label for the x-axis (optional, default value is '')
    ylabel: Label for the y-axis (optional, default value is 'axial force')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    tim = df_force.columns[1]
    values_x = list(df_force[tim])
    values_y = list(df_force['axial force'])
    if values_x[0] != 0:
        values_x.insert(0, 0.0)
        values_y.insert(0, 0.0)
    ax.plot(values_x, values_y, lw=linewidth, color=color)
    ax.grid(grid)
    if xlabel == '':
        xlabel=tim
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)

def plot_section_axialstrain(ax, df_force, xlabel='', ylabel='axial strain', title='', linewidth=1.0, color='k', grid=True):
    """
    Function to plot axial strain history for a beam-column element section.

    ax: The axes on which the plot is drawn.
    df_force: A DataFrame is returned by section_force function.
    xlabel: Label for the x-axis (optional, default value is '')
    ylabel: Label for the y-axis (optional, default value is 'axial strain')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    tim = df_force.columns[1]
    values_x = list(df_force[tim])
    values_y = list(df_force['axial strain'])
    if values_x[0] != 0:
        values_x.insert(0, 0.0)
        values_y.insert(0, 0.0)
    ax.plot(values_x, values_y, lw=linewidth, color=color)
    ax.grid(grid)
    if xlabel == '':
        xlabel=tim
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


def plot_section_axialforcestrain(ax, df_force, xlabel='axial strain', ylabel='axial force', title='', linewidth=1.0,
                               color='k', grid=True):
    """
    Function to plot axial force vs axial strain curve for a beam-column element section.

    ax: The axes on which the plot is drawn.
    df_force: A DataFrame is returned by section_force function.
    xlabel: Label for the x-axis (optional, default value is 'axial strain')
    ylabel: Label for the y-axis (optional, default value is 'axial force')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """

    values_x = list(df_force['axial strain'])
    values_y = list(df_force['axial force'])
    if values_x[0] != 0:
        values_x.insert(0, 0.0)
        values_y.insert(0, 0.0)
    ax.plot(values_x, values_y, lw=linewidth, color=color)
    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


def plot_section_moment(ax, df_force, xlabel='', ylabel='moment', title='', linewidth=1.0, color='k', grid=True):
    """
    Function to plot moment history for a beam-column element section.

    ax: The axes on which the plot is drawn.
    df_force: A DataFrame is returned by section_force function.
    xlabel: Label for the x-axis (optional, default value is '')
    ylabel: Label for the y-axis (optional, default value is 'moment')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    tim = df_force.columns[1]
    values_x = list(df_force[tim])
    values_y = list(df_force['moment'])
    if values_x[0] != 0:
        values_x.insert(0, 0.0)
        values_y.insert(0, 0.0)
    ax.plot(values_x, values_y, lw=linewidth, color=color)
    ax.grid(grid)
    if xlabel == '':
        xlabel=tim
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


def plot_section_curvature(ax, df_force, xlabel='', ylabel='curvature', title='', linewidth=1.0, color='k', grid=True):
    """
    Function to plot curvature history for a beam-column element section.

    ax: The axes on which the plot is drawn.
    df_force: A DataFrame is returned by section_force function.
    xlabel: Label for the x-axis (optional, default value is '')
    ylabel: Label for the y-axis (optional, default value is 'curvature')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    tim = df_force.columns[1]
    values_x = list(df_force[tim])
    values_y = list(df_force['curvature'])
    if values_x[0] != 0:
        values_x.insert(0, 0.0)
        values_y.insert(0, 0.0)

    ax.plot(values_x, values_y, lw=linewidth, color=color)
    ax.grid(grid)
    if xlabel == '':
        xlabel=tim
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)


def plot_section_momentcurvature(ax, df_force, xlabel='curvature', ylabel='moment', title='', linewidth=1.0, color='k',
                              grid=True):
    """
    Function to plot moment - curvature curve for a beam-column element section.

    ax: The axes on which the plot is drawn.
    df_force: A DataFrame is returned by section_force function.
    xlabel: Label for the x-axis (optional, default value is 'curvature')
    ylabel: Label for the y-axis (optional, default value is 'moment')
    title:  Title for the axes (optional, default value is '')
    grid: Whether to show the grid lines (optional, default value is False)
    linewidth: Line width of the graph plot (optional, default value is 1.0)
    color: Color of the graph plot (optional, default value is 'b')

    """
    values_x = list(df_force['curvature'])
    values_y = list(df_force['moment'])
    if values_x[0] != 0:
        values_x.insert(0, 0.0)
        values_y.insert(0, 0.0)
    ax.plot(values_x, values_y, lw=linewidth, color=color)
    ax.grid(grid)
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    ax.set_title(title)

###################################################################

def __create_workbook(file_path):
    if not os.path.exists(file_path):
        book = workbook.Workbook()
        book.save(file_path)
    else:
        book = load_workbook(file_path)

    return book


def __sendtoexcel(file_path, sheet_name, df_data):
    book = __create_workbook(file_path)

    sheetnames = book.sheetnames
    if sheet_name in sheetnames:
        del book[sheet_name]

    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book

    df_data.to_excel(writer, sheet_name=sheet_name, index_label='row')
    writer.close()

def getDoc():
    webbrowser.open('https://drive.google.com/file/d/1ILaQ44qE-GH-SyZbCGnKbGb2au42aO3r/view?usp=sharing')

def getUserManual():
    webbrowser.open('https://drive.google.com/file/d/1viVBX74uQY4qQCasRQ_w004S1zoYcyIE/view?usp=sharing')
